# -*- coding: utf-8 -*-
"""Export dialogs and helpers for MainWindow."""

import concurrent.futures
import math
import os
import sqlite3
import tempfile
import threading
import time
import unicodedata
import zipfile

from typing import Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from qt_compat import (
    Qt, QThread, Signal, QLabel, QSpinBox, QCheckBox, QHBoxLayout, QVBoxLayout,
    QTableWidget, QTableWidgetItem, QMessageBox, QFileDialog, QHeaderView,
    QTimer, QDialog, QListWidget, QListWidgetItem, QAbstractItemView,
    QDialogButtonBox, QProgressBar, QPushButton,
)

class HistoryDbDialog(QDialog):
    def __init__(self, parent, data_dir: str, export_cb):
        super().__init__(parent)
        self._data_dir = data_dir
        self._export_cb = export_cb
        self._bulk_updating = False
        self.setWindowTitle("管理数据库")
        self.resize(620, 460)

        root = QVBoxLayout(self)
        info = QLabel("管理历史数据库，可导出或删除。")
        root.addWidget(info)

        top_row = QHBoxLayout()
        self.select_all_chk = QCheckBox("全选")
        try:
            self.select_all_chk.setTristate(True)
        except Exception:
            pass
        top_row.addWidget(self.select_all_chk)
        top_row.addStretch(1)
        root.addLayout(top_row)

        self.list = QListWidget()
        self.list.setSelectionMode(QAbstractItemView.ExtendedSelection)
        root.addWidget(self.list, 1)

        btns = QDialogButtonBox()
        self.btn_refresh = QPushButton("刷新")
        self.btn_export = QPushButton("导出")
        self.btn_delete = QPushButton("删除")
        self.btn_close = QPushButton("关闭")
        btns.addButton(self.btn_refresh, QDialogButtonBox.ActionRole)
        btns.addButton(self.btn_export, QDialogButtonBox.AcceptRole)
        btns.addButton(self.btn_delete, QDialogButtonBox.DestructiveRole)
        btns.addButton(self.btn_close, QDialogButtonBox.RejectRole)
        root.addWidget(btns)

        self.btn_refresh.clicked.connect(self.reload)
        self.btn_export.clicked.connect(self.export_selected)
        self.btn_delete.clicked.connect(self.delete_selected)
        self.btn_close.clicked.connect(self.reject)
        self.list.itemDoubleClicked.connect(lambda *_: self.export_selected())
        self.select_all_chk.clicked.connect(self._on_select_all_changed)
        self.list.itemChanged.connect(self._on_item_changed)

        self.reload()

    def _iter_items(self):
        for i in range(self.list.count()):
            item = self.list.item(i)
            if not item:
                continue
            path = item.data(Qt.UserRole)
            if not path:
                continue
            yield item, path

    def _checked_paths(self):
        paths = []
        for item, path in self._iter_items():
            if item.checkState() == Qt.Checked:
                paths.append(path)
        return paths

    def _selected_paths(self):
        paths = []
        for item in self.list.selectedItems() or []:
            path = item.data(Qt.UserRole)
            if path:
                paths.append(path)
        return paths

    def _get_action_paths(self):
        paths = self._checked_paths()
        if paths:
            return paths
        return self._selected_paths()

    def _update_select_all_state(self):
        if self._bulk_updating:
            return
        self._bulk_updating = True
        try:
            total = 0
            checked = 0
            for item, _ in self._iter_items():
                total += 1
                if item.checkState() == Qt.Checked:
                    checked += 1
            if total == 0:
                self.select_all_chk.setCheckState(Qt.Unchecked)
                self.select_all_chk.setEnabled(False)
            else:
                self.select_all_chk.setEnabled(True)
                if checked == 0:
                    self.select_all_chk.setCheckState(Qt.Unchecked)
                elif checked == total:
                    self.select_all_chk.setCheckState(Qt.Checked)
                else:
                    self.select_all_chk.setCheckState(Qt.PartiallyChecked)
        finally:
            self._bulk_updating = False

    def _on_select_all_changed(self, state):
        if self._bulk_updating:
            return
        self._bulk_updating = True
        try:
            if isinstance(state, bool):
                checked = state
            else:
                checked = state == Qt.Checked
            for item, _ in self._iter_items():
                item.setCheckState(Qt.Checked if checked else Qt.Unchecked)
        finally:
            self._bulk_updating = False
        self._update_select_all_state()

    def _on_item_changed(self, *_):
        self._update_select_all_state()

    def reload(self):
        self.list.clear()
        if not os.path.isdir(self._data_dir):
            item = QListWidgetItem("(未找到 data_logs 目录)")
            item.setFlags(Qt.NoItemFlags)
            self.list.addItem(item)
            self._update_select_all_state()
            return

        db_files = []
        try:
            for name in os.listdir(self._data_dir):
                if not name.lower().endswith((".sqlite", ".db")):
                    continue
                full = os.path.join(self._data_dir, name)
                if os.path.isfile(full):
                    try:
                        mtime = os.path.getmtime(full)
                    except Exception:
                        mtime = 0.0
                    db_files.append((mtime, full))
        except Exception:
            db_files = []

        if not db_files:
            item = QListWidgetItem("(无历史数据库)")
            item.setFlags(Qt.NoItemFlags)
            self.list.addItem(item)
            self._update_select_all_state()
            return

        db_files.sort(key=lambda x: x[0], reverse=True)
        for mtime, full in db_files:
            base = os.path.splitext(os.path.basename(full))[0]
            try:
                ts_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(float(mtime)))
            except Exception:
                ts_str = "未知时间"
            label = f"{ts_str}  |  {base}"
            item = QListWidgetItem(label)
            item.setData(Qt.UserRole, full)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable | Qt.ItemIsSelectable | Qt.ItemIsEnabled)
            item.setCheckState(Qt.Unchecked)
            self.list.addItem(item)

        self._update_select_all_state()

    def export_selected(self):
        paths = self._get_action_paths()
        if not paths:
            QMessageBox.information(self, "提示", "请先选择要导出的数据库。")
            return
        parent = self.parent()
        if parent and hasattr(parent, "queue_export_db_paths"):
            parent.queue_export_db_paths(paths)
            return
        QMessageBox.warning(self, "提示", "无法导出，请在主界面操作。")

    def delete_selected(self):
        paths = self._get_action_paths()
        if not paths:
            QMessageBox.information(self, "提示", "请先选择要删除的数据库。")
            return
        count = len(paths)
        tip = f"确定删除选中的 {count} 个数据库文件吗？\n此操作不可恢复。"
        if QMessageBox.question(self, "确认删除", tip) != QMessageBox.Yes:
            return
        failed = []
        for path in paths:
            try:
                os.remove(path)
            except Exception:
                failed.append(path)
        if failed:
            QMessageBox.warning(self, "删除失败", "以下文件删除失败：\n" + "\n".join(failed))
        self.reload()

class ExportQueueWorker(QThread):
    progress = Signal(str, int, int)
    status = Signal(str, str)
    phase = Signal(str, str)
    error = Signal(str, str)
    finished = Signal(list, list, str)

    def __init__(self, tasks, export_func, max_workers=8, zip_path=""):
        super().__init__()
        self._tasks = tasks or []
        self._export_func = export_func
        self._max_workers = max(1, int(max_workers))
        self._zip_path = zip_path or ""
        self._cancel_event = threading.Event()
        self._pause_event = threading.Event()

    def cancel(self):
        try:
            self._cancel_event.set()
        except Exception:
            pass

    def set_paused(self, paused: bool):
        try:
            if paused:
                self._pause_event.set()
            else:
                self._pause_event.clear()
        except Exception:
            pass

    def run(self):
        exported = []
        failed = []
        zip_path = self._zip_path

        def _wait_if_paused():
            while self._pause_event.is_set():
                if self._cancel_event.is_set():
                    raise RuntimeError("cancelled")
                time.sleep(0.1)

        def progress_cb(ctx, done, total):
            _wait_if_paused()
            if self._cancel_event.is_set():
                raise RuntimeError("cancelled")
            try:
                total_i = int(total) if total is not None else -1
            except Exception:
                total_i = -1
            try:
                done_i = int(done)
            except Exception:
                done_i = 0
            self.progress.emit(str(ctx), done_i, total_i)

        def run_one(task):
            if self._cancel_event.is_set():
                return False, (task.get("db_path") or "", "")
            db_path = task.get("db_path") or ""
            out_path = task.get("out_path") or ""
            if not db_path:
                return False, ("", "")
            self.status.emit(db_path, "导出中")
            try:
                _wait_if_paused()
                if self._cancel_event.is_set():
                    return False, (db_path, "")
                self._export_func(
                    db_path,
                    out_path,
                    progress_cb=progress_cb,
                    progress_ctx=db_path,
                    phase_cb=lambda p: self.phase.emit(db_path, str(p)),
                )
                if self._cancel_event.is_set():
                    return False, (db_path, "")
                self.status.emit(db_path, "完成")
                return True, (db_path, out_path)
            except Exception as e:
                try:
                    self.error.emit(db_path, str(e))
                except Exception:
                    pass
                self.status.emit(db_path, "失败")
                return False, (db_path, "")

        tmp_ctx = None
        if zip_path:
            tmp_ctx = tempfile.TemporaryDirectory()
            for task in self._tasks:
                if task.get("out_path"):
                    continue
                db_path = task.get("db_path") or ""
                base = os.path.splitext(os.path.basename(db_path))[0]
                task["out_path"] = os.path.join(tmp_ctx.name, base + ".xlsx")

        try:
            with concurrent.futures.ThreadPoolExecutor(max_workers=self._max_workers) as ex:
                futures = []
                for task in self._tasks:
                    if self._cancel_event.is_set():
                        break
                    futures.append(ex.submit(run_one, task))
                for fut in concurrent.futures.as_completed(futures):
                    if self._cancel_event.is_set():
                        break
                    ok, payload = fut.result()
                    if ok:
                        exported.append(payload)
                    else:
                        failed.append(payload[0])

            if zip_path and exported and (not self._cancel_event.is_set()):
                self.phase.emit("__zip__", "打包中")
                try:
                    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
                        for db_path, out_path in exported:
                            if not out_path:
                                continue
                            arc_name = os.path.basename(out_path)
                            zf.write(out_path, arcname=arc_name)
                    self.phase.emit("__zip__", "打包完成")
                except Exception:
                    self.phase.emit("__zip__", "打包失败")
        finally:
            try:
                if tmp_ctx is not None:
                    tmp_ctx.cleanup()
            except Exception:
                pass

        self.finished.emit([p for p, _ in exported], failed, zip_path)


class ExportQueueDialog(QDialog):
    def __init__(self, parent, export_func):
        super().__init__(parent)
        self._export_func = export_func
        self._tasks = []
        self._task_info = {}
        self._worker = None
        self._zip_path = ""
        self._start_ts = None
        self._size_map = {}
        self._progress_bars = {}

        self.setWindowTitle("导出队列")
        self.resize(720, 480)

        root = QVBoxLayout(self)

        top_row = QHBoxLayout()
        top_row.addWidget(QLabel("最大线程数："))
        self.thread_spin = QSpinBox()
        self.thread_spin.setRange(1, 64)
        self.thread_spin.setValue(8)
        top_row.addWidget(self.thread_spin)
        top_row.addStretch(1)
        root.addLayout(top_row)

        self.status_label = QLabel("状态：空闲")
        self.eta_label = QLabel("预计剩余时间：-")
        root.addWidget(self.status_label)
        root.addWidget(self.eta_label)

        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        root.addWidget(self.progress_bar)

        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["数据库", "状态", "进度条", "进度"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        root.addWidget(self.table, 1)

        btn_row = QHBoxLayout()
        self.btn_pause = QPushButton("暂停")
        self.btn_cancel = QPushButton("取消导出")
        self.btn_close = QPushButton("关闭")
        self.btn_pause.setEnabled(False)
        self.btn_cancel.setEnabled(False)
        btn_row.addStretch(1)
        btn_row.addWidget(self.btn_pause)
        btn_row.addWidget(self.btn_cancel)
        btn_row.addWidget(self.btn_close)
        root.addLayout(btn_row)

        self.btn_close.clicked.connect(self.hide)
        self.btn_pause.clicked.connect(self._toggle_pause)
        self.btn_cancel.clicked.connect(self._cancel_export)

        self._timer = QTimer(self)
        self._timer.setInterval(1000)
        self._timer.timeout.connect(self._update_eta)

    def _estimate_finish_seconds(self, total_bytes: int) -> float:
        if total_bytes <= 0:
            return 2.0
        size_mb = float(total_bytes) / (1024.0 * 1024.0)
        seconds = size_mb * 0.4 * 20.0
        if seconds < 1.0:
            seconds = 1.0
        if seconds > 60.0:
            seconds = 60.0
        return seconds

    def _toggle_pause(self):
        if not self._worker or not self._worker.isRunning():
            return
        paused = self.btn_pause.text() == "暂停"
        try:
            self._worker.set_paused(paused)
        except Exception:
            pass
        if paused:
            self.btn_pause.setText("继续")
            self.status_label.setText("状态：已暂停")
        else:
            self.btn_pause.setText("暂停")
            self.status_label.setText("状态：导出中")

    def _cancel_export(self):
        if not self._worker or not self._worker.isRunning():
            return
        self.force_stop_exports()
        self._tasks = []
        self._task_info = {}
        self._size_map = {}
        self._progress_bars = {}
        self.table.setRowCount(0)
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.status_label.setText("状态：已取消")
        self.eta_label.setText("预计剩余时间：--")
        self.thread_spin.setEnabled(True)
        self.btn_cancel.setEnabled(False)
        self.btn_pause.setEnabled(False)
        self.btn_pause.setText("暂停")

    def enqueue_exports(self, tasks, zip_path=""):
        if self._worker and self._worker.isRunning():
            QMessageBox.information(self, "提示", "正在导出，请等待完成。")
            return False
        if not tasks:
            return False
        self._zip_path = zip_path or ""
        self._tasks = tasks
        self._size_map = {}
        self._progress_bars = {}
        for task in tasks:
            db_path = task.get("db_path") or ""
            try:
                if db_path and os.path.isfile(db_path):
                    self._size_map[db_path] = os.path.getsize(db_path)
            except Exception:
                self._size_map[db_path] = 0
        self._build_table()
        self.start_export()
        return True

    def _build_table(self):
        self._task_info = {}
        self.table.setRowCount(len(self._tasks))
        for row, task in enumerate(self._tasks):
            db_path = task.get("db_path") or ""
            name = os.path.basename(db_path)
            self.table.setItem(row, 0, QTableWidgetItem(name))
            self.table.setItem(row, 1, QTableWidgetItem("等待"))
            self.table.setItem(row, 3, QTableWidgetItem("0%"))
            bar = QProgressBar()
            bar.setRange(0, 100)
            bar.setValue(0)
            bar.setFormat("%p%")
            bar.setTextVisible(True)
            self.table.setCellWidget(row, 2, bar)
            self._progress_bars[db_path] = bar
            size_b = self._size_map.get(db_path, 0)
            self._task_info[db_path] = {
                "row": row,
                "done": 0,
                "total": 0,
                "status": "等待",
                "pending_finish": False,
                "finish_mode": False,
                "finish_start_ts": None,
                "finish_seconds": self._estimate_finish_seconds(size_b),
            }

    def show_completed_task(self, name: str):
        self._tasks = [{"db_path": name, "out_path": ""}]
        self._task_info = {name: {"row": 0, "done": 1, "total": 1, "status": "完成"}}
        self.table.setRowCount(1)
        self.table.setItem(0, 0, QTableWidgetItem(os.path.basename(name)))
        self.table.setItem(0, 1, QTableWidgetItem("完成"))
        self.table.setItem(0, 3, QTableWidgetItem("100%"))
        bar = QProgressBar()
        bar.setRange(0, 100)
        bar.setValue(100)
        bar.setFormat("%p%")
        bar.setTextVisible(True)
        self.table.setCellWidget(0, 2, bar)
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(100)
        self.status_label.setText("状态：已完成")
        self.eta_label.setText("预计剩余时间：00:00")
        self.thread_spin.setEnabled(True)

    def start_export(self):
        self.btn_pause.setEnabled(True)
        self.btn_cancel.setEnabled(True)
        self.btn_pause.setText("暂停")
        if not self._tasks:
            return
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.status_label.setText("状态：导出中")
        self.eta_label.setText("预计剩余时间：计算中")
        self._start_ts = time.time()
        for info in self._task_info.values():
            info["pending_finish"] = False
            info["finish_mode"] = False
            info["finish_start_ts"] = None
        self.thread_spin.setEnabled(False)

        self._worker = ExportQueueWorker(self._tasks, self._export_func, self.thread_spin.value(), self._zip_path)
        self._worker.progress.connect(self._on_task_progress)
        self._worker.status.connect(self._on_task_status)
        self._worker.phase.connect(self._on_phase)
        self._worker.error.connect(self._on_task_error)
        self._worker.finished.connect(self._on_finished)
        self._worker.start()
        self._timer.start()

    def is_exporting(self) -> bool:
        return self._worker is not None and self._worker.isRunning()

    def force_stop_exports(self, wait_ms: int = 3000):
        if not self._worker:
            return
        try:
            self._worker.cancel()
        except Exception:
            pass
        try:
            self._timer.stop()
        except Exception:
            pass
        try:
            if self._worker.isRunning():
                self._worker.wait(wait_ms)
        except Exception:
            pass
        try:
            if self._worker.isRunning():
                self._worker.terminate()
                self._worker.wait(1000)
        except Exception:
            pass

    def _on_task_progress(self, db_path, done, total):
        info = self._task_info.get(db_path)
        if not info:
            return
        info["done"] = max(0, int(done))
        if total is not None and int(total) > 0:
            info["total"] = int(total)
        if info.get("pending_finish") and (not info.get("finish_mode")):
            if info.get("total", 0) > 0 and info.get("done", 0) >= info.get("total", 0):
                info["finish_mode"] = True
                info["finish_start_ts"] = time.time()
        self._update_row_progress(db_path)

    def _on_task_status(self, db_path, status):
        info = self._task_info.get(db_path)
        if not info:
            return
        info["status"] = status
        row = info["row"]
        item = self.table.item(row, 1)
        if item:
            item.setText(status)
        if status == "完成":
            bar = self._progress_bars.get(db_path)
            if bar:
                bar.setValue(100)
        self._update_overall_status()

    def _on_task_error(self, db_path, msg):
        if msg:
            self.status_label.setText(f"状态：失败 - {msg}")
            QMessageBox.warning(self, "导出失败", f"数据库：{os.path.basename(db_path)}\n错误：{msg}")

    def _on_phase(self, db_path, phase):
        if db_path == "__zip__":
            self.status_label.setText(f"状态：{phase}")
            return
        info = self._task_info.get(db_path)
        if not info:
            return
        if phase == "收尾中":
            info["pending_finish"] = True
            if info.get("total", 0) > 0 and info.get("done", 0) >= info.get("total", 0):
                if not info.get("finish_mode"):
                    info["finish_mode"] = True
                    info["finish_start_ts"] = time.time()
            status_item = self.table.item(info["row"], 1)
            if status_item:
                status_item.setText("收尾中")
        self._update_overall_status()

    def _on_finished(self, ok_paths, failed_paths, zip_path):
        self._timer.stop()
        self.thread_spin.setEnabled(True)
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(100)
        self.btn_cancel.setEnabled(False)
        self.btn_pause.setEnabled(False)
        self.btn_pause.setText("暂停")
        for db_path, bar in self._progress_bars.items():
            if bar:
                bar.setValue(100)
        if failed_paths:
            self.status_label.setText("状态：已完成（部分失败）")
        else:
            self.status_label.setText("状态：已完成")
        self._update_overall_status(force=True)

    def _update_row_progress(self, db_path):
        info = self._task_info.get(db_path)
        if not info:
            return
        row = info["row"]
        done = info.get("done", 0)
        total = info.get("total", 0)
        if total > 0:
            pct = min(100.0, 100.0 * float(done) / float(total))
            text = f"{pct:.1f}% ({done}/{total})"
        else:
            text = f"{done}"
        item = self.table.item(row, 3)
        if item:
            item.setText(text)
        bar = self._progress_bars.get(db_path)
        if bar:
            now_ts = time.time()
            pct_hidden = self._calc_task_progress(info, now_ts)
            if info.get("status") == "完成":
                bar.setValue(100)
            else:
                bar.setValue(int(max(0, min(99, pct_hidden))))

    def _calc_task_progress(self, info, now_ts: float) -> float:
        if info.get("status") == "完成":
            return 100.0
        total = info.get("total", 0)
        done = info.get("done", 0)
        if total > 0:
            pct_export = 90.0 * float(done) / float(total)
        else:
            pct_export = 0.0
        pct_export = max(0.0, min(90.0, pct_export))
        if info.get("finish_mode") and info.get("finish_start_ts") is not None:
            fin_elapsed = max(0.0, now_ts - float(info.get("finish_start_ts")))
            fin_total = max(0.1, float(info.get("finish_seconds", 1.0)))
            fin_ratio = max(0.0, min(1.0, fin_elapsed / fin_total))
            fin_pct = 90.0 + 9.0 * fin_ratio
            return max(pct_export, min(99.0, fin_pct))
        return pct_export

    def _update_eta(self):
        try:
            total = 0
            done = 0
            for info in self._task_info.values():
                if info.get("total", 0) > 0:
                    total += info.get("total", 0)
                    done += min(info.get("done", 0), info.get("total", 0))
            if total <= 0:
                self.eta_label.setText("预计剩余时间：计算中")
                return
            if self._start_ts is None:
                return
            elapsed = max(0.001, time.time() - float(self._start_ts))
            rate = float(done) / elapsed if done > 0 else 0.0
            if rate <= 0:
                self.eta_label.setText("预计剩余时间：计算中")
                return
            remain = max(0.0, float(total - done) / rate)
            finish_remain = 0.0
            now_ts = time.time()
            for info in self._task_info.values():
                if info.get("finish_mode") and info.get("finish_start_ts") is not None:
                    fin_total = max(0.1, float(info.get("finish_seconds", 1.0)))
                    fin_elapsed = max(0.0, now_ts - float(info.get("finish_start_ts")))
                    finish_remain += max(0.0, fin_total - fin_elapsed)
            mm = int(remain // 60)
            ss = int(remain % 60)
            total_remain = remain + finish_remain
            mm = int(total_remain // 60)
            ss = int(total_remain % 60)
            self.eta_label.setText(f"预计剩余时间：{mm:02d}:{ss:02d}")

            if self.progress_bar.maximum() <= 0:
                return
            if not self._task_info:
                self.progress_bar.setValue(0)
                return
            now_ts = time.time()
            total_pct = 0.0
            for info in self._task_info.values():
                total_pct += self._calc_task_progress(info, now_ts)
            for db_path, info in self._task_info.items():
                bar = self._progress_bars.get(db_path)
                if bar:
                    pct_hidden = self._calc_task_progress(info, now_ts)
                    if info.get("status") == "完成":
                        bar.setValue(100)
                    else:
                        bar.setValue(int(max(0, min(99, pct_hidden))))
            avg_pct = total_pct / max(1, len(self._task_info))
            self.progress_bar.setValue(int(min(99, max(0, avg_pct))))
            self._update_overall_status()
        except Exception:
            pass

    def _update_overall_status(self, force: bool = False):
        if not self._task_info:
            if force:
                self.status_label.setText("状态：空闲")
            return
        statuses = [info.get("status", "") for info in self._task_info.values()]
        if any(s == "导出中" for s in statuses):
            self.status_label.setText("状态：导出中")
            return
        if any(s == "收尾中" for s in statuses) or any(info.get("finish_mode") for info in self._task_info.values()):
            self.status_label.setText("状态：收尾中")
            return
        if all(s == "完成" for s in statuses):
            self.status_label.setText("状态：完成")
            return
        if any(s == "失败" for s in statuses):
            self.status_label.setText("状态：失败")
            return




class ExportMixin:
    def _db_has_data(self, path: str) -> bool:
        if not path:
            return False
        if not os.path.isfile(path):
            return False
        try:
            conn = sqlite3.connect(path)
            try:
                cur = conn.execute("SELECT 1 FROM data LIMIT 1")
                return cur.fetchone() is not None
            finally:
                conn.close()
        except Exception:
            return False

    def _format_export_time(self, wall_ts, rel_ts) -> str:
        try:
            if wall_ts is not None and math.isfinite(float(wall_ts)):
                return time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(float(wall_ts)))
        except Exception:
            pass
        try:
            return f"{float(rel_ts):.3f}s"
        except Exception:
            return ""

    def _unit_label(self, unit: str) -> str:
        u = (unit or "").strip()
        if not u:
            return ""
        ul = u.lower()
        if ul == "g":
            return "g【克】"
        if ul == "n":
            return "N【牛】"
        if u in ("无量纲", "-"):
            return "无量纲"
        return f"{u}【单位】"

    def _split_sheet_name(self, base_title: str, idx: int) -> str:
        base = self._safe_sheet_name(base_title)
        name = f"{base}_{idx}" if idx > 1 else base
        if len(name) > 31:
            name = name[:31]
        return name

    def _create_sheet_with_header_cells(self, wb, base_title: str, idx: int, headers):
        ws = wb.create_sheet(title=self._split_sheet_name(base_title, idx))
        for col, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=h)
        return ws

    def _create_sheet_with_header_append(self, wb, base_title: str, idx: int, headers):
        ws = wb.create_sheet(self._split_sheet_name(base_title, idx))
        ws.append(headers)
        return ws

    def _export_xlsx_from_ring(self, path: str, xs, ys_map, xs_wall, qf_vals=None):
        wb = Workbook()
        max_rows = 1048576

        headers = ["Time"]
        units = list(getattr(self, "_log_units", []))
        avg_header = "平均张力(N【牛】)"
        avg_inserted = False
        for idx, name in enumerate(self.channel_names):
            unit = units[idx] if idx < len(units) else ""
            unit_label = self._unit_label(unit)
            headers.append(f"{name}({unit_label})" if unit_label else name)
            if name == "CH2":
                headers.append(avg_header)
                avg_inserted = True
        if not avg_inserted:
            headers.append(avg_header)
        headers += ["摩擦力(N【牛】)", "摩擦系数", self._quality_flag_label]

        ws_all_idx = 1
        ws_all = wb.active
        ws_all.title = self._split_sheet_name("All", ws_all_idx)
        for i, h in enumerate(headers, start=1):
            ws_all.cell(row=1, column=i, value=h)
        ws_all_row = 2

        nrows = len(xs)
        hi_name = (getattr(self, "_fric_high_name", "") or "").strip()
        lo_name = (getattr(self, "_fric_low_name", "") or "").strip()

        for i, rel_ts in enumerate(xs):
            if ws_all_row > max_rows:
                ws_all_idx += 1
                ws_all = self._create_sheet_with_header_cells(wb, "All", ws_all_idx, headers)
                ws_all_row = 2
            wall_ts = xs_wall[i] if xs_wall and i < len(xs_wall) else None
            t_str = self._format_export_time(wall_ts, rel_ts)
            ws_all.cell(row=ws_all_row, column=1, value=t_str)

            row_vals = []
            for c, name in enumerate(self.channel_names, start=2):
                col = ys_map.get(name, [])
                v = col[i] if i < len(col) else None
                row_vals.append(v)

            high_v = None
            low_v = None
            if hi_name and hi_name in self.channel_names:
                try:
                    high_v = row_vals[self.channel_names.index(hi_name)]
                except Exception:
                    high_v = None
            if lo_name and lo_name in self.channel_names:
                try:
                    low_v = row_vals[self.channel_names.index(lo_name)]
                except Exception:
                    low_v = None
            avg_v = self._calc_avg_tension(high_v, low_v)

            col_idx = 2
            avg_inserted = False
            for c, name in enumerate(self.channel_names, start=2):
                v = row_vals[c - 2] if (c - 2) < len(row_vals) else None
                ws_all.cell(row=ws_all_row, column=col_idx, value=v)
                col_idx += 1
                if name == "CH2":
                    ws_all.cell(row=ws_all_row, column=col_idx, value=avg_v)
                    col_idx += 1
                    avg_inserted = True
            if not avg_inserted:
                ws_all.cell(row=ws_all_row, column=col_idx, value=avg_v)
                col_idx += 1

            fric_n, mu = self._calc_fric_mu(high_v, low_v)
            ws_all.cell(row=ws_all_row, column=col_idx, value=fric_n)
            ws_all.cell(row=ws_all_row, column=col_idx + 1, value=mu)
            col_idx += 2

            qf_v = qf_vals[i] if (qf_vals is not None and i < len(qf_vals)) else None
            ws_all.cell(row=ws_all_row, column=col_idx, value=qf_v)

            ws_all_row += 1

        for i, name in enumerate(self.channel_names):
            unit = units[i] if i < len(units) else ""
            unit_label = self._unit_label(unit)
            header = ["Time", f"{name}({unit_label})" if unit_label else name]
            ws_idx = 1
            ws = self._create_sheet_with_header_cells(wb, name, ws_idx, header)
            row_idx = 2
            vals = ys_map.get(name, [])
            for r in range(nrows):
                if row_idx > max_rows:
                    ws_idx += 1
                    ws = self._create_sheet_with_header_cells(wb, name, ws_idx, header)
                    row_idx = 2
                wall_ts = xs_wall[r] if xs_wall and r < len(xs_wall) else None
                t_str = self._format_export_time(wall_ts, xs[r])
                ws.cell(row=row_idx, column=1, value=t_str)
                v = vals[r] if r < len(vals) else None
                ws.cell(row=row_idx, column=2, value=v)
                row_idx += 1

        if qf_vals is not None:
            q_header = ["Time", self._quality_flag_label]
            ws_q_idx = 1
            ws_q = self._create_sheet_with_header_cells(wb, self._quality_flag_name, ws_q_idx, q_header)
            row_idx = 2
            for r in range(nrows):
                if row_idx > max_rows:
                    ws_q_idx += 1
                    ws_q = self._create_sheet_with_header_cells(wb, self._quality_flag_name, ws_q_idx, q_header)
                    row_idx = 2
                wall_ts = xs_wall[r] if xs_wall and r < len(xs_wall) else None
                t_str = self._format_export_time(wall_ts, xs[r])
                ws_q.cell(row=row_idx, column=1, value=t_str)
                qv = qf_vals[r] if r < len(qf_vals) else None
                ws_q.cell(row=row_idx, column=2, value=qv)
                row_idx += 1

        avg_header = ["Time", "平均张力(N【牛】)"]
        f_header = ["Time", "摩擦力(N【牛】)", self._quality_flag_label]
        mu_header = ["Time", "摩擦系数", self._quality_flag_label]
        ws_f_idx = 1
        ws_mu_idx = 1
        ws_avg_idx = 1
        ws_f = self._create_sheet_with_header_cells(wb, "摩擦力", ws_f_idx, f_header)
        ws_mu = self._create_sheet_with_header_cells(wb, "摩擦系数", ws_mu_idx, mu_header)
        ws_avg = self._create_sheet_with_header_cells(wb, "平均张力", ws_avg_idx, avg_header)
        ws_f_row = 2
        ws_mu_row = 2
        ws_avg_row = 2
        for i, rel_ts in enumerate(xs):
            if ws_f_row > max_rows:
                ws_f_idx += 1
                ws_f = self._create_sheet_with_header_cells(wb, "摩擦力", ws_f_idx, f_header)
                ws_f_row = 2
            if ws_mu_row > max_rows:
                ws_mu_idx += 1
                ws_mu = self._create_sheet_with_header_cells(wb, "摩擦系数", ws_mu_idx, mu_header)
                ws_mu_row = 2
            if ws_avg_row > max_rows:
                ws_avg_idx += 1
                ws_avg = self._create_sheet_with_header_cells(wb, "平均张力", ws_avg_idx, avg_header)
                ws_avg_row = 2
            wall_ts = xs_wall[i] if xs_wall and i < len(xs_wall) else None
            t_str = self._format_export_time(wall_ts, rel_ts)
            fric_n = None
            mu = None
            if hi_name and lo_name and hi_name in ys_map and lo_name in ys_map:
                try:
                    hv = ys_map.get(hi_name, [])[i]
                    lv = ys_map.get(lo_name, [])[i]
                except Exception:
                    hv = None
                    lv = None
                fric_n, mu = self._calc_fric_mu(hv, lv)
            ws_f.cell(row=ws_f_row, column=1, value=t_str)
            ws_f.cell(row=ws_f_row, column=2, value=fric_n)
            qf_v = qf_vals[i] if (qf_vals is not None and i < len(qf_vals)) else None
            ws_f.cell(row=ws_f_row, column=3, value=qf_v)
            ws_mu.cell(row=ws_mu_row, column=1, value=t_str)
            ws_mu.cell(row=ws_mu_row, column=2, value=mu)
            ws_mu.cell(row=ws_mu_row, column=3, value=qf_v)
            avg_v = self._calc_avg_tension(hv, lv)
            ws_avg.cell(row=ws_avg_row, column=1, value=t_str)
            ws_avg.cell(row=ws_avg_row, column=2, value=avg_v)
            ws_f_row += 1
            ws_mu_row += 1
            ws_avg_row += 1

        for ws in wb.worksheets:
            self._autosize_sheet(ws)

        wb.save(path)

    def _export_xlsx_from_db(self, db_path: str, path: str, progress_cb=None, progress_ctx=None, phase_cb=None):
        try:
            if self._data_logger:
                self._data_logger.flush(wait=True, timeout=3.0)
        except Exception:
            pass

        conn = sqlite3.connect(db_path)
        total_rows = None
        if progress_cb:
            try:
                total_rows = conn.execute("SELECT COUNT(*) FROM data").fetchone()[0]
                if total_rows is None:
                    total_rows = 0
            except Exception:
                total_rows = None
            try:
                if total_rows is not None:
                    total_rows = max(1, int(total_rows))
                    progress_cb(progress_ctx, 0, total_rows)
            except Exception:
                pass
        try:
            try:
                cur = conn.execute("SELECT idx, name, unit FROM channels ORDER BY idx")
                channel_rows = cur.fetchall()
                channel_names = [row[1] for row in channel_rows]
                channel_units = [row[2] if len(row) > 2 else "" for row in channel_rows]
            except Exception:
                cur = conn.execute("SELECT idx, name FROM channels ORDER BY idx")
                channel_rows = cur.fetchall()
                channel_names = [row[1] for row in channel_rows]
                channel_units = ["" for _ in channel_rows]
            n_ch = len(channel_names)
            col_names = [f"ch{i}" for i in range(n_ch)]
            cols_sql = ", ".join(["ts"] + col_names) if col_names else "ts"
            query = f"SELECT {cols_sql} FROM data ORDER BY id"

            max_rows = 1048576
            wb = Workbook(write_only=True)

            headers = ["Time"]
            avg_header = "平均张力(N【牛】)"
            avg_inserted = False
            for idx, name in enumerate(channel_names):
                if name == self._quality_flag_name:
                    continue
                unit = channel_units[idx] if idx < len(channel_units) else ""
                unit_label = self._unit_label(unit)
                headers.append(f"{name}({unit_label})" if unit_label else name)
                if name == "CH2":
                    headers.append(avg_header)
                    avg_inserted = True
            if not avg_inserted:
                headers.append(avg_header)
            headers += ["摩擦力(N【牛】)", "摩擦系数", self._quality_flag_label]

            ws_all_idx = 1
            ws_all = self._create_sheet_with_header_append(wb, "All", ws_all_idx, headers)
            ws_all_rows = 1
            avg_header = ["Time", "平均张力(N【牛】)"]
            f_header = ["Time", "摩擦力(N【牛】)", self._quality_flag_label]
            mu_header = ["Time", "摩擦系数", self._quality_flag_label]
            ws_f_idx = 1
            ws_mu_idx = 1
            ws_avg_idx = 1
            ws_f = self._create_sheet_with_header_append(wb, "摩擦力", ws_f_idx, f_header)
            ws_mu = self._create_sheet_with_header_append(wb, "摩擦系数", ws_mu_idx, mu_header)
            ws_avg = self._create_sheet_with_header_append(wb, "平均张力", ws_avg_idx, avg_header)
            ws_f_rows = 1
            ws_mu_rows = 1
            ws_avg_rows = 1

            hi_name = (getattr(self, "_fric_high_name", "") or "").strip()
            lo_name = (getattr(self, "_fric_low_name", "") or "").strip()
            name_to_idx = {name: i for i, name in enumerate(channel_names)}
            hi_idx = name_to_idx.get(hi_name, None)
            lo_idx = name_to_idx.get(lo_name, None)

            cur = conn.execute(query)
            done_rows = 0
            while True:
                rows = cur.fetchmany(1000)
                if not rows:
                    break
                for row in rows:
                    ts_val = row[0]
                    vals = list(row[1:]) if n_ch > 0 else []
                    t_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(float(ts_val)))
                    high_v = vals[hi_idx] if (hi_idx is not None and hi_idx < len(vals)) else None
                    low_v = vals[lo_idx] if (lo_idx is not None and lo_idx < len(vals)) else None
                    row_out = [t_str]
                    qf_val = None
                    avg_v = self._calc_avg_tension(high_v, low_v)
                    avg_inserted = False
                    for idx, name in enumerate(channel_names):
                        v = vals[idx] if idx < len(vals) else None
                        if name == self._quality_flag_name:
                            qf_val = v
                            continue
                        row_out.append(v)
                        if name == "CH2":
                            row_out.append(avg_v)
                            avg_inserted = True
                    if not avg_inserted:
                        row_out.append(avg_v)

                    fric_n, mu = self._calc_fric_mu(high_v, low_v)
                    row_out += [fric_n, mu, qf_val]

                    if ws_all_rows >= max_rows:
                        ws_all_idx += 1
                        ws_all = self._create_sheet_with_header_append(wb, "All", ws_all_idx, headers)
                        ws_all_rows = 1
                    ws_all.append(row_out)
                    ws_all_rows += 1

                    if ws_f_rows >= max_rows:
                        ws_f_idx += 1
                        ws_f = self._create_sheet_with_header_append(wb, "摩擦力", ws_f_idx, f_header)
                        ws_f_rows = 1
                    ws_f.append([t_str, fric_n, qf_val])
                    ws_f_rows += 1

                    if ws_mu_rows >= max_rows:
                        ws_mu_idx += 1
                        ws_mu = self._create_sheet_with_header_append(wb, "摩擦系数", ws_mu_idx, mu_header)
                        ws_mu_rows = 1
                    if ws_avg_rows >= max_rows:
                        ws_avg_idx += 1
                        ws_avg = self._create_sheet_with_header_append(wb, "平均张力", ws_avg_idx, avg_header)
                        ws_avg_rows = 1
                    ws_mu.append([t_str, mu, qf_val])
                    ws_mu_rows += 1
                    avg_v = self._calc_avg_tension(high_v, low_v)
                    ws_avg.append([t_str, avg_v])
                    ws_avg_rows += 1

                done_rows += len(rows)
                if progress_cb and total_rows is not None:
                    try:
                        progress_cb(progress_ctx, done_rows, total_rows)
                    except Exception:
                        pass
            wb.save(path)
        finally:
            conn.close()
        try:
            if phase_cb:
                phase_cb("收尾中")
        except Exception:
            pass
        try:
            self._autosize_workbook(path)
        except Exception:
            pass

    def export_history_db(self):
        data_dir = os.path.join(os.getcwd(), "data_logs")
        if not os.path.isdir(data_dir):
            QMessageBox.information(self, "提示", "未找到 data_logs 目录。请先采集数据生成历史数据库。")
            return

        db_path, _ = QFileDialog.getOpenFileName(
            self,
            "选择历史数据库",
            data_dir,
            "SQLite DB (*.sqlite *.db);;All Files (*.*)"
        )
        if not db_path:
            return
        if not os.path.isfile(db_path):
            QMessageBox.warning(self, "提示", "数据库文件不存在。")
            return

        self.queue_export_db_paths([db_path])

    def open_history_dialog(self):
        data_dir = os.path.join(os.getcwd(), "data_logs")
        if not hasattr(self, "_history_db_dialog") or self._history_db_dialog is None:
            self._history_db_dialog = HistoryDbDialog(self, data_dir, self._export_history_db_path)
        else:
            try:
                self._history_db_dialog._data_dir = data_dir
                self._history_db_dialog.reload()
            except Exception:
                pass
        dlg = self._history_db_dialog
        dlg.show()
        try:
            dlg.raise_()
            dlg.activateWindow()
        except Exception:
            pass

    def _get_export_queue_dialog(self):
        if not hasattr(self, "_export_queue_dialog") or self._export_queue_dialog is None:
            self._export_queue_dialog = ExportQueueDialog(self, self._export_xlsx_from_db)
        return self._export_queue_dialog

    def open_export_queue_dialog(self):
        dlg = self._get_export_queue_dialog()
        dlg.show()
        try:
            dlg.raise_()
            dlg.activateWindow()
        except Exception:
            pass

    def queue_export_db_paths(self, db_paths: List[str]):
        if not db_paths:
            return

        if len(db_paths) == 1:
            db_path = db_paths[0]
            base = os.path.splitext(os.path.basename(db_path))[0]
            default_dir = os.path.dirname(db_path) or os.path.join(os.getcwd(), "data_logs")
            out_path, _ = QFileDialog.getSaveFileName(
                self,
                "导出为 XLSX",
                os.path.join(default_dir, f"{base}.xlsx"),
                "Excel Files (*.xlsx)"
            )
            if not out_path:
                return
            if not out_path.lower().endswith(".xlsx"):
                out_path += ".xlsx"
            tasks = [{"db_path": db_path, "out_path": out_path}]
            dlg = self._get_export_queue_dialog()
            dlg.show()
            try:
                dlg.raise_()
                dlg.activateWindow()
            except Exception:
                pass
            dlg.enqueue_exports(tasks)
            return

        default_dir = os.path.dirname(db_paths[0]) or os.path.join(os.getcwd(), "data_logs")
        zip_path, _ = QFileDialog.getSaveFileName(
            self,
            "导出压缩包",
            os.path.join(default_dir, "history_databases.zip"),
            "Zip Files (*.zip)"
        )
        if not zip_path:
            return
        if not zip_path.lower().endswith(".zip"):
            zip_path += ".zip"
        tasks = [{"db_path": p, "out_path": ""} for p in db_paths]
        dlg = self._get_export_queue_dialog()
        dlg.show()
        try:
            dlg.raise_()
            dlg.activateWindow()
        except Exception:
            pass
        dlg.enqueue_exports(tasks, zip_path=zip_path)

    def _build_history_menu(self):
        if not hasattr(self, "hist_menu") or self.hist_menu is None:
            return
        self.hist_menu.clear()

        act_refresh = self.hist_menu.addAction("刷新历史数据库")
        act_refresh.triggered.connect(self._build_history_menu)

        act_pick = self.hist_menu.addAction("管理数据库")
        act_pick.triggered.connect(self.open_history_dialog)
        act_queue = self.hist_menu.addAction("导出队列")
        act_queue.triggered.connect(self.open_export_queue_dialog)


        self.hist_menu.addSeparator()

        data_dir = os.path.join(os.getcwd(), "data_logs")
        if not os.path.isdir(data_dir):
            act_empty = self.hist_menu.addAction("(未找到 data_logs 目录)")
            act_empty.setEnabled(False)
            return

        db_files = []
        try:
            for name in os.listdir(data_dir):
                if not name.lower().endswith((".sqlite", ".db")):
                    continue
                full = os.path.join(data_dir, name)
                if os.path.isfile(full):
                    try:
                        mtime = os.path.getmtime(full)
                    except Exception:
                        mtime = 0.0
                    db_files.append((mtime, full))
        except Exception:
            db_files = []

        if not db_files:
            act_none = self.hist_menu.addAction("(无历史数据库)")
            act_none.setEnabled(False)
            return

        db_files.sort(key=lambda x: x[0], reverse=True)
        max_items = 10
        for mtime, full in db_files[:max_items]:
            base = os.path.splitext(os.path.basename(full))[0]
            try:
                ts_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(float(mtime)))
            except Exception:
                ts_str = "未知时间"
            label = f"{ts_str}  |  {base}"
            act = self.hist_menu.addAction(label)
            act.triggered.connect(lambda _=False, p=full: self._export_history_db_path(p))

    def _export_history_db_path(self, db_path: str):
        if not db_path or not os.path.isfile(db_path):
            QMessageBox.warning(self, "提示", "数据库文件不存在。")
            return
        self.queue_export_db_paths([db_path])

    def save_xlsx(self):
        db_path = self._log_db_path if getattr(self, "_log_db_path", "") else ""
        use_db = self._db_has_data(db_path)

        xs = []
        ys_map = {}
        xs_wall = []
        if not use_db:
            xs, ys_map, xs_wall, qf_vals = self._snapshot_ring(include_wall=True, include_quality=True)
            if not xs or not self.channel_names:
                QMessageBox.information(self, "提示", "当前没有可保存的数据。请先开始采集。")
                return

        path, _ = QFileDialog.getSaveFileName(self, "保存为 XLSX", "modbus_data.xlsx", "Excel Files (*.xlsx)")
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        try:
            if use_db:
                dlg = self._get_export_queue_dialog()
                dlg.show()
                try:
                    dlg.raise_()
                    dlg.activateWindow()
                except Exception:
                    pass
                tasks = [{"db_path": db_path, "out_path": path}]
                dlg.enqueue_exports(tasks)
            else:
                self._export_xlsx_from_ring(path, xs, ys_map, xs_wall, qf_vals=qf_vals)
                dlg = self._get_export_queue_dialog()
                dlg.show()
                try:
                    dlg.raise_()
                    dlg.activateWindow()
                except Exception:
                    pass
                dlg.show_completed_task(path)
            self.set_status(f"已保存：{path}")
        except Exception as e:
            QMessageBox.critical(self, "保存失败", f"保存 xlsx 失败：\n{e}")

    @staticmethod
    def _safe_sheet_name(name: str) -> str:
        bad = ['\\', '/', '*', '[', ']', ':', '?']
        s = "".join("_" if ch in bad else ch for ch in name).strip() or "Sheet"
        return s[:31]

    @staticmethod
    def _autosize_sheet(ws):
        max_rows = min(ws.max_row, 2)
        for col in range(1, ws.max_column + 1):
            max_len = 0
            for row in range(1, max_rows + 1):
                v = ws.cell(row=row, column=col).value
                if v is None:
                    continue
                s = str(v)
                # 将 CJK 宽字符按宽度 2 计算，便于 Excel 列宽调整
                width = 0
                for ch in s:
                    width += 2 if unicodedata.east_asian_width(ch) in ("W", "F") else 1
                max_len = max(max_len, width)
            ws.column_dimensions[get_column_letter(col)].width = min(max(10, max_len + 2), 50)

    def _autosize_workbook(self, path: str):
        wb = load_workbook(path)
        try:
            for ws in wb.worksheets:
                self._autosize_sheet(ws)
            wb.save(path)
        finally:
            try:
                wb.close()
            except Exception:
                pass

