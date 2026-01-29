# -*- coding: utf-8 -*-
"""Modbus 助手主界面窗口。"""

import math
import concurrent.futures
import os
import sqlite3
import time
import threading
import tempfile
import zipfile
import unicodedata
from typing import Dict, List, Optional, Tuple

import pyqtgraph as pg
from pyqtgraph.graphicsItems.DateAxisItem import DateAxisItem

import serial
from serial.tools import list_ports

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# 可选：numpy 在高采样率下可加速绘图
try:
    import numpy as np
except Exception:  # pragma: no cover
    np = None

from qt_compat import (
    Qt, QThread, Signal, QMainWindow, QWidget, QLabel, QComboBox, QPushButton, QLineEdit,
    QSpinBox, QDoubleSpinBox, QCheckBox, QHBoxLayout, QVBoxLayout, QGridLayout,
    QGroupBox, QTableWidget, QTableWidgetItem, QMessageBox, QFileDialog,
    QHeaderView, QDockWidget, QTabWidget, QTextEdit, QPlainTextEdit, QSplitter,
    QSizePolicy, QTimer, QSettings, QPoint, QTextCursor, QGuiApplication, QApplication,
    Slot, QDialog, QListWidget, QListWidgetItem, QAbstractItemView, QDialogButtonBox, QProgressBar,
)

from modbus_utils import ChannelConfig, DTYPE_INFO, hex_bytes
from rs485 import Rs485CtrlConfig, Rs485CtrlMode
from virtual_serial import SIM_REGISTRY
from worker import ModbusRtuWorker
from sim_window import SerialSimManagerWindow
from data_logger import DataLogger

class HistoryDbDialog(QDialog):
    def __init__(self, parent, data_dir: str, export_cb):
        super().__init__(parent)
        self._data_dir = data_dir
        self._export_cb = export_cb
        self._bulk_updating = False
        self.setWindowTitle("管理数据库")
        self.resize(620, 460)

        root = QVBoxLayout(self)
        info = QLabel("管理历史数据库，可导出或删除。")
        root.addWidget(info)

        top_row = QHBoxLayout()
        self.select_all_chk = QCheckBox("全选")
        try:
            self.select_all_chk.setTristate(True)
        except Exception:
            pass
        top_row.addWidget(self.select_all_chk)
        top_row.addStretch(1)
        root.addLayout(top_row)

        self.list = QListWidget()
        self.list.setSelectionMode(QAbstractItemView.ExtendedSelection)
        root.addWidget(self.list, 1)

        btns = QDialogButtonBox()
        self.btn_refresh = QPushButton("刷新")
        self.btn_export = QPushButton("导出")
        self.btn_delete = QPushButton("删除")
        self.btn_close = QPushButton("关闭")
        btns.addButton(self.btn_refresh, QDialogButtonBox.ActionRole)
        btns.addButton(self.btn_export, QDialogButtonBox.AcceptRole)
        btns.addButton(self.btn_delete, QDialogButtonBox.DestructiveRole)
        btns.addButton(self.btn_close, QDialogButtonBox.RejectRole)
        root.addWidget(btns)

        self.btn_refresh.clicked.connect(self.reload)
        self.btn_export.clicked.connect(self.export_selected)
        self.btn_delete.clicked.connect(self.delete_selected)
        self.btn_close.clicked.connect(self.reject)
        self.list.itemDoubleClicked.connect(lambda *_: self.export_selected())
        self.select_all_chk.clicked.connect(self._on_select_all_changed)
        self.list.itemChanged.connect(self._on_item_changed)

        self.reload()

    def _iter_items(self):
        for i in range(self.list.count()):
            item = self.list.item(i)
            if not item:
                continue
            path = item.data(Qt.UserRole)
            if not path:
                continue
            yield item, path

    def _checked_paths(self):
        paths = []
        for item, path in self._iter_items():
            if item.checkState() == Qt.Checked:
                paths.append(path)
        return paths

    def _selected_paths(self):
        paths = []
        for item in self.list.selectedItems() or []:
            path = item.data(Qt.UserRole)
            if path:
                paths.append(path)
        return paths

    def _get_action_paths(self):
        paths = self._checked_paths()
        if paths:
            return paths
        return self._selected_paths()

    def _update_select_all_state(self):
        if self._bulk_updating:
            return
        self._bulk_updating = True
        try:
            total = 0
            checked = 0
            for item, _ in self._iter_items():
                total += 1
                if item.checkState() == Qt.Checked:
                    checked += 1
            if total == 0:
                self.select_all_chk.setCheckState(Qt.Unchecked)
                self.select_all_chk.setEnabled(False)
            else:
                self.select_all_chk.setEnabled(True)
                if checked == 0:
                    self.select_all_chk.setCheckState(Qt.Unchecked)
                elif checked == total:
                    self.select_all_chk.setCheckState(Qt.Checked)
                else:
                    self.select_all_chk.setCheckState(Qt.PartiallyChecked)
        finally:
            self._bulk_updating = False

    def _on_select_all_changed(self, state):
        if self._bulk_updating:
            return
        self._bulk_updating = True
        try:
            if isinstance(state, bool):
                checked = state
            else:
                checked = state == Qt.Checked
            for item, _ in self._iter_items():
                item.setCheckState(Qt.Checked if checked else Qt.Unchecked)
        finally:
            self._bulk_updating = False
        self._update_select_all_state()

    def _on_item_changed(self, *_):
        self._update_select_all_state()

    def reload(self):
        self.list.clear()
        if not os.path.isdir(self._data_dir):
            item = QListWidgetItem("(未找到 data_logs 目录)")
            item.setFlags(Qt.NoItemFlags)
            self.list.addItem(item)
            self._update_select_all_state()
            return

        db_files = []
        try:
            for name in os.listdir(self._data_dir):
                if not name.lower().endswith((".sqlite", ".db")):
                    continue
                full = os.path.join(self._data_dir, name)
                if os.path.isfile(full):
                    try:
                        mtime = os.path.getmtime(full)
                    except Exception:
                        mtime = 0.0
                    db_files.append((mtime, full))
        except Exception:
            db_files = []

        if not db_files:
            item = QListWidgetItem("(无历史数据库)")
            item.setFlags(Qt.NoItemFlags)
            self.list.addItem(item)
            self._update_select_all_state()
            return

        db_files.sort(key=lambda x: x[0], reverse=True)
        for mtime, full in db_files:
            base = os.path.splitext(os.path.basename(full))[0]
            try:
                ts_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(float(mtime)))
            except Exception:
                ts_str = "未知时间"
            label = f"{ts_str}  |  {base}"
            item = QListWidgetItem(label)
            item.setData(Qt.UserRole, full)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable | Qt.ItemIsSelectable | Qt.ItemIsEnabled)
            item.setCheckState(Qt.Unchecked)
            self.list.addItem(item)

        self._update_select_all_state()

    def export_selected(self):
        paths = self._get_action_paths()
        if not paths:
            QMessageBox.information(self, "提示", "请先选择要导出的数据库。")
            return
        parent = self.parent()
        if parent and hasattr(parent, "queue_export_db_paths"):
            parent.queue_export_db_paths(paths)
            return
        QMessageBox.warning(self, "提示", "无法导出，请在主界面操作。")

    def delete_selected(self):
        paths = self._get_action_paths()
        if not paths:
            QMessageBox.information(self, "提示", "请先选择要删除的数据库。")
            return
        count = len(paths)
        tip = f"确定删除选中的 {count} 个数据库文件吗？\n此操作不可恢复。"
        if QMessageBox.question(self, "确认删除", tip) != QMessageBox.Yes:
            return
        failed = []
        for path in paths:
            try:
                os.remove(path)
            except Exception:
                failed.append(path)
        if failed:
            QMessageBox.warning(self, "删除失败", "以下文件删除失败：\n" + "\n".join(failed))
        self.reload()

class ExportQueueWorker(QThread):
    progress = Signal(str, int, int)
    status = Signal(str, str)
    phase = Signal(str, str)
    error = Signal(str, str)
    finished = Signal(list, list, str)

    def __init__(self, tasks, export_func, max_workers=8, zip_path=""):
        super().__init__()
        self._tasks = tasks or []
        self._export_func = export_func
        self._max_workers = max(1, int(max_workers))
        self._zip_path = zip_path or ""
        self._cancel_event = threading.Event()
        self._pause_event = threading.Event()

    def cancel(self):
        try:
            self._cancel_event.set()
        except Exception:
            pass

    def set_paused(self, paused: bool):
        try:
            if paused:
                self._pause_event.set()
            else:
                self._pause_event.clear()
        except Exception:
            pass

    def run(self):
        exported = []
        failed = []
        zip_path = self._zip_path

        def _wait_if_paused():
            while self._pause_event.is_set():
                if self._cancel_event.is_set():
                    raise RuntimeError("cancelled")
                time.sleep(0.1)

        def progress_cb(ctx, done, total):
            _wait_if_paused()
            if self._cancel_event.is_set():
                raise RuntimeError("cancelled")
            try:
                total_i = int(total) if total is not None else -1
            except Exception:
                total_i = -1
            try:
                done_i = int(done)
            except Exception:
                done_i = 0
            self.progress.emit(str(ctx), done_i, total_i)

        def run_one(task):
            if self._cancel_event.is_set():
                return False, (task.get("db_path") or "", "")
            db_path = task.get("db_path") or ""
            out_path = task.get("out_path") or ""
            if not db_path:
                return False, ("", "")
            self.status.emit(db_path, "导出中")
            try:
                _wait_if_paused()
                if self._cancel_event.is_set():
                    return False, (db_path, "")
                self._export_func(
                    db_path,
                    out_path,
                    progress_cb=progress_cb,
                    progress_ctx=db_path,
                    phase_cb=lambda p: self.phase.emit(db_path, str(p)),
                )
                if self._cancel_event.is_set():
                    return False, (db_path, "")
                self.status.emit(db_path, "完成")
                return True, (db_path, out_path)
            except Exception as e:
                try:
                    self.error.emit(db_path, str(e))
                except Exception:
                    pass
                self.status.emit(db_path, "失败")
                return False, (db_path, "")

        tmp_ctx = None
        if zip_path:
            tmp_ctx = tempfile.TemporaryDirectory()
            for task in self._tasks:
                if task.get("out_path"):
                    continue
                db_path = task.get("db_path") or ""
                base = os.path.splitext(os.path.basename(db_path))[0]
                task["out_path"] = os.path.join(tmp_ctx.name, base + ".xlsx")

        try:
            with concurrent.futures.ThreadPoolExecutor(max_workers=self._max_workers) as ex:
                futures = []
                for task in self._tasks:
                    if self._cancel_event.is_set():
                        break
                    futures.append(ex.submit(run_one, task))
                for fut in concurrent.futures.as_completed(futures):
                    if self._cancel_event.is_set():
                        break
                    ok, payload = fut.result()
                    if ok:
                        exported.append(payload)
                    else:
                        failed.append(payload[0])

            if zip_path and exported and (not self._cancel_event.is_set()):
                self.phase.emit("__zip__", "打包中")
                try:
                    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
                        for db_path, out_path in exported:
                            if not out_path:
                                continue
                            arc_name = os.path.basename(out_path)
                            zf.write(out_path, arcname=arc_name)
                    self.phase.emit("__zip__", "打包完成")
                except Exception:
                    self.phase.emit("__zip__", "打包失败")
        finally:
            try:
                if tmp_ctx is not None:
                    tmp_ctx.cleanup()
            except Exception:
                pass

        self.finished.emit([p for p, _ in exported], failed, zip_path)


class ExportQueueDialog(QDialog):
    def __init__(self, parent, export_func):
        super().__init__(parent)
        self._export_func = export_func
        self._tasks = []
        self._task_info = {}
        self._worker = None
        self._zip_path = ""
        self._start_ts = None
        self._size_map = {}
        self._progress_bars = {}

        self.setWindowTitle("导出队列")
        self.resize(720, 480)

        root = QVBoxLayout(self)

        top_row = QHBoxLayout()
        top_row.addWidget(QLabel("最大线程数："))
        self.thread_spin = QSpinBox()
        self.thread_spin.setRange(1, 64)
        self.thread_spin.setValue(8)
        top_row.addWidget(self.thread_spin)
        top_row.addStretch(1)
        root.addLayout(top_row)

        self.status_label = QLabel("状态：空闲")
        self.eta_label = QLabel("预计剩余时间：-")
        root.addWidget(self.status_label)
        root.addWidget(self.eta_label)

        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        root.addWidget(self.progress_bar)

        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["数据库", "状态", "进度条", "进度"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        root.addWidget(self.table, 1)

        btn_row = QHBoxLayout()
        self.btn_pause = QPushButton("暂停")
        self.btn_cancel = QPushButton("取消导出")
        self.btn_close = QPushButton("关闭")
        self.btn_pause.setEnabled(False)
        self.btn_cancel.setEnabled(False)
        btn_row.addStretch(1)
        btn_row.addWidget(self.btn_pause)
        btn_row.addWidget(self.btn_cancel)
        btn_row.addWidget(self.btn_close)
        root.addLayout(btn_row)

        self.btn_close.clicked.connect(self.hide)
        self.btn_pause.clicked.connect(self._toggle_pause)
        self.btn_cancel.clicked.connect(self._cancel_export)

        self._timer = QTimer(self)
        self._timer.setInterval(1000)
        self._timer.timeout.connect(self._update_eta)

    def _estimate_finish_seconds(self, total_bytes: int) -> float:
        if total_bytes <= 0:
            return 2.0
        size_mb = float(total_bytes) / (1024.0 * 1024.0)
        seconds = size_mb * 0.4 * 5.0
        if seconds < 1.0:
            seconds = 1.0
        if seconds > 60.0:
            seconds = 60.0
        return seconds

    def _toggle_pause(self):
        if not self._worker or not self._worker.isRunning():
            return
        paused = self.btn_pause.text() == "暂停"
        try:
            self._worker.set_paused(paused)
        except Exception:
            pass
        if paused:
            self.btn_pause.setText("继续")
            self.status_label.setText("状态：已暂停")
        else:
            self.btn_pause.setText("暂停")
            self.status_label.setText("状态：导出中")

    def _cancel_export(self):
        if not self._worker or not self._worker.isRunning():
            return
        self.force_stop_exports()
        self._tasks = []
        self._task_info = {}
        self._size_map = {}
        self._progress_bars = {}
        self.table.setRowCount(0)
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.status_label.setText("状态：已取消")
        self.eta_label.setText("预计剩余时间：--")
        self.thread_spin.setEnabled(True)
        self.btn_cancel.setEnabled(False)
        self.btn_pause.setEnabled(False)
        self.btn_pause.setText("暂停")

    def enqueue_exports(self, tasks, zip_path=""):
        if self._worker and self._worker.isRunning():
            QMessageBox.information(self, "提示", "正在导出，请等待完成。")
            return False
        if not tasks:
            return False
        self._zip_path = zip_path or ""
        self._tasks = tasks
        self._size_map = {}
        self._progress_bars = {}
        for task in tasks:
            db_path = task.get("db_path") or ""
            try:
                if db_path and os.path.isfile(db_path):
                    self._size_map[db_path] = os.path.getsize(db_path)
            except Exception:
                self._size_map[db_path] = 0
        self._build_table()
        self.start_export()
        return True

    def _build_table(self):
        self._task_info = {}
        self.table.setRowCount(len(self._tasks))
        for row, task in enumerate(self._tasks):
            db_path = task.get("db_path") or ""
            name = os.path.basename(db_path)
            self.table.setItem(row, 0, QTableWidgetItem(name))
            self.table.setItem(row, 1, QTableWidgetItem("等待"))
            self.table.setItem(row, 3, QTableWidgetItem("0%"))
            bar = QProgressBar()
            bar.setRange(0, 100)
            bar.setValue(0)
            bar.setFormat("%p%")
            bar.setTextVisible(True)
            self.table.setCellWidget(row, 2, bar)
            self._progress_bars[db_path] = bar
            size_b = self._size_map.get(db_path, 0)
            self._task_info[db_path] = {
                "row": row,
                "done": 0,
                "total": 0,
                "status": "等待",
                "pending_finish": False,
                "finish_mode": False,
                "finish_start_ts": None,
                "finish_seconds": self._estimate_finish_seconds(size_b),
            }

    def show_completed_task(self, name: str):
        self._tasks = [{"db_path": name, "out_path": ""}]
        self._task_info = {name: {"row": 0, "done": 1, "total": 1, "status": "完成"}}
        self.table.setRowCount(1)
        self.table.setItem(0, 0, QTableWidgetItem(os.path.basename(name)))
        self.table.setItem(0, 1, QTableWidgetItem("完成"))
        self.table.setItem(0, 3, QTableWidgetItem("100%"))
        bar = QProgressBar()
        bar.setRange(0, 100)
        bar.setValue(100)
        bar.setFormat("%p%")
        bar.setTextVisible(True)
        self.table.setCellWidget(0, 2, bar)
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(100)
        self.status_label.setText("状态：已完成")
        self.eta_label.setText("预计剩余时间：00:00")
        self.thread_spin.setEnabled(True)

    def start_export(self):
        self.btn_pause.setEnabled(True)
        self.btn_cancel.setEnabled(True)
        self.btn_pause.setText("暂停")
        if not self._tasks:
            return
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.status_label.setText("状态：导出中")
        self.eta_label.setText("预计剩余时间：计算中")
        self._start_ts = time.time()
        for info in self._task_info.values():
            info["pending_finish"] = False
            info["finish_mode"] = False
            info["finish_start_ts"] = None
        self.thread_spin.setEnabled(False)

        self._worker = ExportQueueWorker(self._tasks, self._export_func, self.thread_spin.value(), self._zip_path)
        self._worker.progress.connect(self._on_task_progress)
        self._worker.status.connect(self._on_task_status)
        self._worker.phase.connect(self._on_phase)
        self._worker.error.connect(self._on_task_error)
        self._worker.finished.connect(self._on_finished)
        self._worker.start()
        self._timer.start()

    def is_exporting(self) -> bool:
        return self._worker is not None and self._worker.isRunning()

    def force_stop_exports(self, wait_ms: int = 3000):
        if not self._worker:
            return
        try:
            self._worker.cancel()
        except Exception:
            pass
        try:
            self._timer.stop()
        except Exception:
            pass
        try:
            if self._worker.isRunning():
                self._worker.wait(wait_ms)
        except Exception:
            pass
        try:
            if self._worker.isRunning():
                self._worker.terminate()
                self._worker.wait(1000)
        except Exception:
            pass

    def _on_task_progress(self, db_path, done, total):
        info = self._task_info.get(db_path)
        if not info:
            return
        info["done"] = max(0, int(done))
        if total is not None and int(total) > 0:
            info["total"] = int(total)
        if info.get("pending_finish") and (not info.get("finish_mode")):
            if info.get("total", 0) > 0 and info.get("done", 0) >= info.get("total", 0):
                info["finish_mode"] = True
                info["finish_start_ts"] = time.time()
        self._update_row_progress(db_path)

    def _on_task_status(self, db_path, status):
        info = self._task_info.get(db_path)
        if not info:
            return
        info["status"] = status
        row = info["row"]
        item = self.table.item(row, 1)
        if item:
            item.setText(status)
        if status == "完成":
            bar = self._progress_bars.get(db_path)
            if bar:
                bar.setValue(100)

    def _on_task_error(self, db_path, msg):
        if msg:
            self.status_label.setText(f"状态：失败 - {msg}")
            QMessageBox.warning(self, "导出失败", f"数据库：{os.path.basename(db_path)}\n错误：{msg}")

    def _on_phase(self, db_path, phase):
        if db_path == "__zip__":
            self.status_label.setText(f"状态：{phase}")
            return
        info = self._task_info.get(db_path)
        if not info:
            return
        if phase == "收尾中":
            info["pending_finish"] = True
            if info.get("total", 0) > 0 and info.get("done", 0) >= info.get("total", 0):
                if not info.get("finish_mode"):
                    info["finish_mode"] = True
                    info["finish_start_ts"] = time.time()
            status_item = self.table.item(info["row"], 1)
            if status_item:
                status_item.setText("收尾中")
        self.status_label.setText(f"状态：{phase}")

    def _on_finished(self, ok_paths, failed_paths, zip_path):
        self._timer.stop()
        self.thread_spin.setEnabled(True)
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(100)
        self.btn_cancel.setEnabled(False)
        self.btn_pause.setEnabled(False)
        self.btn_pause.setText("暂停")
        for db_path, bar in self._progress_bars.items():
            if bar:
                bar.setValue(100)
        if failed_paths:
            self.status_label.setText("状态：已完成（部分失败）")
        else:
            self.status_label.setText("状态：已完成")

    def _update_row_progress(self, db_path):
        info = self._task_info.get(db_path)
        if not info:
            return
        row = info["row"]
        done = info.get("done", 0)
        total = info.get("total", 0)
        if total > 0:
            pct = min(100.0, 100.0 * float(done) / float(total))
            text = f"{pct:.1f}% ({done}/{total})"
        else:
            text = f"{done}"
        item = self.table.item(row, 3)
        if item:
            item.setText(text)
        bar = self._progress_bars.get(db_path)
        if bar:
            now_ts = time.time()
            pct_hidden = self._calc_task_progress(info, now_ts)
            if info.get("status") == "完成":
                bar.setValue(100)
            else:
                bar.setValue(int(max(0, min(99, pct_hidden))))

    def _calc_task_progress(self, info, now_ts: float) -> float:
        if info.get("status") == "完成":
            return 100.0
        total = info.get("total", 0)
        done = info.get("done", 0)
        if total > 0:
            pct_export = 90.0 * float(done) / float(total)
        else:
            pct_export = 0.0
        pct_export = max(0.0, min(90.0, pct_export))
        if info.get("finish_mode") and info.get("finish_start_ts") is not None:
            fin_elapsed = max(0.0, now_ts - float(info.get("finish_start_ts")))
            fin_total = max(0.1, float(info.get("finish_seconds", 1.0)))
            fin_ratio = max(0.0, min(1.0, fin_elapsed / fin_total))
            fin_pct = 90.0 + 9.0 * fin_ratio
            return max(pct_export, min(99.0, fin_pct))
        return pct_export

    def _update_eta(self):
        try:
            total = 0
            done = 0
            for info in self._task_info.values():
                if info.get("total", 0) > 0:
                    total += info.get("total", 0)
                    done += min(info.get("done", 0), info.get("total", 0))
            if total <= 0:
                self.eta_label.setText("预计剩余时间：计算中")
                return
            if self._start_ts is None:
                return
            elapsed = max(0.001, time.time() - float(self._start_ts))
            rate = float(done) / elapsed if done > 0 else 0.0
            if rate <= 0:
                self.eta_label.setText("预计剩余时间：计算中")
                return
            remain = max(0.0, float(total - done) / rate)
            finish_remain = 0.0
            now_ts = time.time()
            for info in self._task_info.values():
                if info.get("finish_mode") and info.get("finish_start_ts") is not None:
                    fin_total = max(0.1, float(info.get("finish_seconds", 1.0)))
                    fin_elapsed = max(0.0, now_ts - float(info.get("finish_start_ts")))
                    finish_remain += max(0.0, fin_total - fin_elapsed)
            mm = int(remain // 60)
            ss = int(remain % 60)
            total_remain = remain + finish_remain
            mm = int(total_remain // 60)
            ss = int(total_remain % 60)
            self.eta_label.setText(f"预计剩余时间：{mm:02d}:{ss:02d}")

            if self.progress_bar.maximum() <= 0:
                return
            if not self._task_info:
                self.progress_bar.setValue(0)
                return
            now_ts = time.time()
            total_pct = 0.0
            for info in self._task_info.values():
                total_pct += self._calc_task_progress(info, now_ts)
            for db_path, info in self._task_info.items():
                bar = self._progress_bars.get(db_path)
                if bar:
                    pct_hidden = self._calc_task_progress(info, now_ts)
                    if info.get("status") == "完成":
                        bar.setValue(100)
                    else:
                        bar.setValue(int(max(0, min(99, pct_hidden))))
            avg_pct = total_pct / max(1, len(self._task_info))
            self.progress_bar.setValue(int(min(99, max(0, avg_pct))))
        except Exception:
            pass

class MainWindow(QMainWindow):

    def __init__(self):
        super().__init__()
        try:
            self.setObjectName('MainWindow')
        except Exception:
            pass
        self.setWindowTitle("Modbus RTU 上位机助手（浅色主题 / 通讯监视）")
        self.resize(1260, 820)

        self.worker: Optional[ModbusRtuWorker] = None
        self.is_connected = False
        self.is_acquiring = False
        self.is_paused = False

        self.channel_names: List[str] = []
        self.curves: Dict[str, pg.PlotDataItem] = {}

        # ---- 绘图环形缓冲区（大小 = 当前窗口最大点数） ----
        self._buf_size = 100
        self._buf_count = 0
        self._buf_idx = 0  # 下一次写入索引
        self._ts_buf = None  # numpy 数组或列表
        self._ts_wall_buf = None  # 墙钟时间秒（epoch）
        self._val_buf_by_channel: Dict[str, object] = {}  # 名称 -> np.ndarray 或列表
        self._plot_x = None  # 用于绘图的连续 x（numpy）
        self._plot_y_by_channel: Dict[str, object] = {}  # 名称 -> np.ndarray 或列表
        self._fric_buf = None
        self._mu_buf = None
        self._fric_plot_y = None
        self._mu_plot_y = None
        self._qf_buf = None
        self._plot_seq = 0

        self._last_plotted_seq = -1
        # 平滑滚动的时间基准（相对秒）。
        self._t0_mono_ts = None           # 首次采样时的 time.monotonic()
        self._last_sample_rel_ts = None   # 上次采样的相对秒
        self._last_sample_mono_ts = None  # 上次采样时的 time.monotonic()

        # 带暂停补偿的单调时间轴（恢复时不跳变）。
        self._mono_pause_accum = 0.0
        self._mono_pause_start = None
        self._settings = QSettings("ModbusAssistant", "ModbusAssistant")
        self._fric_high_name = ""
        self._fric_low_name = ""
        self._wrap_angle_deg = 0.0
        self._wrap_angle_rad = 0.0
        self._quality_flag_name = "质量标志"
        self._quality_flag_label = "质量标志（quality flag）【0为无效，1为有效】"
        self._quality_gap_pending: List[dict] = []
        self._quality_gap_hold_mode = False
        self._quality_gap_start_mono = None
        self._quality_gap_triggered = False
        self._last_valid_row: Optional[Dict[str, float]] = None
        self._last_tension_setpoint: Optional[float] = None
        self._quality_last_source = "mu"
        self._quality_syncing = False
        self._quality_ui_syncing = False
        self._quality_gap_timeout_s = 1.0
        self._data_logger = DataLogger(base_dir=os.path.join(os.getcwd(), "data_logs"))
        self._log_db_path = ""
        self._log_channels: List[str] = []
        self._log_units: List[str] = []
        self._last_unit_map: Dict[str, str] = {}

        root = QWidget()
        self.setCentralWidget(root)
        layout = QHBoxLayout(root)
        layout.setContentsMargins(0, 0, 0, 0)

        # 使用分割器，使左/右大小在控件启用状态或文本变化时保持稳定。
        self.main_splitter = QSplitter(Qt.Horizontal)
        layout.addWidget(self.main_splitter, 1)

        left_widget = QWidget()
        left = QVBoxLayout(left_widget)
        left.setContentsMargins(8, 8, 8, 8)


        self.main_splitter.addWidget(left_widget)
        self.main_splitter.setStretchFactor(0, 1)
        try:
            self.main_splitter.setCollapsible(0, False)
        except Exception:
            pass

        # ---- 串口分组 ----
        serial_box = QGroupBox("串口配置")
        left.addWidget(serial_box)
        sg = QGridLayout(serial_box)

        self.port_combo = QComboBox()
        self.refresh_ports_btn = QPushButton("刷新串口")
        self.refresh_ports_btn.clicked.connect(lambda *_: self.refresh_ports())

        self.baud_combo = QComboBox()
        for b in [9600, 19200, 38400, 57600, 115200, 230400, 460800, 921600]:
            self.baud_combo.addItem(str(b), b)
        self.baud_combo.setCurrentText("38400")  # 默认 9600

        self.parity_combo = QComboBox()
        self.parity_combo.addItems(["N", "E", "O"])

        self.stopbits_combo = QComboBox()
        self.stopbits_combo.addItems(["1", "2"])

        self.bytesize_combo = QComboBox()
        self.bytesize_combo.addItems(["7", "8"])
        self.bytesize_combo.setCurrentText("8")

        self.timeout_spin = QDoubleSpinBox()
        self.timeout_spin.setDecimals(2)
        self.timeout_spin.setRange(0.05, 10.0)
        self.timeout_spin.setSingleStep(0.05)
        self.timeout_spin.setValue(0.5)

        # RS485 方向控制
        self.rs485_mode_combo = QComboBox()
        self.rs485_mode_combo.addItems([Rs485CtrlMode.RTS, Rs485CtrlMode.DTR, Rs485CtrlMode.AUTO])
        self.rs485_mode_combo.setCurrentText(Rs485CtrlMode.RTS)  # 默认尝试用RTS解决TX不亮

        self.pre_tx_spin = QSpinBox()
        self.pre_tx_spin.setRange(0, 200)
        self.pre_tx_spin.setValue(0)
        self.pre_tx_spin.setSuffix(" ms")

        self.post_tx_spin = QSpinBox()
        self.post_tx_spin.setRange(0, 200)
        self.post_tx_spin.setValue(2)
        self.post_tx_spin.setSuffix(" ms")

        # 通讯监视器启用（rx/tx）
        self.mon_rx_chk = QCheckBox("监听")
        self.mon_rx_chk.setChecked(False)
        self.mon_tx_chk = QCheckBox("监听")
        self.mon_tx_chk.setChecked(False)

        sg.addWidget(QLabel("接收串口(Modbus)"), 0, 0)
        sg.addWidget(self.port_combo, 0, 1)
        sg.addWidget(self.refresh_ports_btn, 0, 2)

        sg.addWidget(self.mon_rx_chk, 0, 3)

        self.tx_port_combo = QComboBox()
        self.tx_baud_combo = QComboBox()
        for b in [9600, 19200, 38400, 57600, 115200, 230400, 460800, 921600]:
            self.tx_baud_combo.addItem(str(b), b)
        self.tx_baud_combo.setCurrentText("115200")

        self.tx_interval_spin = QSpinBox()
        self.tx_interval_spin.setRange(5, 100000)
        self.tx_interval_spin.setValue(20)  # 默认发送间隔：20ms（115200下更稳妥，且与轮询周期匹配）
        self.tx_interval_spin.setSuffix(" ms")
        self.enable_tx_chk = QCheckBox("启用数据发送串口")
        self.enable_tx_chk.setChecked(True)
        sg.addWidget(QLabel("发送串口(输出)"), 1, 0)
        sg.addWidget(self.tx_port_combo, 1, 1)
        sg.addWidget(self.enable_tx_chk, 1, 2)

        sg.addWidget(self.mon_tx_chk, 1, 3)
        sg.addWidget(QLabel("发送波特率"), 2, 0)
        sg.addWidget(self.tx_baud_combo, 2, 1)
        sg.addWidget(QLabel("发送间隔"), 2, 2)
        sg.addWidget(self.tx_interval_spin, 2, 3)

        sg.addWidget(QLabel("接收波特率"), 3, 0)
        sg.addWidget(self.baud_combo, 3, 1)

        sg.addWidget(QLabel("接收校验"), 4, 0)
        sg.addWidget(self.parity_combo, 4, 1)

        sg.addWidget(QLabel("接收停止位"), 5, 0)
        sg.addWidget(self.stopbits_combo, 5, 1)

        sg.addWidget(QLabel("接收数据位"), 6, 0)
        sg.addWidget(self.bytesize_combo, 6, 1)

        sg.addWidget(QLabel("接收超时(s)"), 7, 0)
        sg.addWidget(self.timeout_spin, 7, 1)

        sg.addWidget(QLabel("RS485方向控制"), 8, 0)
        sg.addWidget(self.rs485_mode_combo, 8, 1)
        sg.addWidget(QLabel("TX前延时"), 9, 0)
        sg.addWidget(self.pre_tx_spin, 9, 1)
        sg.addWidget(QLabel("TX后延时"), 10, 0)
        sg.addWidget(self.post_tx_spin, 10, 1)

        # 连接/采集按钮
        btns = QHBoxLayout()
        self.connect_btn = QPushButton("连接串口")
        self.connect_btn.clicked.connect(lambda *_: self.toggle_connect())

        self.acquire_btn = QPushButton("开始采集")
        self.acquire_btn.clicked.connect(lambda *_: self.toggle_acquire())
        self.acquire_btn.setEnabled(False)

        self.pause_btn = QPushButton("暂停")
        self.pause_btn.clicked.connect(lambda *_: self.toggle_pause())
        self.pause_btn.setEnabled(False)

        btns.addWidget(self.connect_btn)
        btns.addWidget(self.acquire_btn)
        btns.addWidget(self.pause_btn)
        left.addLayout(btns)

        # ---- Modbus 分组 ----
        modbus_box = QGroupBox("Modbus 配置")
        left.addWidget(modbus_box)
        mg = QGridLayout(modbus_box)

        self.unit_spin = QSpinBox()
        self.unit_spin.setRange(1, 247)
        self.unit_spin.setValue(1)

        self.func_combo = QComboBox()
        self.func_combo.addItem("03 读保持寄存器", 3)
        self.func_combo.addItem("04 读输入寄存器", 4)

        self.poll_spin = QSpinBox()
        self.poll_spin.setRange(20, 100000)
        self.poll_spin.setValue(20)
        self.poll_spin.setSuffix(" ms")

        self.addr_base1_chk = QCheckBox("地址从 1 开始（自动 -1）")
        self.addr_base1_chk.setChecked(False)  # 你的示例帧 start=0x0000

        mg.addWidget(QLabel("从站地址(站号)"), 0, 0)
        mg.addWidget(self.unit_spin, 0, 1)
        mg.addWidget(QLabel("功能码"), 1, 0)
        mg.addWidget(self.func_combo, 1, 1)
        mg.addWidget(QLabel("轮询周期"), 2, 0)
        mg.addWidget(self.poll_spin, 2, 1)
        mg.addWidget(self.addr_base1_chk, 3, 0, 1, 2)

        # ---- 绘图分组 ----
        plot_box = QGroupBox("绘图设置")
        left.addWidget(plot_box)
        pgd = QGridLayout(plot_box)

        self.max_points_spin = QSpinBox()
        self.max_points_spin.setRange(10, 200000)
        self.max_points_spin.setValue(100)

        self.autoscale_chk = QCheckBox("Y轴自动缩放")
        self.autoscale_chk.setChecked(True)

        # 绘图更新做限频；设置变化时标记为脏
        self.max_points_spin.valueChanged.connect(self._on_max_points_changed)
        self.autoscale_chk.toggled.connect(self._mark_plot_dirty)

        self.plot_fps_spin = QSpinBox()
        self.plot_fps_spin.setRange(1, 240)
        self.plot_fps_spin.setValue(60)
        self.plot_fps_spin.setSuffix(" Hz")
        self.plot_fps_spin.valueChanged.connect(self._on_plot_fps_changed)

        self.clear_btn = QPushButton("清空数据")
        self.clear_btn.clicked.connect(lambda *_: self.clear_data())

        self.save_btn = QPushButton("保存为 XLSX")
        self.save_btn.clicked.connect(lambda *_: self.save_xlsx())

        pgd.addWidget(QLabel("当前窗口最大点数"), 0, 0)
        pgd.addWidget(self.max_points_spin, 0, 1)
        pgd.addWidget(QLabel("绘图刷新率"), 1, 0)
        pgd.addWidget(self.plot_fps_spin, 1, 1)
        pgd.addWidget(self.autoscale_chk, 2, 0, 1, 2)
        pgd.addWidget(self.clear_btn, 3, 0)
        pgd.addWidget(self.save_btn, 3, 1)

        # ---- 通道分组 ----
        ch_box = QGroupBox("通道配置（多通道）")
        left.addWidget(ch_box, 1)
        cl = QVBoxLayout(ch_box)

        self.ch_table = QTableWidget(0, 8)
        self.ch_table.setHorizontalHeaderLabels([
            "启用", "名称", "地址", "数据类型", "字节序(Word内)", "字顺序(Word间)", "缩放系数", "单位"
        ])
        self.ch_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.ch_table.horizontalHeader().setStretchLastSection(True)
        self.ch_table.setAlternatingRowColors(True)
        self.ch_table.itemChanged.connect(lambda *_: self._refresh_friction_channel_options())
        cl.addWidget(self.ch_table)

        btn_row = QHBoxLayout()
        self.add_ch_btn = QPushButton("添加通道")
        self.del_ch_btn = QPushButton("删除选中")
        self.add_ch_btn.clicked.connect(lambda *_: self.add_channel_row())
        self.del_ch_btn.clicked.connect(lambda *_: self.delete_selected_rows())
        btn_row.addWidget(self.add_ch_btn)
        btn_row.addWidget(self.del_ch_btn)
        btn_row.addStretch(1)
        cl.addLayout(btn_row)

        # ---- 绘图区域 ----
        # 使用数值时间轴（秒）以实现平滑滚动。
        # DateAxisItem 会将刻度对齐到“整洁”边界（常为 1s），这会
        # 看起来像是 x 轴每秒只移动一次。
        self.plot = pg.PlotWidget()
        # ---- 绘图性能选项 ----
        # 1) 将绘制裁剪到视窗，避免渲染屏外线段
        # 2) 数据密集时启用自动降采样（峰值模式）
        # 3) 保留渲染端点数上限作为安全兜底
        try:
            pi = self.plot.getPlotItem()
            try:
                pi.setClipToView(True)
            except Exception:
                pass
            try:
                # 不同 pyqtgraph 版本使用的参数名是 “mode” 或 “method”
                pi.setDownsampling(auto=True, mode="peak")
            except Exception:
                try:
                    pi.setDownsampling(auto=True, method="peak")
                except Exception:
                    pass
        except Exception:
            pass
        # 每条曲线发送到渲染器的点数上限（0=禁用）
        self._max_display_points = 6000

        self.plot.setLabel("bottom", "时间", units="s")
        self.plot.setLabel("left", "张力")
        self.plot.addLegend()
        self.plot.showGrid(x=True, y=True, alpha=0.25)

        # ---- 摩擦力曲线 ----
        self.friction_plot = pg.PlotWidget()
        try:
            pi = self.friction_plot.getPlotItem()
            try:
                pi.setClipToView(True)
            except Exception:
                pass
            try:
                pi.setDownsampling(auto=True, mode="peak")
            except Exception:
                try:
                    pi.setDownsampling(auto=True, method="peak")
                except Exception:
                    pass
        except Exception:
            pass
        self.friction_plot.setLabel("bottom", "时间", units="s")
        self.friction_plot.setLabel("left", "摩擦力")
        self.friction_plot.addLegend()
        self.friction_plot.showGrid(x=True, y=True, alpha=0.25)
        self.friction_curve = self.friction_plot.plot([], [], name="摩擦力", pen=pg.mkPen(color=(255, 140, 0), width=2))

        # ---- 摩擦系数曲线 ----
        self.mu_plot = pg.PlotWidget()
        try:
            pi = self.mu_plot.getPlotItem()
            try:
                pi.setClipToView(True)
            except Exception:
                pass
            try:
                pi.setDownsampling(auto=True, mode="peak")
            except Exception:
                try:
                    pi.setDownsampling(auto=True, method="peak")
                except Exception:
                    pass
        except Exception:
            pass
        self.mu_plot.setLabel("bottom", "时间", units="s")
        self.mu_plot.setLabel("left", "摩擦系数")
        self.mu_plot.addLegend()
        self.mu_plot.showGrid(x=True, y=True, alpha=0.25)
        self.mu_curve = self.mu_plot.plot([], [], name="摩擦系数", pen=pg.mkPen(color=(0, 120, 220), width=2))

        # ---- 绘图窗口（停靠面板） ----
        self.plot_tabs = QTabWidget()
        tension_tab = QWidget()
        t_layout = QVBoxLayout(tension_tab)
        t_layout.setContentsMargins(0, 0, 0, 0)
        t_layout.addWidget(self.plot, 1)
        self.status_label = QLabel("状态：未连接")
        t_layout.addWidget(self.status_label)
        self.plot_tabs.addTab(tension_tab, "张力")

        self.friction_tab = QWidget()
        f_layout = QVBoxLayout(self.friction_tab)
        f_layout.setContentsMargins(0, 0, 0, 0)
        cfg = QGridLayout()
        self.fric_high_combo = QComboBox()
        self.fric_low_combo = QComboBox()
        self.fric_swap_btn = QPushButton("互换")
        self.wrap_angle_spin = QDoubleSpinBox()
        self.wrap_angle_spin.setDecimals(2)
        self.wrap_angle_spin.setRange(0.0, 360.0)
        self.wrap_angle_spin.setSingleStep(1.0)
        self.wrap_angle_spin.setValue(10.0)
        self.wrap_angle_spin.setSuffix(" °")
        self.rmin_spin = QDoubleSpinBox()
        self.rmin_spin.setDecimals(4)
        self.rmin_spin.setRange(0.0, 1000.0)
        self.rmin_spin.setSingleStep(0.01)
        self.rmin_spin.setValue(1.01)
        self.mu_max_spin = QDoubleSpinBox()
        self.mu_max_spin.setDecimals(4)
        self.mu_max_spin.setRange(0.0, 10.0)
        self.mu_max_spin.setSingleStep(0.01)
        self.mu_max_spin.setValue(0.5)
        self.rmax_spin = QDoubleSpinBox()
        self.rmax_spin.setDecimals(6)
        self.rmax_spin.setRange(0.0, 1e6)
        self.rmax_spin.setSingleStep(0.01)
        self.rmax_spin.setValue(1.0)
        self.rmax_formula_label = QLabel("Rmax=exp(μmax·θ)")
        self.qgap_spin = QDoubleSpinBox()
        self.qgap_spin.setDecimals(2)
        self.qgap_spin.setRange(0.05, 10.0)
        self.qgap_spin.setSingleStep(0.05)
        self.qgap_spin.setValue(1.0)
        self.qgap_spin.setSuffix(" s")
        cfg.addWidget(QLabel("高张力侧"), 0, 0)
        cfg.addWidget(self.fric_high_combo, 0, 1)
        cfg.addWidget(QLabel("低张力侧"), 0, 2)
        cfg.addWidget(self.fric_low_combo, 0, 3)
        cfg.addWidget(self.fric_swap_btn, 0, 4)
        cfg.addWidget(QLabel("包角"), 1, 0)
        cfg.addWidget(self.wrap_angle_spin, 1, 1)
        cfg.addWidget(QLabel("Rmin"), 2, 0)
        cfg.addWidget(self.rmin_spin, 2, 1)
        cfg.addWidget(QLabel("μmax"), 2, 2)
        cfg.addWidget(self.mu_max_spin, 2, 3)
        cfg.addWidget(QLabel("Rmax"), 3, 0)
        cfg.addWidget(self.rmax_spin, 3, 1)
        cfg.addWidget(self.rmax_formula_label, 3, 2, 1, 3)
        cfg.addWidget(QLabel("丢包/解析失败超时(s)"), 4, 0)
        cfg.addWidget(self.qgap_spin, 4, 1)
        cfg.setColumnStretch(5, 1)
        f_layout.addLayout(cfg)
        f_layout.addWidget(self.friction_plot, 1)

        self.mu_tab = QWidget()
        mu_layout = QVBoxLayout(self.mu_tab)
        mu_layout.setContentsMargins(0, 0, 0, 0)
        mu_cfg = QGridLayout()
        self.mu_high_combo = QComboBox()
        self.mu_low_combo = QComboBox()
        self.mu_swap_btn = QPushButton("互换")
        self.mu_wrap_angle_spin = QDoubleSpinBox()
        self.mu_wrap_angle_spin.setDecimals(2)
        self.mu_wrap_angle_spin.setRange(0.0, 360.0)
        self.mu_wrap_angle_spin.setSingleStep(1.0)
        self.mu_wrap_angle_spin.setValue(10.0)
        self.mu_wrap_angle_spin.setSuffix(" °")
        self.rmin_spin_mu = QDoubleSpinBox()
        self.rmin_spin_mu.setDecimals(4)
        self.rmin_spin_mu.setRange(0.0, 1000.0)
        self.rmin_spin_mu.setSingleStep(0.01)
        self.rmin_spin_mu.setValue(1.01)
        self.mu_max_spin_mu = QDoubleSpinBox()
        self.mu_max_spin_mu.setDecimals(4)
        self.mu_max_spin_mu.setRange(0.0, 10.0)
        self.mu_max_spin_mu.setSingleStep(0.01)
        self.mu_max_spin_mu.setValue(0.5)
        self.rmax_spin_mu = QDoubleSpinBox()
        self.rmax_spin_mu.setDecimals(6)
        self.rmax_spin_mu.setRange(0.0, 1e6)
        self.rmax_spin_mu.setSingleStep(0.01)
        self.rmax_spin_mu.setValue(1.0)
        self.rmax_formula_label_mu = QLabel("Rmax=exp(μmax·θ)")
        self.qgap_spin_mu = QDoubleSpinBox()
        self.qgap_spin_mu.setDecimals(2)
        self.qgap_spin_mu.setRange(0.05, 10.0)
        self.qgap_spin_mu.setSingleStep(0.05)
        self.qgap_spin_mu.setValue(1.0)
        self.qgap_spin_mu.setSuffix(" s")
        mu_cfg.addWidget(QLabel("高张力侧"), 0, 0)
        mu_cfg.addWidget(self.mu_high_combo, 0, 1)
        mu_cfg.addWidget(QLabel("低张力侧"), 0, 2)
        mu_cfg.addWidget(self.mu_low_combo, 0, 3)
        mu_cfg.addWidget(self.mu_swap_btn, 0, 4)
        mu_cfg.addWidget(QLabel("包角"), 1, 0)
        mu_cfg.addWidget(self.mu_wrap_angle_spin, 1, 1)
        mu_cfg.addWidget(QLabel("Rmin"), 2, 0)
        mu_cfg.addWidget(self.rmin_spin_mu, 2, 1)
        mu_cfg.addWidget(QLabel("μmax"), 2, 2)
        mu_cfg.addWidget(self.mu_max_spin_mu, 2, 3)
        mu_cfg.addWidget(QLabel("Rmax"), 3, 0)
        mu_cfg.addWidget(self.rmax_spin_mu, 3, 1)
        mu_cfg.addWidget(self.rmax_formula_label_mu, 3, 2, 1, 3)
        mu_cfg.addWidget(QLabel("丢包/解析失败超时(s)"), 4, 0)
        mu_cfg.addWidget(self.qgap_spin_mu, 4, 1)
        mu_cfg.setColumnStretch(5, 1)
        mu_layout.addLayout(mu_cfg)
        mu_layout.addWidget(self.mu_plot, 1)

        self.plot_dock = QDockWidget("绘图窗口", self)
        try:
            self.plot_dock.setObjectName("dock_plot")
        except Exception:
            pass
        self.plot_dock.setAllowedAreas(Qt.BottomDockWidgetArea | Qt.TopDockWidgetArea | Qt.LeftDockWidgetArea | Qt.RightDockWidgetArea)
        plot_container = QWidget()
        plot_layout = QVBoxLayout(plot_container)
        plot_layout.setContentsMargins(6, 6, 6, 6)
        plot_layout.addWidget(self.plot_tabs, 1)
        self.plot_dock.setWidget(plot_container)
        self.addDockWidget(Qt.RightDockWidgetArea, self.plot_dock)

        self.fric_high_combo.currentIndexChanged.connect(self._on_friction_config_changed)
        self.fric_low_combo.currentIndexChanged.connect(self._on_friction_config_changed)
        self.wrap_angle_spin.valueChanged.connect(self._on_friction_config_changed)
        self.rmin_spin.valueChanged.connect(self._on_quality_rmin_changed)
        self.mu_max_spin.valueChanged.connect(self._on_quality_mu_max_changed)
        self.rmax_spin.valueChanged.connect(self._on_quality_rmax_changed)
        self.qgap_spin.valueChanged.connect(self._on_quality_gap_timeout_changed)
        self.fric_swap_btn.clicked.connect(self._swap_friction_channels)
        self.mu_high_combo.currentIndexChanged.connect(self._on_mu_config_changed)
        self.mu_low_combo.currentIndexChanged.connect(self._on_mu_config_changed)
        self.mu_wrap_angle_spin.valueChanged.connect(self._on_mu_config_changed)
        self.mu_swap_btn.clicked.connect(self._swap_mu_channels)
        self.rmin_spin_mu.valueChanged.connect(self._on_quality_rmin_changed_mu)
        self.mu_max_spin_mu.valueChanged.connect(self._on_quality_mu_max_changed_mu)
        self.rmax_spin_mu.valueChanged.connect(self._on_quality_rmax_changed_mu)
        self.qgap_spin_mu.valueChanged.connect(self._on_quality_gap_timeout_changed_mu)

        # ---- 通讯监视器停靠面板 ----
        self.monitor_dock = QDockWidget("通讯监视窗口", self)
        try:
            self.monitor_dock.setObjectName('dock_monitor')
        except Exception:
            pass
        self.monitor_dock.setAllowedAreas(Qt.BottomDockWidgetArea | Qt.TopDockWidgetArea)
        self.monitor_text = QPlainTextEdit()
        self.monitor_text.setReadOnly(True)
        self.monitor_text.setUndoRedoEnabled(False)
        try:
            self.monitor_text.setMaximumBlockCount(2000)
        except Exception:
            pass

        mon_container = QWidget()
        mon_layout = QVBoxLayout(mon_container)
        mon_layout.setContentsMargins(6, 6, 6, 6)
        # 显示模式
        mon_mode_row = QHBoxLayout()
        mon_mode_row.addWidget(QLabel("显示模式"))
        self.monitor_mode_combo = QComboBox()
        self.monitor_mode_combo.addItem("HEX", "hex")
        self.monitor_mode_combo.addItem("文本(UTF-8)", "utf-8")
        self.monitor_mode_combo.addItem("文本(GBK)", "gbk")
        self.monitor_mode_combo.setCurrentIndex(1)
        self.monitor_mode_combo.currentIndexChanged.connect(lambda *_: self.schedule_monitor_render(full=True))
        mon_mode_row.addWidget(self.monitor_mode_combo)
        mon_mode_row.addStretch(1)
        mon_layout.addLayout(mon_mode_row)

        mon_layout.addWidget(self.monitor_text, 1)

        mon_btns = QHBoxLayout()
        self.mon_clear_btn = QPushButton("清空监视")
        self.mon_clear_btn.clicked.connect(lambda *_: self.clear_monitor())
        self.mon_save_btn = QPushButton("保存日志")
        self.mon_save_btn.clicked.connect(lambda *_: self.save_monitor_log())
        mon_btns.addWidget(self.mon_clear_btn)
        mon_btns.addWidget(self.mon_save_btn)
        mon_btns.addStretch(1)
        mon_layout.addLayout(mon_btns)

        self.monitor_dock.setWidget(mon_container)
        self.addDockWidget(Qt.BottomDockWidgetArea, self.monitor_dock)
        self.monitor_dock.visibilityChanged.connect(lambda vis: vis and (self.schedule_monitor_render(full=True) or True))

        # 结构化日志缓存，用于 HEX/TEXT 模式重绘
        self._monitor_entries: List[dict] = []
        self._manual_entries: List[dict] = []
        self._motor_mon_entries: List[dict] = []

        # ---- UI 限频计时器（高帧率下避免卡顿） ----
        self._monitor_dirty = False
        self._manual_dirty = False
        self._monitor_timer = QTimer(self)
        self._monitor_timer.setInterval(50)
        self._monitor_timer.setSingleShot(True)
        self._monitor_timer.timeout.connect(self._flush_monitor_render)

        self._manual_timer = QTimer(self)
        self._manual_timer.setInterval(50)
        self._manual_timer.setSingleShot(True)
        self._manual_timer.timeout.connect(self._flush_manual_render)

        self._motor_mon_dirty = False
        self._motor_mon_timer = QTimer(self)
        self._motor_mon_timer.setInterval(50)
        self._motor_mon_timer.setSingleShot(True)
        self._motor_mon_timer.timeout.connect(self._flush_motor_monitor_render)

        # ---- 绘图限频（由刷新率驱动） ----
        self._plot_dirty = False
        self._plot_timer = QTimer(self)
        self._plot_timer.setInterval(16)  # 将由 _on_plot_fps_changed() 更新
        self._plot_timer.timeout.connect(self._flush_plot)
        # 注意：绘图计时器仅在采集中启动（见 _update_plot_timer_running）
        self._on_plot_fps_changed()

        self._last_xrange_update = 0.0
        self._last_yrange_update = 0.0

        # ---- 自定义发送停靠面板（默认隐藏，显示在右侧） ----
        self.custom_send_dock = QDockWidget("自定义串口发送", self)
        try:
            self.custom_send_dock.setObjectName('dock_custom_send')
        except Exception:
            pass
        self.custom_send_dock.setAllowedAreas(Qt.RightDockWidgetArea | Qt.LeftDockWidgetArea)

        send_container = QWidget()
        send_layout = QVBoxLayout(send_container)
        send_layout.setContentsMargins(8, 8, 8, 8)

        send_grid = QGridLayout()
        self.custom_send_port_combo = QComboBox()
        self.custom_send_line = QLineEdit()
        self.custom_send_line.setPlaceholderText("输入要发送的内容（文本或HEX：01 03 00 00 00 02）")
        self.custom_send_crlf_chk = QCheckBox("末尾添加 \\r\\n")
        self.custom_send_crlf_chk.setChecked(True)
        self.custom_send_btn = QPushButton("发送")

        send_grid.addWidget(QLabel("发送串口"), 0, 0)
        send_grid.addWidget(self.custom_send_port_combo, 0, 1)
        send_grid.addWidget(QLabel("发送内容"), 1, 0)
        send_grid.addWidget(self.custom_send_line, 1, 1)
        send_layout.addLayout(send_grid)

        # 显示模式
        mode_row = QHBoxLayout()
        mode_row.addWidget(QLabel("日志显示"))
        self.custom_send_mode_combo = QComboBox()
        self.custom_send_mode_combo.addItem("HEX", "hex")
        self.custom_send_mode_combo.addItem("文本(UTF-8)", "utf-8")
        self.custom_send_mode_combo.addItem("文本(GBK)", "gbk")
        self.custom_send_mode_combo.setCurrentIndex(1)
        self.custom_send_mode_combo.currentIndexChanged.connect(lambda *_: self.schedule_custom_send_render(full=True))
        mode_row.addWidget(self.custom_send_mode_combo)
        mode_row.addStretch(1)
        send_layout.addLayout(mode_row)

        send_row = QHBoxLayout()
        send_row.addWidget(self.custom_send_crlf_chk)
        send_row.addStretch(1)
        send_row.addWidget(self.custom_send_btn)
        send_layout.addLayout(send_row)

        tip = QLabel("提示：Enter 直接发送；可输入文本，或输入十六进制字节（如：01 03 00 00 00 02 / 010300000002）。")
        tip.setWordWrap(True)
        send_layout.addWidget(tip)

        # 回显/监听区域（过滤所选已打开端口的手动 TX/RX）
        send_layout.addWidget(QLabel("回复 / 日志"))
        self.custom_send_log = QPlainTextEdit()
        self.custom_send_log.setReadOnly(True)
        self.custom_send_log.setUndoRedoEnabled(False)
        try:
            self.custom_send_log.setMaximumBlockCount(1200)
        except Exception:
            pass
        self.custom_send_log.setPlaceholderText("这里会显示自定义发送对应的 TX/RX。")
        send_layout.addWidget(self.custom_send_log, 1)

        log_btns = QHBoxLayout()
        self.custom_send_clear_btn = QPushButton("清空")
        self.custom_send_clear_btn.clicked.connect(lambda *_: self.clear_custom_send_log())
        log_btns.addWidget(self.custom_send_clear_btn)
        log_btns.addStretch(1)
        send_layout.addLayout(log_btns)

        send_layout.addStretch(1)

        self.custom_send_dock.setWidget(send_container)
        self.addDockWidget(Qt.RightDockWidgetArea, self.custom_send_dock)
        self.custom_send_dock.hide()  # 默认关闭

        # ---- 电机控制停靠面板（默认隐藏） ----
        self.motor_dock = QDockWidget("电机控制", self)
        try:
            self.motor_dock.setObjectName('dock_motor')
        except Exception:
            pass
        self.motor_dock.setAllowedAreas(Qt.RightDockWidgetArea | Qt.LeftDockWidgetArea)

        motor_container = QWidget()
        motor_layout = QVBoxLayout(motor_container)
        motor_layout.setContentsMargins(8, 8, 8, 8)

        # 启用/禁用
        en_row = QHBoxLayout()
        self.motor_enable_btn = QPushButton("使能")
        self.motor_disable_btn = QPushButton("去使能")
        self.motor_enable_lamp = QLabel()
        self.motor_enable_lamp.setFixedSize(14, 14)
        self._set_lamp_color(self.motor_enable_lamp, "#777777")
        en_row.addWidget(self.motor_enable_btn)
        en_row.addWidget(self.motor_disable_btn)
        en_row.addWidget(QLabel("使能灯"))
        en_row.addWidget(self.motor_enable_lamp)
        en_row.addStretch(1)
        motor_layout.addLayout(en_row)

        # 方向
        dir_row = QHBoxLayout()
        self.motor_forward_btn = QPushButton("正转")
        self.motor_backward_btn = QPushButton("反转")
        self.motor_dir_lamp = QLabel()
        self.motor_dir_lamp.setFixedSize(14, 14)
        self._set_lamp_color(self.motor_dir_lamp, "#777777")
        dir_row.addWidget(self.motor_forward_btn)
        dir_row.addWidget(self.motor_backward_btn)
        dir_row.addWidget(QLabel("转向灯"))
        dir_row.addWidget(self.motor_dir_lamp)
        dir_row.addStretch(1)
        motor_layout.addLayout(dir_row)
        # 模式选择
        mode_row = QHBoxLayout()
        self.motor_mode_tension_btn = QPushButton("张力模式")
        self.motor_mode_speed_btn = QPushButton("速度模式")
        self.motor_mode_tension_lamp = QLabel()
        self.motor_mode_tension_lamp.setFixedSize(14, 14)
        self._set_lamp_color(self.motor_mode_tension_lamp, "#777777")
        self.motor_mode_speed_lamp = QLabel()
        self.motor_mode_speed_lamp.setFixedSize(14, 14)
        self._set_lamp_color(self.motor_mode_speed_lamp, "#777777")
        mode_row.addWidget(self.motor_mode_tension_btn)
        mode_row.addWidget(self.motor_mode_speed_btn)
        mode_row.addWidget(QLabel("张力模式灯"))
        mode_row.addWidget(self.motor_mode_tension_lamp)
        mode_row.addWidget(QLabel("速度模式灯"))
        mode_row.addWidget(self.motor_mode_speed_lamp)
        mode_row.addStretch(1)
        motor_layout.addLayout(mode_row)

        # 速度控制
        speed_row = QHBoxLayout()
        speed_row.addWidget(QLabel("转速(RPM)"))
        self.motor_speed_edit = QLineEdit()
        self.motor_speed_edit.setPlaceholderText("例如 100")
        self.motor_speed_btn = QPushButton("转速控制")
        speed_row.addWidget(self.motor_speed_edit, 1)
        speed_row.addWidget(self.motor_speed_btn)
        motor_layout.addLayout(speed_row)

        # 张力控制
        tension_row = QHBoxLayout()
        tension_row.addWidget(QLabel("张力(N)"))
        self.motor_tension_edit = QLineEdit()
        self.motor_tension_edit.setPlaceholderText("例如 1")
        self.motor_tension_btn = QPushButton("张力控制")
        tension_row.addWidget(self.motor_tension_edit, 1)
        tension_row.addWidget(self.motor_tension_btn)
        motor_layout.addLayout(tension_row)

        # PID 控制
        pid_row = QHBoxLayout()
        pid_row.addWidget(QLabel("Kp"))
        self.motor_kp_edit = QLineEdit()
        self.motor_kp_edit.setPlaceholderText("Kp:例如0.1")
        pid_row.addWidget(self.motor_kp_edit)
        pid_row.addWidget(QLabel("Ki"))
        self.motor_ki_edit = QLineEdit()
        self.motor_ki_edit.setPlaceholderText("Ki:例如0.02")
        pid_row.addWidget(self.motor_ki_edit)
        pid_row.addWidget(QLabel("Kd"))
        self.motor_kd_edit = QLineEdit()
        self.motor_kd_edit.setPlaceholderText("Kd:例如0.005")
        pid_row.addWidget(self.motor_kd_edit)
        self.motor_pid_btn = QPushButton("PID设置")
        pid_row.addWidget(self.motor_pid_btn)
        motor_layout.addLayout(pid_row)

        # 急停
        self.motor_estop_btn = QPushButton("急停")
        try:
            self.motor_estop_btn.setStyleSheet("background:#d9534f;color:white;font-weight:bold;")
        except Exception:
            pass
        try:
            self.motor_estop_btn.setFixedSize(160, 120)
        except Exception:
            pass
        estop_row = QHBoxLayout()
        estop_row.addStretch(1)
        estop_row.addWidget(self.motor_estop_btn)
        estop_row.addStretch(1)
        motor_layout.addLayout(estop_row)

        # 电机控制的 TX 监视
        motor_mon_group = QGroupBox("发送串口监视")
        motor_mon_layout = QVBoxLayout(motor_mon_group)
        motor_mon_layout.setContentsMargins(6, 6, 6, 6)
        motor_mon_mode_row = QHBoxLayout()
        motor_mon_mode_row.addWidget(QLabel("显示方式"))
        self.motor_mon_mode_combo = QComboBox()
        self.motor_mon_mode_combo.addItem("文本(UTF-8)", "utf-8")
        self.motor_mon_mode_combo.addItem("HEX", "hex")
        self.motor_mon_mode_combo.addItem("文本(GBK)", "gbk")
        self.motor_mon_mode_combo.setCurrentIndex(0)
        self.motor_mon_mode_combo.currentIndexChanged.connect(lambda *_: self.schedule_motor_monitor_render(full=True))
        motor_mon_mode_row.addWidget(self.motor_mon_mode_combo)
        motor_mon_mode_row.addStretch(1)
        self.motor_mon_clear_btn = QPushButton("清空")
        self.motor_mon_clear_btn.clicked.connect(self.clear_motor_monitor)
        motor_mon_mode_row.addWidget(self.motor_mon_clear_btn)
        motor_mon_layout.addLayout(motor_mon_mode_row)
        self.motor_mon_text = QPlainTextEdit()
        self.motor_mon_text.setReadOnly(True)
        self.motor_mon_text.setUndoRedoEnabled(False)
        try:
            self.motor_mon_text.setMaximumBlockCount(2000)
        except Exception:
            pass
        self.motor_mon_text.setPlaceholderText("这里显示发送串口返回的数据。")
        motor_mon_layout.addWidget(self.motor_mon_text, 1)
        motor_layout.addWidget(motor_mon_group, 1)


        self.motor_dock.setWidget(motor_container)
        self.addDockWidget(Qt.RightDockWidgetArea, self.motor_dock)
        self.motor_dock.hide()  # 默认隐藏
        self.motor_dock.visibilityChanged.connect(lambda vis: vis and (self.schedule_motor_monitor_render(full=True) or True))

        self.motor_enable_btn.clicked.connect(lambda *_: self.on_motor_enable())
        self.motor_disable_btn.clicked.connect(lambda *_: self.on_motor_disable())
        self.motor_forward_btn.clicked.connect(lambda *_: self.on_motor_forward())
        self.motor_backward_btn.clicked.connect(lambda *_: self.on_motor_backward())
        self.motor_speed_btn.clicked.connect(lambda *_: self.on_motor_speed())
        self.motor_tension_btn.clicked.connect(lambda *_: self.on_motor_tension())
        self.motor_pid_btn.clicked.connect(lambda *_: self.on_motor_pid())
        self.motor_estop_btn.clicked.connect(lambda *_: self.on_motor_estop())
        self.motor_mode_tension_btn.clicked.connect(lambda *_: self.on_motor_mode_tension())
        self.motor_mode_speed_btn.clicked.connect(lambda *_: self.on_motor_mode_speed())
        self.motor_mode = None

        self.custom_send_btn.clicked.connect(lambda *_: self.send_custom_serial())
        self.custom_send_line.returnPressed.connect(self.send_custom_serial)
        self.custom_send_dock.visibilityChanged.connect(lambda vis: vis and (self.update_custom_send_ports() or True) and (self.schedule_custom_send_render(full=True) or True))
        self.custom_send_port_combo.currentIndexChanged.connect(lambda *_: self.schedule_custom_send_render(full=True))

        # ---- 工作区菜单（显示/隐藏面板） ----
        ws_menu = self.menuBar().addMenu("工作区")

        act_serial = ws_menu.addAction("串口配置")
        act_serial.setCheckable(True)
        act_serial.setChecked(True)
        act_serial.toggled.connect(serial_box.setVisible)

        act_modbus = ws_menu.addAction("Modbus 配置")
        act_modbus.setCheckable(True)
        act_modbus.setChecked(True)
        act_modbus.toggled.connect(modbus_box.setVisible)

        act_plotset = ws_menu.addAction("绘图设置")
        act_plotset.setCheckable(True)
        act_plotset.setChecked(True)
        act_plotset.toggled.connect(plot_box.setVisible)

        act_channels = ws_menu.addAction("通道配置")
        act_channels.setCheckable(True)
        act_channels.setChecked(True)
        act_channels.toggled.connect(ch_box.setVisible)

        ws_menu.addSeparator()
        # 停靠面板自带的切换动作
        ws_menu.addAction(self.monitor_dock.toggleViewAction())
        ws_menu.addAction(self.custom_send_dock.toggleViewAction())

        # 监视器启用开关
        self.mon_rx_chk.toggled.connect(self.schedule_monitor_render)
        self.mon_tx_chk.toggled.connect(self.on_tx_monitor_toggled)

        # ---- 控制菜单 ----
        ctrl_menu = self.menuBar().addMenu("控制")
        act_motor = ctrl_menu.addAction("电机控制")
        act_motor.triggered.connect(self.open_motor_control)


        # ---- 绘图窗口菜单 ----
        plot_menu = self.menuBar().addMenu("绘图窗口")
        self.act_plot_window = plot_menu.addAction("显示绘图窗口")
        self.act_plot_window.setCheckable(True)
        self.act_plot_window.setChecked(True)
        self.act_plot_window.toggled.connect(lambda on: self.plot_dock.setVisible(on))
        self.plot_dock.visibilityChanged.connect(lambda vis: self.act_plot_window.setChecked(vis))

        self.act_friction_plot = plot_menu.addAction("摩擦力绘图窗口")
        self.act_friction_plot.setCheckable(True)
        self.act_friction_plot.setChecked(False)
        self.act_friction_plot.toggled.connect(lambda on: self._set_plot_tab_visible(self.friction_tab, "摩擦力", on))

        self.act_mu_plot = plot_menu.addAction("摩擦系数绘图窗口")
        self.act_mu_plot.setCheckable(True)
        self.act_mu_plot.setChecked(False)
        self.act_mu_plot.toggled.connect(lambda on: self._set_plot_tab_visible(self.mu_tab, "摩擦系数", on))

        # ---- 历史数据菜单 ----
        self.hist_menu = self.menuBar().addMenu("历史数据")
        self._build_history_menu()

        # ---- 串口模拟器菜单 ----
        sim_menu = self.menuBar().addMenu('串口仿真')
        act_sim = sim_menu.addAction('打开仿真串口界面')
        self.sim_manager = SerialSimManagerWindow(self)
        self.sim_manager.hide()
        act_sim.triggered.connect(self.open_simulator)
        self.sim_manager.ports_changed.connect(self.refresh_ports)


        # 初始化
        self.refresh_ports()
        # 默认两通道：两个 int16（01 03 00 00 00 02 ...）
        self.add_channel_row(default_name="CH1", default_addr=0, default_dtype="int16")
        self.add_channel_row(default_name="CH2", default_addr=1, default_dtype="int16")
        self._refresh_friction_channel_options()

        # 保持布局稳定（左侧面板宽度不抖动）并恢复上次工作区状态。
        self._apply_stable_widget_sizing()

        # 仅在窗口真正显示后恢复上次工作区状态。
        # 过早执行（在 __init__ 期间）可能导致部分系统首帧布局不完整。
        self._restored_once = False

    def showEvent(self, e):
        super().showEvent(e)
        if getattr(self, "_restored_once", False):
            return
        self._restored_once = True
        QTimer.singleShot(0, self._restore_after_show)
    def _restore_after_show(self):
        self._restore_window_layout()

        # 修复：保存的窗口位置可能部分在屏幕外（负 Y 等）
        self._ensure_frame_on_screen()
        QTimer.singleShot(50, self._ensure_frame_on_screen)
        QTimer.singleShot(180, self._ensure_frame_on_screen)

        # 强制进行几次布局/绘制。
        # 在某些系统（Windows + 高 DPI，或基于 OpenGL 的控件）上，第一帧
        # 在发生 resizeEvent（如用户拖动边框）之前可能不会完全布局/绘制。
        # 我们在不改变可见尺寸的情况下模拟一次。
        self._force_first_layout_pass()
        QTimer.singleShot(0, self._force_first_layout_pass)

    def _force_first_layout_pass(self):
        # 激活中央布局
        try:
            cw = self.centralWidget()
            if cw:
                try:
                    cw.updateGeometry()
                except Exception:
                    pass
                if cw.layout():
                    cw.layout().activate()
        except Exception:
            pass

        # 稳定分割器几何
        try:
            if hasattr(self, "main_splitter") and self.main_splitter is not None:
                try:
                    self.main_splitter.updateGeometry()
                except Exception:
                    pass
                # 重新应用当前尺寸以强制内部重新计算
                self.main_splitter.setSizes(self.main_splitter.sizes())
                # 防止恢复出错的分割器状态（如一侧几乎收缩）。
                try:
                    sizes = self.main_splitter.sizes()
                    if isinstance(sizes, (list, tuple)) and len(sizes) >= 2:
                        total = int(sizes[0]) + int(sizes[1])
                        if total > 0 and (int(sizes[0]) < 180 or int(sizes[1]) < 180):
                            left = max(360, int(total * 0.35))
                            right = max(420, total - left)
                            self.main_splitter.setSizes([left, right])
                except Exception:
                    pass

        except Exception:
            pass

        # 刷新一次待处理事件（仅首次显示）
        try:
            QApplication.processEvents()
        except Exception:
            pass

        # 微调窗口尺寸（触发类似手动拖边的 resizeEvent）
        try:
            is_max = getattr(self, "isMaximized", lambda: False)()
            is_full = getattr(self, "isFullScreen", lambda: False)()
            if not is_max and not is_full:
                w, h = int(self.width()), int(self.height())
                self.resize(w + 1, h + 1)
                self.resize(w, h)
        except Exception:
            pass

        # 让绘图控件稳定下来（pyqtgraph / OpenGL）
        try:
            if hasattr(self, "plot") and self.plot is not None:
                try:
                    self.plot.updateGeometry()
                except Exception:
                    pass
                self.plot.update()

                try:
                    self.plot.repaint()
                except Exception:
                    pass
        except Exception:
            pass

        try:
            self.update()
            self.repaint()
        except Exception:
            pass

        # 确保恢复的窗口框架在当前屏幕完全可见
        # （修复 QSettings 恢复出负 Y 等情况）
        self._ensure_frame_on_screen()

    def _ensure_frame_on_screen(self):
        """确保窗口框架在当前屏幕可用区域内。

        这用于修复 QSettings 恢复的窗口位置出现负 Y（顶部被裁剪）或部分在屏幕外的情况。
        我们使用 *frameGeometry*（含标题栏）而不是客户端区域进行限制。
        """
        try:
            # 不干预最大化/全屏状态
            try:
                if getattr(self, "isMaximized", lambda: False)() or getattr(self, "isFullScreen", lambda: False)():
                    return
            except Exception:
                pass

            screen = None
            try:
                screen = self.screen()
            except Exception:
                screen = None
            if screen is None:
                try:
                    screen = QApplication.primaryScreen()
                except Exception:
                    screen = None
            if screen is None:
                return

            avail = screen.availableGeometry()
            fg = self.frameGeometry()

            x, y = int(fg.x()), int(fg.y())
            w, h = int(fg.width()), int(fg.height())

            # 如果完全在屏幕外，则重新居中
            if (x + w) < (avail.left() + 20) or x > (avail.left() + avail.width() - 20) or (y + h) < (avail.top() + 20) or y > (avail.top() + avail.height() - 20):
                new_x = int(avail.left() + max(0, (avail.width() - w) // 2))
                new_y = int(avail.top() + max(0, (avail.height() - h) // 2))
                self.move(new_x, new_y)
                return

            # 限制到可用区域
            max_x = int(avail.left() + max(0, avail.width() - w))
            max_y = int(avail.top() + max(0, avail.height() - h))
            new_x = min(max(x, int(avail.left())), max_x)
            new_y = min(max(y, int(avail.top())), max_y)

            if new_x != x or new_y != y:
                self.move(new_x, new_y)
        except Exception:
            pass

    def open_simulator(self):
        """打开串口仿真界面（可同时仿真多个串口）。"""
        if not hasattr(self, 'sim_manager') or self.sim_manager is None:
            self.sim_manager = SerialSimManagerWindow(self)
        self.sim_manager.show()
        try:
            self.sim_manager.raise_()
            self.sim_manager.activateWindow()
        except Exception:
            pass


    def open_motor_control(self):
        if not hasattr(self, 'motor_dock') or self.motor_dock is None:
            return
        self.motor_dock.show()
        try:
            self.motor_dock.raise_()
            self.motor_dock.activateWindow()
        except Exception:
            pass

    def _apply_stable_widget_sizing(self):
        """防止下拉框文本长短变化导致左侧面板宽度抖动。

        这用于在点击连接/刷新等操作时保持当前工作区布局稳定。
        """
        combos = [
            getattr(self, 'port_combo', None),
            getattr(self, 'tx_port_combo', None),
            getattr(self, 'custom_send_port_combo', None),
        ]
        for cb in combos:
            if cb is None:
                continue
            # 优先使用不会随当前文本改变尺寸的策略。
            pol = None
            try:
                pol = QComboBox.SizeAdjustPolicy.AdjustToMinimumContentsLengthWithIcon
            except Exception:
                pol = getattr(QComboBox, 'AdjustToMinimumContentsLengthWithIcon', None)
                if pol is None:
                    pol = getattr(QComboBox, 'AdjustToMinimumContentsLength', None)
            try:
                if pol is not None:
                    cb.setSizeAdjustPolicy(pol)
            except Exception:
                pass
            try:
                cb.setMinimumContentsLength(28)
            except Exception:
                pass
            try:
                cb.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            except Exception:
                pass

        try:
            self.status_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        except Exception:
            pass

    def _apply_safe_geometry(self, x, y, w, h):
        """将几何尺寸限制到当前屏幕可用区域。

        这用于避免 Qt 警告，例如：
        QWindowsWindow::setGeometry: Unable to set geometry ...
        当恢复的窗口尺寸超过当前可用屏幕区域
        （任务栏、DPI 缩放、显示器变化等）时会出现该问题。
        """
        try:
            x = int(x); y = int(y); w = int(w); h = int(h)
        except Exception:
            return

        # 若最大化/全屏，不与窗口管理器对抗。
        try:
            if getattr(self, 'isMaximized', lambda: False)() or getattr(self, 'isFullScreen', lambda: False)():
                return
        except Exception:
            pass

        try:
            cx = x + max(0, w // 2)
            cy = y + max(0, h // 2)
            try:
                center = QPoint(cx, cy)
            except Exception:
                center = None

            screen = None
            try:
                if center is not None:
                    screen = QGuiApplication.screenAt(center)
            except Exception:
                screen = None
            if screen is None:
                try:
                    screen = QGuiApplication.primaryScreen()
                except Exception:
                    screen = None

            if screen is None:
                try:
                    self.setGeometry(x, y, w, h)
                except Exception:
                    pass
                return

            avail = screen.availableGeometry()

            try:
                minw = int(self.minimumWidth()) if int(self.minimumWidth()) > 0 else 640
            except Exception:
                minw = 640
            try:
                minh = int(self.minimumHeight()) if int(self.minimumHeight()) > 0 else 480
            except Exception:
                minh = 480

            w = max(minw, min(w, int(avail.width())))
            h = max(minh, min(h, int(avail.height())))

            x = min(max(x, int(avail.left())), int(avail.right()) - w + 1)
            y = min(max(y, int(avail.top())), int(avail.bottom()) - h + 1)

            self.setGeometry(x, y, w, h)
        except Exception:
            try:
                self.setGeometry(x, y, w, h)
            except Exception:
                pass

    def _ensure_window_on_screen(self):
        """限制当前窗口几何，使其在某个屏幕可见。"""
        try:
            if getattr(self, 'isMaximized', lambda: False)() or getattr(self, 'isFullScreen', lambda: False)():
                return
        except Exception:
            pass
        try:
            g = self.geometry()
            self._apply_safe_geometry(g.x(), g.y(), g.width(), g.height())
        except Exception:
            pass

    def _save_window_layout(self):
        s = getattr(self, '_settings', None)
        if s is None:
            return
        try:
            s.setValue('main/geometry', self.saveGeometry())
            s.setValue('main/state', self.saveState())
            try:
                s.setValue('main/wstate', int(self.windowState()))
            except Exception:
                pass
            try:
                g = self.normalGeometry() if getattr(self, 'isMaximized', lambda: False)() else self.geometry()
                s.setValue('main/rect', [int(g.x()), int(g.y()), int(g.width()), int(g.height())])
            except Exception:
                pass
            if hasattr(self, 'main_splitter'):
                s.setValue('main/splitter', self.main_splitter.saveState())
        except Exception:
            pass

    def _restore_window_layout(self):
        s = getattr(self, '_settings', None)
        if s is None:
            return
        had_split = False
        try:
            had_split = bool(s.value('main/splitter'))
        except Exception:
            had_split = False

        # --- 恢复主窗口矩形（优先） ---
        rect = None
        try:
            rect = s.value('main/rect')
        except Exception:
            rect = None
        if rect and isinstance(rect, (list, tuple)) and len(rect) >= 4:
            try:
                x, y, w, h = rect[:4]
                self._apply_safe_geometry(x, y, w, h)
            except Exception:
                pass
        else:
            # 兼容旧版本：回退到 Qt 的 saveGeometry/restoreGeometry
            try:
                geom = s.value('main/geometry')
                if geom:
                    self.restoreGeometry(geom)
            except Exception:
                pass
            # 限制以防恢复的几何尺寸不适配当前屏幕/DPI
            self._ensure_window_on_screen()

        # --- 恢复停靠面板/工具状态 ---
        try:
            state = s.value('main/state')
            if state:
                self.restoreState(state)
        except Exception:
            pass

        # --- 恢复分割器尺寸 ---
        try:
            sp = s.value('main/splitter')
            if sp and hasattr(self, 'main_splitter'):
                self.main_splitter.restoreState(sp)
        except Exception:
            pass

        # 恢复窗口最大化状态（在几何恢复之后）
        try:
            ws = s.value('main/wstate')
            ws_i = int(ws) if ws is not None else 0
            if ws_i & int(Qt.WindowMaximized):
                self.setWindowState(self.windowState() | Qt.WindowMaximized)
        except Exception:
            pass

        # 在 restoreState 之后再次限制（停靠/最小尺寸可能变化）
        self._ensure_window_on_screen()

        # 首次运行兜底：给左侧面板一个合理宽度。
        if not had_split:
            try:
                if hasattr(self, 'main_splitter'):
                    try:
                        if int(self.main_splitter.count()) >= 2:
                            self.main_splitter.setSizes([420, 1000])
                        else:
                            self.main_splitter.setSizes([1000])
                    except Exception:
                        pass
            except Exception:
                pass


    # ---------- UI 限频辅助 ----------
    def schedule_monitor_render(self, full: bool = False):
        """限制通讯监视器 UI 更新频率。

        full=True 强制全量重绘（如显示模式改变/面板显示）。
        """
        if full:
            setattr(self, "_monitor_force_full", True)
        self._monitor_dirty = True
        try:
            if not self._monitor_timer.isActive():
                self._monitor_timer.start()
        except Exception:
            self.render_monitor(force_full=bool(full))

    def schedule_custom_send_render(self, full: bool = False):
        if full:
            setattr(self, "_manual_force_full", True)
        self._manual_dirty = True
        try:
            if not self._manual_timer.isActive():
                self._manual_timer.start()
        except Exception:
            self.render_custom_send_log(force_full=bool(full))


    def schedule_motor_monitor_render(self, full: bool = False):
        if full:
            setattr(self, "_motor_mon_force_full", True)
        self._motor_mon_dirty = True
        try:
            if not self._motor_mon_timer.isActive():
                self._motor_mon_timer.start()
        except Exception:
            self.render_motor_monitor(force_full=bool(full))
    def _flush_monitor_render(self):
        if not getattr(self, "_monitor_dirty", False):
            return
        # 面板隐藏时，延迟到可见时再渲染
        if hasattr(self, "monitor_dock") and not self.monitor_dock.isVisible():
            self._monitor_dirty = False
            self._monitor_force_full = True
            return

        self._monitor_dirty = False
        force_full = bool(getattr(self, "_monitor_force_full", False))
        try:
            self.render_monitor(force_full=force_full)
        finally:
            self._monitor_force_full = False

        if getattr(self, "_monitor_dirty", False):
            try:
                if not self._monitor_timer.isActive():
                    self._monitor_timer.start()
            except Exception:
                pass


    def _flush_motor_monitor_render(self):
        if not getattr(self, "_motor_mon_dirty", False):
            return
        if hasattr(self, "motor_dock") and not self.motor_dock.isVisible():
            self._motor_mon_dirty = False
            self._motor_mon_force_full = True
            return

        self._motor_mon_dirty = False
        force_full = bool(getattr(self, "_motor_mon_force_full", False))
        try:
            self.render_motor_monitor(force_full=force_full)
        finally:
            self._motor_mon_force_full = False

        if getattr(self, "_motor_mon_dirty", False):
            try:
                if not self._motor_mon_timer.isActive():
                    self._motor_mon_timer.start()
            except Exception:
                pass
    def _flush_manual_render(self):
        if not getattr(self, "_manual_dirty", False):
            return
        if hasattr(self, "custom_send_dock") and not self.custom_send_dock.isVisible():
            self._manual_dirty = False
            self._manual_force_full = True
            return

        self._manual_dirty = False
        force_full = bool(getattr(self, "_manual_force_full", False))
        try:
            self.render_custom_send_log(force_full=force_full)
        finally:
            self._manual_force_full = False

        if getattr(self, "_manual_dirty", False):
            try:
                if not self._manual_timer.isActive():
                    self._manual_timer.start()
            except Exception:
                pass

    def _mark_plot_dirty(self, *args, **kwargs):
        self._plot_dirty = True

    def _set_plot_tab_visible(self, tab_widget: QWidget, title: str, visible: bool):
        if not hasattr(self, "plot_tabs"):
            return
        idx = self.plot_tabs.indexOf(tab_widget)
        if visible and idx < 0:
            self.plot_tabs.addTab(tab_widget, title)
            try:
                self.plot_tabs.setCurrentWidget(tab_widget)
            except Exception:
                pass
            try:
                self.plot_dock.show()
                self.plot_dock.raise_()
                self.plot_dock.activateWindow()
            except Exception:
                pass
        elif (not visible) and idx >= 0:
            try:
                self.plot_tabs.removeTab(idx)
            except Exception:
                pass

    def _refresh_friction_channel_options(self):
        if not hasattr(self, "fric_high_combo") or not hasattr(self, "fric_low_combo"):
            return
        names = []
        try:
            rows = int(self.ch_table.rowCount()) if hasattr(self, "ch_table") else 0
        except Exception:
            rows = 0
        for r in range(rows):
            name = None
            try:
                item = self.ch_table.item(r, 1) if hasattr(self, "ch_table") else None
                name = (item.text() if item else "").strip()
            except Exception:
                name = ""
            if not name:
                name = f"CH{r+1}"
            if name and name not in names:
                names.append(name)

        cur_high = self.fric_high_combo.currentText() if hasattr(self, "fric_high_combo") else ""
        cur_low = self.fric_low_combo.currentText() if hasattr(self, "fric_low_combo") else ""
        cur_high_mu = self.mu_high_combo.currentText() if hasattr(self, "mu_high_combo") else ""
        cur_low_mu = self.mu_low_combo.currentText() if hasattr(self, "mu_low_combo") else ""

        for combo in [self.fric_high_combo, self.fric_low_combo, getattr(self, "mu_high_combo", None), getattr(self, "mu_low_combo", None)]:
            if combo is None:
                continue
            try:
                combo.blockSignals(True)
                combo.clear()
                if not names:
                    combo.addItem("(无通道)")
                    combo.setEnabled(False)
                else:
                    for n in names:
                        combo.addItem(n)
                    combo.setEnabled(True)
            except Exception:
                pass
            finally:
                try:
                    combo.blockSignals(False)
                except Exception:
                    pass

        # 如可能则恢复选择
        try:
            if cur_high and cur_high in names:
                self.fric_high_combo.setCurrentText(cur_high)
            if cur_low and cur_low in names:
                self.fric_low_combo.setCurrentText(cur_low)
            if cur_high_mu and cur_high_mu in names and hasattr(self, "mu_high_combo"):
                self.mu_high_combo.setCurrentText(cur_high_mu)
            if cur_low_mu and cur_low_mu in names and hasattr(self, "mu_low_combo"):
                self.mu_low_combo.setCurrentText(cur_low_mu)
        except Exception:
            pass

        self._on_friction_config_changed()

    def _swap_friction_channels(self):
        try:
            if not self.fric_high_combo.isEnabled() or not self.fric_low_combo.isEnabled():
                return
            hi = self.fric_high_combo.currentIndex()
            lo = self.fric_low_combo.currentIndex()
            if hi < 0 or lo < 0:
                return
            self.fric_high_combo.setCurrentIndex(lo)
            self.fric_low_combo.setCurrentIndex(hi)
        except Exception:
            pass
        self._on_friction_config_changed()

    def _swap_mu_channels(self):
        try:
            if not self.mu_high_combo.isEnabled() or not self.mu_low_combo.isEnabled():
                return
            hi = self.mu_high_combo.currentIndex()
            lo = self.mu_low_combo.currentIndex()
            if hi < 0 or lo < 0:
                return
            self.mu_high_combo.setCurrentIndex(lo)
            self.mu_low_combo.setCurrentIndex(hi)
        except Exception:
            pass
        self._on_mu_config_changed()

    def _sync_mu_from_fric(self):
        if not hasattr(self, "mu_high_combo"):
            return
        try:
            self.mu_high_combo.blockSignals(True)
            self.mu_low_combo.blockSignals(True)
            self.mu_wrap_angle_spin.blockSignals(True)
            if self.mu_high_combo.isEnabled():
                self.mu_high_combo.setCurrentText(self.fric_high_combo.currentText())
            if self.mu_low_combo.isEnabled():
                self.mu_low_combo.setCurrentText(self.fric_low_combo.currentText())
            if self.mu_wrap_angle_spin.isEnabled():
                self.mu_wrap_angle_spin.setValue(self.wrap_angle_spin.value())
        except Exception:
            pass
        finally:
            try:
                self.mu_high_combo.blockSignals(False)
                self.mu_low_combo.blockSignals(False)
                self.mu_wrap_angle_spin.blockSignals(False)
            except Exception:
                pass

    def _sync_fric_from_mu(self):
        if not hasattr(self, "fric_high_combo"):
            return
        try:
            self.fric_high_combo.blockSignals(True)
            self.fric_low_combo.blockSignals(True)
            self.wrap_angle_spin.blockSignals(True)
            if self.fric_high_combo.isEnabled():
                self.fric_high_combo.setCurrentText(self.mu_high_combo.currentText())
            if self.fric_low_combo.isEnabled():
                self.fric_low_combo.setCurrentText(self.mu_low_combo.currentText())
            if self.wrap_angle_spin.isEnabled():
                self.wrap_angle_spin.setValue(self.mu_wrap_angle_spin.value())
        except Exception:
            pass
        finally:
            try:
                self.fric_high_combo.blockSignals(False)
                self.fric_low_combo.blockSignals(False)
                self.wrap_angle_spin.blockSignals(False)
            except Exception:
                pass

    def _on_friction_config_changed(self, *args):
        try:
            self._fric_high_name = (self.fric_high_combo.currentText() or "").strip()
            self._fric_low_name = (self.fric_low_combo.currentText() or "").strip()
        except Exception:
            self._fric_high_name = ""
            self._fric_low_name = ""
        try:
            self._wrap_angle_deg = float(self.wrap_angle_spin.value()) if hasattr(self, "wrap_angle_spin") else 0.0
        except Exception:
            self._wrap_angle_deg = 0.0
        try:
            self._wrap_angle_rad = math.radians(float(self._wrap_angle_deg)) if float(self._wrap_angle_deg) > 0 else 0.0
        except Exception:
            self._wrap_angle_rad = 0.0

        self._sync_quality_from_wrap()

        self._sync_mu_from_fric()
        self._recalc_friction_buffers()
        try:
            self._plot_seq = int(getattr(self, "_plot_seq", 0) or 0) + 1
        except Exception:
            pass
        self._plot_dirty = True
        try:
            self.update_plot()
        except Exception:
            pass

    def _on_mu_config_changed(self, *args):
        self._sync_fric_from_mu()
        self._on_friction_config_changed()

    def _on_quality_rmin_changed(self, *args):
        if getattr(self, '_quality_ui_syncing', False):
            return
        self._sync_quality_ui('main')
        # Rmin only affects quality check; no derived sync needed.
        try:
            self._plot_seq = int(getattr(self, '_plot_seq', 0) or 0) + 1
        except Exception:
            pass
        self._plot_dirty = True

    def _on_quality_rmin_changed_mu(self, *args):
        if getattr(self, '_quality_ui_syncing', False):
            return
        self._sync_quality_ui('mu')
        self._on_quality_rmin_changed()

    def _on_quality_mu_max_changed(self, *args):
        if getattr(self, '_quality_syncing', False) or getattr(self, '_quality_ui_syncing', False):
            return
        self._quality_last_source = 'mu'
        self._sync_quality_from_mu()
        self._sync_quality_ui('main')

    def _on_quality_mu_max_changed_mu(self, *args):
        if getattr(self, '_quality_ui_syncing', False):
            return
        self._sync_quality_ui('mu')
        self._on_quality_mu_max_changed()

    def _on_quality_rmax_changed(self, *args):
        if getattr(self, '_quality_syncing', False) or getattr(self, '_quality_ui_syncing', False):
            return
        self._quality_last_source = 'rmax'
        self._sync_quality_from_rmax()
        self._sync_quality_ui('main')

    def _on_quality_rmax_changed_mu(self, *args):
        if getattr(self, '_quality_ui_syncing', False):
            return
        self._sync_quality_ui('mu')
        self._on_quality_rmax_changed()

    def _on_quality_gap_timeout_changed(self, *args):
        if getattr(self, '_quality_ui_syncing', False):
            return
        try:
            self._quality_gap_timeout_s = float(self.qgap_spin.value()) if hasattr(self, 'qgap_spin') else 1.0
        except Exception:
            self._quality_gap_timeout_s = 1.0
        self._sync_quality_ui('main')

    def _on_quality_gap_timeout_changed_mu(self, *args):
        if getattr(self, '_quality_ui_syncing', False):
            return
        self._sync_quality_ui('mu')
        self._on_quality_gap_timeout_changed()

    def _sync_quality_ui(self, source: str):
        if getattr(self, '_quality_ui_syncing', False):
            return
        self._quality_ui_syncing = True
        try:
            if source == 'main':
                if hasattr(self, 'rmin_spin_mu'):
                    self.rmin_spin_mu.setValue(self.rmin_spin.value())
                if hasattr(self, 'mu_max_spin_mu'):
                    self.mu_max_spin_mu.setValue(self.mu_max_spin.value())
                if hasattr(self, 'rmax_spin_mu'):
                    self.rmax_spin_mu.setValue(self.rmax_spin.value())
                if hasattr(self, 'qgap_spin_mu'):
                    self.qgap_spin_mu.setValue(self.qgap_spin.value())
            elif source == 'mu':
                if hasattr(self, 'rmin_spin_mu'):
                    self.rmin_spin.setValue(self.rmin_spin_mu.value())
                if hasattr(self, 'mu_max_spin_mu'):
                    self.mu_max_spin.setValue(self.mu_max_spin_mu.value())
                if hasattr(self, 'rmax_spin_mu'):
                    self.rmax_spin.setValue(self.rmax_spin_mu.value())
                if hasattr(self, 'qgap_spin_mu'):
                    self.qgap_spin.setValue(self.qgap_spin_mu.value())
        finally:
            self._quality_ui_syncing = False

    def _sync_quality_from_wrap(self):
        if getattr(self, '_quality_syncing', False):
            return
        if getattr(self, '_quality_last_source', 'mu') == 'rmax':
            self._sync_quality_from_rmax()
        else:
            self._sync_quality_from_mu()

    def _sync_quality_from_mu(self):
        if getattr(self, '_quality_syncing', False):
            return
        self._quality_syncing = True
        try:
            mu_max = float(self.mu_max_spin.value()) if hasattr(self, 'mu_max_spin') else 0.0
        except Exception:
            mu_max = 0.0
        try:
            theta = float(getattr(self, '_wrap_angle_rad', 0.0) or 0.0)
        except Exception:
            theta = 0.0
        try:
            rmax = math.exp(mu_max * theta) if theta > 0 else 1.0
        except Exception:
            rmax = 1.0
        try:
            if hasattr(self, 'rmax_spin'):
                self.rmax_spin.setValue(float(rmax))
        finally:
            self._quality_syncing = False
        self._sync_quality_ui('main')

    def _sync_quality_from_rmax(self):
        if getattr(self, '_quality_syncing', False):
            return
        self._quality_syncing = True
        try:
            rmax = float(self.rmax_spin.value()) if hasattr(self, 'rmax_spin') else 1.0
        except Exception:
            rmax = 1.0
        try:
            theta = float(getattr(self, '_wrap_angle_rad', 0.0) or 0.0)
        except Exception:
            theta = 0.0
        if theta > 0 and rmax > 0:
            try:
                mu_max = math.log(rmax) / theta
            except Exception:
                mu_max = 0.0
        else:
            mu_max = 0.0
        try:
            if hasattr(self, 'mu_max_spin'):
                self.mu_max_spin.setValue(float(mu_max))
        finally:
            self._quality_syncing = False
        self._sync_quality_ui('main')

    def _get_quality_gap_timeout(self) -> float:
        try:
            return float(getattr(self, '_quality_gap_timeout_s', 1.0) or 1.0)
        except Exception:
            return 1.0
    def _calc_fric_mu(self, high_v, low_v):
        try:
            if high_v is None or low_v is None:
                return None, None
            high = float(high_v)
            low = float(low_v)
        except Exception:
            return None, None
        try:
            if not math.isfinite(high) or not math.isfinite(low):
                return None, None
        except Exception:
            pass
        fric = high - low
        mu = None
        theta = float(getattr(self, "_wrap_angle_rad", 0.0) or 0.0)
        if theta > 0 and low > 0 and high > 0:
            try:
                ratio = high / low
                if ratio > 0:
                    mu = math.log(ratio) / theta
            except Exception:
                mu = None
        return fric, mu

    def _update_friction_buffers_at_index(self, idx: int, row: dict):
        if self._fric_buf is None or self._mu_buf is None:
            return
        high_name = (getattr(self, "_fric_high_name", "") or "").strip()
        low_name = (getattr(self, "_fric_low_name", "") or "").strip()
        if (not high_name) or (not low_name):
            fric, mu = None, None
        else:
            fric, mu = self._calc_fric_mu(row.get(high_name), row.get(low_name))
        if np is not None:
            try:
                self._fric_buf[idx] = (np.nan if fric is None else float(fric))
                self._mu_buf[idx] = (np.nan if mu is None else float(mu))
            except Exception:
                pass
        else:
            self._fric_buf[idx] = fric
            self._mu_buf[idx] = mu

    def _recalc_friction_buffers(self):
        size = int(getattr(self, "_buf_size", 0) or 0)
        if size <= 0 or self._fric_buf is None or self._mu_buf is None:
            return
        high_name = (getattr(self, "_fric_high_name", "") or "").strip()
        low_name = (getattr(self, "_fric_low_name", "") or "").strip()
        high_buf = self._val_buf_by_channel.get(high_name) if high_name else None
        low_buf = self._val_buf_by_channel.get(low_name) if low_name else None
        for i in range(size):
            if high_buf is None or low_buf is None:
                fric, mu = None, None
            else:
                try:
                    hv = high_buf[i]
                    lv = low_buf[i]
                except Exception:
                    hv = None
                    lv = None
                fric, mu = self._calc_fric_mu(hv, lv)
            if np is not None:
                try:
                    self._fric_buf[i] = (np.nan if fric is None else float(fric))
                    self._mu_buf[i] = (np.nan if mu is None else float(mu))
                except Exception:
                    pass
            else:
                self._fric_buf[i] = fric
                self._mu_buf[i] = mu

    def _update_friction_plots(self, xs, idx: int, full: bool, count: int, scroll_live: bool, x_left: float, x_right: float):
        if xs is None or self._fric_buf is None or self._mu_buf is None:
            return
        size = int(getattr(self, "_buf_size", 0) or 0)
        if size <= 0:
            return

        # X 轴范围同步
        if scroll_live:
            try:
                self.friction_plot.setXRange(x_left, x_right, padding=0.0)
            except Exception:
                pass
            try:
                self.mu_plot.setXRange(x_left, x_right, padding=0.0)
            except Exception:
                pass

        if np is not None:
            if full:
                first = size - idx
                if self._fric_plot_y is None or getattr(self._fric_plot_y, "shape", (0,))[0] != size:
                    self._fric_plot_y = np.empty(size, dtype=float)
                if self._mu_plot_y is None or getattr(self._mu_plot_y, "shape", (0,))[0] != size:
                    self._mu_plot_y = np.empty(size, dtype=float)
                self._fric_plot_y[:first] = self._fric_buf[idx:]
                self._fric_plot_y[first:] = self._fric_buf[:idx]
                self._mu_plot_y[:first] = self._mu_buf[idx:]
                self._mu_plot_y[first:] = self._mu_buf[:idx]
                ys_fric = self._fric_plot_y
                ys_mu = self._mu_plot_y
            else:
                ys_fric = self._fric_buf[:count]
                ys_mu = self._mu_buf[:count]
            try:
                self.friction_curve.setData(xs, ys_fric, connect="finite", skipFiniteCheck=True)
            except Exception:
                try:
                    self.friction_curve.setData(xs, ys_fric, connect="finite")
                except Exception:
                    pass
            try:
                self.mu_curve.setData(xs, ys_mu, connect="finite", skipFiniteCheck=True)
            except Exception:
                try:
                    self.mu_curve.setData(xs, ys_mu, connect="finite")
                except Exception:
                    pass
        else:
            if full:
                fric_raw = list(self._fric_buf[idx:]) + list(self._fric_buf[:idx])
                mu_raw = list(self._mu_buf[idx:]) + list(self._mu_buf[:idx])
            else:
                fric_raw = list(self._fric_buf[:count])
                mu_raw = list(self._mu_buf[:count])
            xs_f, ys_f = [], []
            xs_m, ys_m = [], []
            for t, v in zip(xs, fric_raw):
                if v is None:
                    continue
                xs_f.append(t)
                ys_f.append(v)
            for t, v in zip(xs, mu_raw):
                if v is None:
                    continue
                xs_m.append(t)
                ys_m.append(v)
            try:
                self.friction_curve.setData(xs_f, ys_f)
            except Exception:
                pass
            try:
                self.mu_curve.setData(xs_m, ys_m)
            except Exception:
                pass

        # 派生曲线的自动缩放处理
        try:
            auto = bool(self.autoscale_chk.isChecked()) if hasattr(self, "autoscale_chk") else True
        except Exception:
            auto = True
        try:
            self.friction_plot.enableAutoRange(axis="y", enable=auto)
            self.mu_plot.enableAutoRange(axis="y", enable=auto)
        except Exception:
            pass
    def _flush_plot(self):
        # 绘图刷新由计时器（Hz）驱动，我们始终调用 update_plot()，
        # 它仅在有新采样时才重新上传曲线数据，但
        # 会在指定刷新率下保持 X 平滑滚动。
        self.update_plot()


    def _on_plot_fps_changed(self, *args):
        """将绘图刷新率（Hz）应用到绘图计时器。"""
        try:
            hz = int(self.plot_fps_spin.value()) if hasattr(self, 'plot_fps_spin') else 60
        except Exception:
            hz = 60
        hz = max(1, min(240, hz))
        interval_ms = max(1, int(round(1000.0 / float(hz))))
        try:
            self._plot_timer.setInterval(interval_ms)
        except Exception:
            pass

        # 根据当前状态应用启停
        try:
            self._update_plot_timer_running()
        except Exception:
            pass

    def _update_plot_timer_running(self):
        """根据采集状态启动/停止绘图计时器（节省 CPU，冻结滚动）。"""
        try:
            live = bool(getattr(self, "is_acquiring", False)) and (not bool(getattr(self, "is_paused", False)))
        except Exception:
            live = False

        if live:
            try:
                if not self._plot_timer.isActive():
                    self._plot_timer.start()
            except Exception:
                pass
        else:
            try:
                if self._plot_timer.isActive():
                    self._plot_timer.stop()
            except Exception:
                pass


    def _on_max_points_changed(self, *args):
        """最大点数变化时调整环形缓冲区大小。"""
        try:
            new_size = int(self.max_points_spin.value())
        except Exception:
            return
        self._resize_ring_buffers(new_size)
        self._mark_plot_dirty()

    def _resize_ring_buffers(self, new_size: int):
        new_size = int(max(10, new_size))
        old_size = int(getattr(self, '_buf_size', 0) or 0)
        if new_size <= 0 or new_size == old_size:
            return
        xs, ys_map, xs_wall, qf_vals = self._snapshot_ring(include_wall=True, include_quality=True)
        self._alloc_ring_buffers(
            new_size,
            list(self.channel_names),
            keep_last=True,
            xs=xs,
            ys_map=ys_map,
            xs_wall=xs_wall,
            qf_vals=qf_vals,
        )

    def _alloc_ring_buffers(self, size: int, channel_names: list, keep_last: bool = False, xs=None, ys_map=None, xs_wall=None, qf_vals=None):
        """分配环形缓冲区。

        当 keep_last=True 时，将最后 min(len(xs), size) 个样本复制到新缓冲区。
        """
        size = int(max(10, size))
        self._buf_size = size
        self._buf_count = 0
        self._buf_idx = 0
        self._plot_seq = 0
        self._last_plotted_seq = -1

        if np is not None:
            self._ts_buf = np.full(size, np.nan, dtype=float)
            self._ts_wall_buf = np.full(size, np.nan, dtype=float)
            self._plot_x = np.empty(size, dtype=float)
        else:
            self._ts_buf = [None] * size
            self._ts_wall_buf = [None] * size
            self._plot_x = None

        if np is not None:
            self._fric_buf = np.full(size, np.nan, dtype=float)
            self._mu_buf = np.full(size, np.nan, dtype=float)
            self._fric_plot_y = np.empty(size, dtype=float)
            self._mu_plot_y = np.empty(size, dtype=float)
        else:
            self._fric_buf = [None] * size
            self._mu_buf = [None] * size
            self._fric_plot_y = None
            self._mu_plot_y = None
        if np is not None:
            self._qf_buf = np.full(size, np.nan, dtype=float)
        else:
            self._qf_buf = [None] * size

        self._val_buf_by_channel = {}
        self._plot_y_by_channel = {}
        for name in channel_names:
            if np is not None:
                self._val_buf_by_channel[name] = np.full(size, np.nan, dtype=float)
                self._plot_y_by_channel[name] = np.empty(size, dtype=float)
            else:
                self._val_buf_by_channel[name] = [None] * size

        if keep_last and xs:
            try:
                k = min(len(xs), size)
            except Exception:
                k = 0
            if k > 0:
                tail_x = xs[-k:]
                if np is not None:
                    self._ts_buf[:k] = np.asarray(tail_x, dtype=float)
                else:
                    self._ts_buf[:k] = list(tail_x)
                if xs_wall:
                    try:
                        tail_w = xs_wall[-k:]
                    except Exception:
                        tail_w = []
                else:
                    tail_w = []
                if tail_w:
                    if np is not None:
                        self._ts_wall_buf[:k] = np.asarray(tail_w, dtype=float)
                    else:
                        self._ts_wall_buf[:k] = list(tail_w)

                for name in channel_names:
                    ys = (ys_map or {}).get(name, [])
                    tail_y = ys[-k:] if ys else [None] * k
                    if np is not None:
                        arr = np.asarray([(np.nan if v is None else float(v)) for v in tail_y], dtype=float)
                        self._val_buf_by_channel[name][:k] = arr
                    else:
                        self._val_buf_by_channel[name][:k] = list(tail_y)

                if qf_vals:
                    tail_qf = qf_vals[-k:] if len(qf_vals) >= k else list(qf_vals)
                    if len(tail_qf) < k:
                        tail_qf = ([None] * (k - len(tail_qf))) + list(tail_qf)
                    if np is not None:
                        try:
                            arr_qf = np.asarray([(np.nan if v is None else float(v)) for v in tail_qf], dtype=float)
                            self._qf_buf[:k] = arr_qf
                        except Exception:
                            pass
                    else:
                        self._qf_buf[:k] = list(tail_qf)

                # 对保留样本重新计算摩擦相关缓冲
                high_name = (getattr(self, "_fric_high_name", "") or "").strip()
                low_name = (getattr(self, "_fric_low_name", "") or "").strip()
                if high_name and low_name:
                    tail_high = (ys_map or {}).get(high_name, [])
                    tail_low = (ys_map or {}).get(low_name, [])
                    tail_high = tail_high[-k:] if tail_high else [None] * k
                    tail_low = tail_low[-k:] if tail_low else [None] * k
                    for j in range(k):
                        fric, mu = self._calc_fric_mu(tail_high[j], tail_low[j])
                        if np is not None:
                            try:
                                self._fric_buf[j] = (np.nan if fric is None else float(fric))
                                self._mu_buf[j] = (np.nan if mu is None else float(mu))
                            except Exception:
                                pass
                        else:
                            self._fric_buf[j] = fric
                            self._mu_buf[j] = mu
                else:
                    if np is not None:
                        try:
                            self._fric_buf[:k] = np.nan
                            self._mu_buf[:k] = np.nan
                        except Exception:
                            pass
                    else:
                        self._fric_buf[:k] = [None] * k
                        self._mu_buf[:k] = [None] * k
                self._buf_count = k
                self._buf_idx = k % size

    def _snapshot_ring(self, include_wall: bool = False, include_quality: bool = False):
        """将环形缓冲区快照为按时间排序的 Python 列表（用于调整大小/导出）。"""
        count = int(getattr(self, '_buf_count', 0) or 0)
        size = int(getattr(self, '_buf_size', 0) or 0)
        if count <= 0 or size <= 0 or self._ts_buf is None:
            if include_wall:
                if include_quality:
                    return [], {}, [], []
                return [], {}, []
            if include_quality:
                return [], {}, []
            return [], {}
        idx = int(getattr(self, '_buf_idx', 0) or 0)

        if count < size:
            # 未环绕
            if np is not None:
                xs = [float(x) for x in self._ts_buf[:count]]
            else:
                xs = list(self._ts_buf[:count])
        else:
            # 已环绕：最旧数据在 idx 处
            if np is not None:
                xs = [float(x) for x in self._ts_buf[idx:]] + [float(x) for x in self._ts_buf[:idx]]
            else:
                xs = list(self._ts_buf[idx:]) + list(self._ts_buf[:idx])

        xs_wall = []
        if include_wall:
            buf = self._ts_wall_buf
            if buf is None:
                xs_wall = []
            else:
                if count < size:
                    if np is not None:
                        arr = buf[:count]
                        xs_wall = [None if (not np.isfinite(v)) else float(v) for v in arr]
                    else:
                        xs_wall = list(buf[:count])
                else:
                    if np is not None:
                        arr = list(buf[idx:]) + list(buf[:idx])
                        xs_wall = [None if (not np.isfinite(v)) else float(v) for v in arr]
                    else:
                        xs_wall = list(buf[idx:]) + list(buf[:idx])

        ys_map = {}
        for name in list(self.channel_names):
            buf = self._val_buf_by_channel.get(name)
            if buf is None:
                continue
            if np is not None:
                if count < size:
                    arr = buf[:count]
                else:
                    arr = np.concatenate((buf[idx:], buf[:idx]))
                ys = [None if (not np.isfinite(v)) else float(v) for v in arr]
            else:
                if count < size:
                    ys = list(buf[:count])
                else:
                    ys = list(buf[idx:]) + list(buf[:idx])
            ys_map[name] = ys
        qf_vals = []
        if include_quality and self._qf_buf is not None:
            buf = self._qf_buf
            if np is not None:
                if count < size:
                    arr = buf[:count]
                else:
                    arr = np.concatenate((buf[idx:], buf[:idx]))
                qf_vals = [None if (not np.isfinite(v)) else float(v) for v in arr]
            else:
                if count < size:
                    qf_vals = list(buf[:count])
                else:
                    qf_vals = list(buf[idx:]) + list(buf[:idx])
        if include_wall:
            if include_quality:
                return xs, ys_map, xs_wall, qf_vals
            return xs, ys_map, xs_wall
        if include_quality:
            return xs, ys_map, qf_vals
        return xs, ys_map

    def _safe_float(self, v):
        try:
            fv = float(v)
        except Exception:
            return None
        try:
            if not math.isfinite(fv):
                return None
        except Exception:
            pass
        return fv

    def _row_data_ok(self, row: dict) -> bool:
        for name in self.channel_names:
            if self._safe_float(row.get(name, None)) is None:
                return False
        return True

    def _sanitize_row(self, row: dict) -> Dict[str, Optional[float]]:
        out: Dict[str, Optional[float]] = {}
        for name in self.channel_names:
            out[name] = self._safe_float(row.get(name, None))
        return out

    def _get_quality_params(self) -> Tuple[float, float]:
        try:
            rmin = float(self.rmin_spin.value()) if hasattr(self, "rmin_spin") else 1.01
        except Exception:
            rmin = 1.01
        try:
            rmax = float(self.rmax_spin.value()) if hasattr(self, "rmax_spin") else 1.0
        except Exception:
            rmax = 1.0
        return rmin, rmax

    def _calc_quality_flag(self, row: dict, data_ok: bool) -> int:
        if not data_ok:
            return 0

        high_name = (getattr(self, "_fric_high_name", "") or "").strip()
        low_name = (getattr(self, "_fric_low_name", "") or "").strip()
        if not high_name or not low_name:
            return 0

        high_v = self._safe_float(row.get(high_name, None))
        low_v = self._safe_float(row.get(low_name, None))
        if high_v is None or low_v is None:
            return 0

        if high_v <= 0 or low_v <= 0:
            return 0

        rmin, rmax = self._get_quality_params()
        try:
            ratio = float(high_v) / float(low_v)
        except Exception:
            return 0
        if ratio < rmin or ratio > rmax:
            return 0

        if getattr(self, "motor_mode", None) == 0:
            last_t = getattr(self, "_last_tension_setpoint", None)
            try:
                last_t = float(last_t) if last_t is not None else None
            except Exception:
                last_t = None
            if last_t is not None and last_t > 0:
                tmin = 0.05 * last_t
                if high_v < tmin or low_v < tmin:
                    return 0

        return 1

    def _commit_sample(self, mono_ts: float, wall_ts: float, row: dict, quality_flag: int):
        # Compute relative time based on the sample's monotonic timestamp.
        if self._t0_mono_ts is None:
            self._t0_mono_ts = float(mono_ts)
        pause_accum = float(getattr(self, '_mono_pause_accum', 0.0) or 0.0)
        rel_ts = float(float(mono_ts) - float(self._t0_mono_ts) - pause_accum)
        if rel_ts < 0.0:
            rel_ts = 0.0

        self._last_sample_rel_ts = rel_ts
        self._last_sample_mono_ts = float(mono_ts)

        size = int(getattr(self, '_buf_size', 0) or 0)
        if size <= 0:
            return
        i = int(getattr(self, '_buf_idx', 0) or 0) % size

        if np is not None:
            try:
                self._ts_buf[i] = float(rel_ts)
            except Exception:
                self._ts_buf[i] = np.nan
            try:
                if self._ts_wall_buf is not None:
                    self._ts_wall_buf[i] = float(wall_ts)
            except Exception:
                try:
                    if self._ts_wall_buf is not None:
                        self._ts_wall_buf[i] = np.nan
                except Exception:
                    pass
            for name in self.channel_names:
                v = row.get(name, None)
                try:
                    self._val_buf_by_channel[name][i] = (np.nan if v is None else float(v))
                except Exception:
                    self._val_buf_by_channel[name][i] = np.nan
            try:
                if self._qf_buf is not None:
                    self._qf_buf[i] = float(quality_flag)
            except Exception:
                pass
        else:
            self._ts_buf[i] = rel_ts
            if self._ts_wall_buf is not None:
                self._ts_wall_buf[i] = wall_ts
            for name in self.channel_names:
                self._val_buf_by_channel[name][i] = row.get(name, None)
            try:
                if self._qf_buf is not None:
                    self._qf_buf[i] = int(quality_flag)
            except Exception:
                pass

        # Update derived buffers (friction/mu) using the committed row values.
        self._update_friction_buffers_at_index(i, row)

        try:
            if getattr(self, "_log_db_path", ""):
                row_for_log = {name: row.get(name, None) for name in self.channel_names}
                row_for_log[self._quality_flag_name] = int(quality_flag)
                self._data_logger.append(wall_ts, row_for_log)
        except Exception:
            pass

        self._buf_idx = (i + 1) % size
        if int(getattr(self, '_buf_count', 0) or 0) < size:
            self._buf_count += 1

        self._plot_seq = int(getattr(self, '_plot_seq', 0) or 0) + 1
        self._plot_dirty = True

    def _trigger_comm_gap_stop(self):
        if getattr(self, "_quality_gap_triggered", False):
            return
        self._quality_gap_triggered = True
        try:
            self.on_motor_estop()
        except Exception:
            pass
        try:
            self.stop_acquire()
        except Exception:
            pass
        try:
            QMessageBox.warning(self, "通信异常", "连续通信丢包或解析失败超过 1 秒，已急停并停止采集。")
        except Exception:
            pass

    def _process_quality_sample(self, mono_ts: float, wall_ts: float, row: dict):
        data_ok = self._row_data_ok(row)

        if not data_ok:
            if self._quality_gap_start_mono is None:
                self._quality_gap_start_mono = float(mono_ts)
                self._quality_gap_triggered = False

            if (float(mono_ts) - float(self._quality_gap_start_mono)) >= self._get_quality_gap_timeout():
                self._trigger_comm_gap_stop()
                return

            pending = self._quality_gap_pending
            pending.append({"mono": mono_ts, "wall": wall_ts, "row": row})

            if self._quality_gap_hold_mode:
                hold_row = self._last_valid_row or self._sanitize_row(row)
                self._commit_sample(mono_ts, wall_ts, hold_row, 0)
                return

            if len(pending) > 3:
                hold_row = self._last_valid_row
                if hold_row is None:
                    for s in pending:
                        self._commit_sample(s["mono"], s["wall"], self._sanitize_row(s["row"]), 0)
                else:
                    for s in pending:
                        self._commit_sample(s["mono"], s["wall"], hold_row, 0)
                pending.clear()
                self._quality_gap_hold_mode = True
            return

        # data ok: flush pending gaps if any
        if self._quality_gap_start_mono is not None:
            self._quality_gap_start_mono = None
            self._quality_gap_triggered = False
        self._quality_gap_hold_mode = False

        pending = self._quality_gap_pending
        if pending:
            m = len(pending)
            if self._last_valid_row is not None and m <= 3:
                last_row = self._last_valid_row
                cur_row = self._sanitize_row(row)
                for idx, s in enumerate(pending, start=1):
                    frac = float(idx) / float(m + 1)
                    interp_row: Dict[str, Optional[float]] = {}
                    for name in self.channel_names:
                        v0 = last_row.get(name, None)
                        v1 = cur_row.get(name, None)
                        if v0 is None or v1 is None:
                            v = v0 if v0 is not None else v1
                        else:
                            v = float(v0) + (float(v1) - float(v0)) * frac
                        interp_row[name] = v
                    self._commit_sample(s["mono"], s["wall"], interp_row, 0)
            else:
                hold_row = self._last_valid_row
                for s in pending:
                    if hold_row is None:
                        self._commit_sample(s["mono"], s["wall"], self._sanitize_row(s["row"]), 0)
                    else:
                        self._commit_sample(s["mono"], s["wall"], hold_row, 0)
            pending.clear()
            self._quality_gap_hold_mode = False

        clean_row = self._sanitize_row(row)
        qf = self._calc_quality_flag(clean_row, True)
        self._commit_sample(mono_ts, wall_ts, clean_row, qf)
        self._last_valid_row = clean_row

    # ---------- 监视 ----------
    def append_monitor(self, s: str):
        """追加一行纯文本信息（非帧）。

        注意：为避免高帧率卡顿，UI 渲染会限频。
        """
        self._monitor_entries.append({"kind": "INFO", "data": b"", "tag": "", "note": str(s)})
        self.schedule_monitor_render()

    def _custom_send_current_tag(self) -> str:
        """返回自定义发送面板的当前过滤标签（如 'rx:COM3' / 'tx:COM4'）。"""
        if not self.is_connected or self.worker is None:
            return ""
        if not hasattr(self, "custom_send_port_combo"):
            return ""
        target = self.custom_send_port_combo.currentData()
        if target == "rx":
            p = getattr(self.worker, "port", "") or ""
            return f"rx:{p}" if p else ""
        if target == "tx":
            p = getattr(self.worker, "tx_port", "") or ""
            return f"tx:{p}" if p else ""
        return ""
    def on_tx_monitor_toggled(self, checked: bool):
        """启用/禁用 TX 端口异步 RX 监听钩子。

        仅影响通讯监视窗口对发送串口的 RX 监听（窥探式读取）。
        """
        if self.is_connected and self.worker is not None:
            try:
                self.worker.set_tx_tap_enabled(bool(checked))
            except Exception:
                pass
        self.schedule_monitor_render()


    def clear_monitor(self):
        self._monitor_entries.clear()
        self.monitor_text.clear()
        self._monitor_render_idx = 0
        self._monitor_render_mode = None

    def clear_custom_send_log(self):
        self._manual_entries.clear()
        self.custom_send_log.clear()
        self._manual_render_idx = 0
        self._manual_render_mode = None
        self._manual_render_tag = None


    def clear_motor_monitor(self):
        self._motor_mon_entries.clear()
        if hasattr(self, "motor_mon_text"):
            self.motor_mon_text.clear()
        self._motor_mon_render_idx = 0
        self._motor_mon_render_mode = None
    def _decode_bytes(self, data: bytes, mode: str) -> str:
        if mode == "hex":
            return hex_bytes(data)
        enc = "utf-8" if mode == "utf-8" else "gbk"
        try:
            return data.decode(enc, errors="replace")
        except Exception:
            return data.decode("utf-8", errors="replace")

    def _format_entry(self, e: dict, mode: str) -> str:
        kind = e.get("kind", "")
        tag = e.get("tag", "")
        note = e.get("note", "")
        data = e.get("data", b"") or b""
        prefix = kind
        if tag:
            prefix += f"[{tag}]"
        if data:
            payload = self._decode_bytes(data, mode)
            if note:
                return f"{prefix}: {payload}  {note}"
            return f"{prefix}: {payload}"
        # 无数据
        return f"{prefix}: {note}" if prefix else str(note)

    def render_monitor(self, force_full: bool = False):
        mode = "hex"
        if hasattr(self, "monitor_mode_combo"):
            mode = self.monitor_mode_combo.currentData() or "hex"

        # 记录模式/索引用于增量追加渲染
        last_mode = getattr(self, "_monitor_render_mode", None)
        render_idx = int(getattr(self, "_monitor_render_idx", 0) or 0)

        # 若模式变更或强制刷新，则重建最近 N 行（少见路径）
        if force_full or last_mode != mode or render_idx <= 0:
            max_lines = 2000
            entries = self._monitor_entries[-max_lines:]
            try:
                self.monitor_text.blockSignals(True)
                self.monitor_text.setPlainText("\n".join(self._format_entry(e, mode) for e in entries))
                self.monitor_text.blockSignals(False)
                self.monitor_text.moveCursor(QTextCursor.End)
            except Exception:
                pass
            self._monitor_render_mode = mode
            self._monitor_render_idx = len(self._monitor_entries)
            return

        # 增量追加新行
        if render_idx < len(self._monitor_entries):
            new_entries = self._monitor_entries[render_idx:]
            # 批量追加以减少 UI 更新
            lines = [self._format_entry(e, mode) for e in new_entries]
            if lines:
                try:
                    self.monitor_text.blockSignals(True)
                    self.monitor_text.appendPlainText("\n".join(lines))
                    self.monitor_text.blockSignals(False)
                    self.monitor_text.moveCursor(QTextCursor.End)
                except Exception:
                    pass
        self._monitor_render_mode = mode
        self._monitor_render_idx = len(self._monitor_entries)

    def render_custom_send_log(self, force_full: bool = False):
        if not hasattr(self, "custom_send_mode_combo"):
            return
        mode = self.custom_send_mode_combo.currentData() or "hex"
        tag_filter = self._custom_send_current_tag()

        last_mode = getattr(self, "_manual_render_mode", None)
        last_tag = getattr(self, "_manual_render_tag", None)
        render_idx = int(getattr(self, "_manual_render_idx", 0) or 0)

        if force_full or last_mode != mode or last_tag != tag_filter or render_idx <= 0:
            max_lines = 1200
            entries = [e for e in self._manual_entries if (not tag_filter or e.get("tag") == tag_filter)]
            entries = entries[-max_lines:]
            try:
                self.custom_send_log.blockSignals(True)
                self.custom_send_log.setPlainText("\n".join(self._format_entry(e, mode) for e in entries))
                self.custom_send_log.blockSignals(False)
                self.custom_send_log.moveCursor(QTextCursor.End)
            except Exception:
                pass
            self._manual_render_mode = mode
            self._manual_render_tag = tag_filter
            self._manual_render_idx = len(self._manual_entries)
            return

        if render_idx < len(self._manual_entries):
            new_entries = self._manual_entries[render_idx:]
            lines = []
            for e in new_entries:
                if (not tag_filter) or (e.get("tag") == tag_filter):
                    lines.append(self._format_entry(e, mode))
            if lines:
                try:
                    self.custom_send_log.blockSignals(True)
                    self.custom_send_log.appendPlainText("\n".join(lines))
                    self.custom_send_log.blockSignals(False)
                    self.custom_send_log.moveCursor(QTextCursor.End)
                except Exception:
                    pass
        self._manual_render_mode = mode
        self._manual_render_tag = tag_filter
        self._manual_render_idx = len(self._manual_entries)


    def render_motor_monitor(self, force_full: bool = False):
        mode = "utf-8"
        if hasattr(self, "motor_mon_mode_combo"):
            mode = self.motor_mon_mode_combo.currentData() or "utf-8"
        last_mode = getattr(self, "_motor_mon_render_mode", None)
        render_idx = int(getattr(self, "_motor_mon_render_idx", 0) or 0)
        max_lines = 2000

        if force_full or last_mode != mode or render_idx > len(self._motor_mon_entries):
            entries = self._motor_mon_entries[-max_lines:]
            try:
                self.motor_mon_text.blockSignals(True)
                self.motor_mon_text.setPlainText("\n".join(self._format_entry(e, mode) for e in entries))
                self.motor_mon_text.blockSignals(False)
                self.motor_mon_text.moveCursor(QTextCursor.End)
            except Exception:
                pass
            self._motor_mon_render_mode = mode
            self._motor_mon_render_idx = len(self._motor_mon_entries)
            return

        if render_idx < len(self._motor_mon_entries):
            new_entries = self._motor_mon_entries[render_idx:]
            lines_out = [self._format_entry(e, mode) for e in new_entries]
            if lines_out:
                try:
                    self.motor_mon_text.blockSignals(True)
                    self.motor_mon_text.appendPlainText("\n".join(lines_out))
                    self.motor_mon_text.blockSignals(False)
                    self.motor_mon_text.moveCursor(QTextCursor.End)
                except Exception:
                    pass
        self._motor_mon_render_mode = mode
        self._motor_mon_render_idx = len(self._motor_mon_entries)
    @Slot(str, bytes, str, str)
    def on_frame(self, kind: str, data: bytes, tag: str, note: str):
        e = {"kind": str(kind), "data": bytes(data or b""), "tag": str(tag or ""), "note": str(note or "")}

        # 为自定义发送面板始终保留手动 TX/RX 记录
        is_manual = str(kind).startswith('TX_MANUAL') or str(kind).startswith('RX_MANUAL')
        if is_manual:
            self._manual_entries.append(e)
            if self.custom_send_dock.isVisible():
                self.schedule_custom_send_render()

            # 限制日志内存占用
            if len(self._manual_entries) > 6000:
                overflow = len(self._manual_entries) - 6000
                del self._manual_entries[:overflow]
                try:
                    self._manual_render_idx = max(0, int(getattr(self, '_manual_render_idx', 0)) - overflow)
                except Exception:
                    self._manual_render_idx = 0

        # 通讯监视器按端口过滤（rx/tx 监听复选框）
        t = e.get('tag', '') or ''
        allow = True
        if t.startswith('rx:') and hasattr(self, 'mon_rx_chk'):
            allow = bool(self.mon_rx_chk.isChecked())
        elif t.startswith('tx:') and hasattr(self, 'mon_tx_chk'):
            allow = bool(self.mon_tx_chk.isChecked())

        if allow:
            self._monitor_entries.append(e)
            self.schedule_monitor_render()

            if len(self._monitor_entries) > 12000:
                overflow = len(self._monitor_entries) - 12000
                del self._monitor_entries[:overflow]
                try:
                    self._monitor_render_idx = max(0, int(getattr(self, '_monitor_render_idx', 0)) - overflow)
                except Exception:
                    self._monitor_render_idx = 0



        # 电机 TX 监视：仅显示 tx 端口的 RX 帧
        t2 = e.get("tag", "") or ""
        k2 = str(kind)
        if t2.startswith("tx:") and k2.startswith("RX"):
            self._motor_mon_entries.append(e)
            self.schedule_motor_monitor_render()
            if len(self._motor_mon_entries) > 8000:
                overflow = len(self._motor_mon_entries) - 8000
                del self._motor_mon_entries[:overflow]
                try:
                    self._motor_mon_render_idx = max(0, int(getattr(self, "_motor_mon_render_idx", 0)) - overflow)
                except Exception:
                    self._motor_mon_render_idx = 0
    def save_monitor_log(self):
        path, _ = QFileDialog.getSaveFileName(self, "保存通讯日志", "comm_log.txt", "Text Files (*.txt)")
        if not path:
            return
        if not path.lower().endswith(".txt"):
            path += ".txt"
        try:
            with open(path, "w", encoding="utf-8") as f:
                f.write(self.monitor_text.toPlainText())
            self.set_status(f"通讯日志已保存：{path}")
        except Exception as e:
            QMessageBox.critical(self, "保存失败", f"保存通讯日志失败：\n{e}")

    # ---------- 状态 ----------
    def set_status(self, msg: str):
        self.status_label.setText(f"状态：{msg}")

    # ---------- 端口 ----------
    def refresh_ports(self):
        """刷新物理 COM 端口和应用内模拟端口。"""
        # 保留当前选择，避免 UI 跳动。
        cur_rx = self.port_combo.currentData()
        cur_tx = self.tx_port_combo.currentData()

        self.port_combo.clear()
        self.tx_port_combo.clear()

        items = []

        # 物理端口
        try:
            ports = list(list_ports.comports())
        except Exception:
            ports = []

        for p in ports:
            desc = p.description or ""
            manu = getattr(p, "manufacturer", "") or ""
            prod = getattr(p, "product", "") or ""
            hwid = getattr(p, "hwid", "") or ""
            extra = " | ".join([x for x in [desc, manu, prod] if x and x.lower() != "n/a"])
            label = f"{p.device}"
            if extra:
                label += f"  —  {extra}"
            elif hwid:
                label += f"  —  {hwid}"
            items.append((label, p.device))

        # 模拟端口
        for info in SIM_REGISTRY.list_infos():
            label = f"{info.com}  —  仿真串口"
            items.append((label, info.key))

        if not items:
            self.port_combo.addItem("(未发现串口)", "")
            self.tx_port_combo.addItem("(未发现串口)", "")
        else:
            for label, key in items:
                self.port_combo.addItem(label, key)
                self.tx_port_combo.addItem(label, key)
        # 如可能则恢复选择
        try:
            if cur_rx:
                idx = self.port_combo.findData(cur_rx)
                if idx >= 0:
                    self.port_combo.setCurrentIndex(idx)
            if cur_tx:
                idx2 = self.tx_port_combo.findData(cur_tx)
                if idx2 >= 0:
                    self.tx_port_combo.setCurrentIndex(idx2)
        except Exception:
            pass

        # 保持自定义发送面板端口列表同步
        self.update_custom_send_ports()
    def update_custom_send_ports(self):
        """只显示本程序已打开的端口。"""
        if not hasattr(self, "custom_send_port_combo"):
            return

        combo = self.custom_send_port_combo
        combo.blockSignals(True)
        combo.clear()

        items = []
        if self.is_connected and self.worker is not None:
            # 当 worker 发出 connected(True) 时，rx/modbus 端口总是已连接
            rx_port = getattr(self.worker, "port", "")
            if rx_port:
                items.append((f"接收串口(Modbus)：{rx_port}", "rx"))

            # tx/输出端口可能被启用
            if getattr(self.worker, "tx_enabled", False) and getattr(self.worker, "_tx_ser", None) is not None:
                tx_port = getattr(self.worker, "tx_port", "")
                if tx_port and tx_port != rx_port:
                    items.append((f"发送串口(输出)：{tx_port}", "tx"))
                elif tx_port:
                    items.append((f"发送串口(输出)：{tx_port}", "tx"))

        if not items:
            combo.addItem("(未连接串口)", "")
            if hasattr(self, "custom_send_line"):
                self.custom_send_line.setEnabled(False)
            if hasattr(self, "custom_send_btn"):
                self.custom_send_btn.setEnabled(False)
        else:
            for text, data in items:
                combo.addItem(text, data)
            if hasattr(self, "custom_send_line"):
                self.custom_send_line.setEnabled(True)
            if hasattr(self, "custom_send_btn"):
                self.custom_send_btn.setEnabled(True)

        combo.blockSignals(False)

    def send_custom_serial(self):
        if not self.is_connected or self.worker is None:
            QMessageBox.information(self, "提示", "请先连接串口后再发送。")
            return

        target = self.custom_send_port_combo.currentData()
        if not target:
            QMessageBox.warning(self, "提示", "当前没有可发送的已连接串口。")
            return

        text = (self.custom_send_line.text() if self.custom_send_line else "").strip()
        if not text:
            return

        add_crlf = bool(self.custom_send_crlf_chk.isChecked())
        self.worker.enqueue_custom_send(str(target), text, add_crlf)
        # 不清空输入（方便连续修改/重复发送）
        self.custom_send_line.setFocus()

    # ---------- 电机控制 ----------
    def _set_lamp_color(self, label: QLabel, color: str):
        if label is None:
            return
        try:
            label.setStyleSheet(f"background:{color};border-radius:7px;border:1px solid #444;")
        except Exception:
            pass


    def _set_motor_mode_lamps(self, mode: Optional[int]):
        if mode == 0:
            self._set_lamp_color(self.motor_mode_tension_lamp, "#5cb85c")
            self._set_lamp_color(self.motor_mode_speed_lamp, "#777777")
        elif mode == 1:
            self._set_lamp_color(self.motor_mode_tension_lamp, "#777777")
            self._set_lamp_color(self.motor_mode_speed_lamp, "#5cb85c")
        else:
            self._set_lamp_color(self.motor_mode_tension_lamp, "#777777")
            self._set_lamp_color(self.motor_mode_speed_lamp, "#777777")

    def _require_motor_mode(self, required: Optional[int] = None) -> bool:
        if getattr(self, "motor_mode", None) is None:
            QMessageBox.warning(self, "提示", "请先选择控制模式。")
            return False
        if required is not None and int(self.motor_mode) != int(required):
            mode_name = "张力模式" if int(required) == 0 else "速度模式"
            QMessageBox.warning(self, "提示", f"请先切换到{mode_name}。")
            return False
        return True

    def _motor_can_send(self) -> bool:
        if not self.is_connected or self.worker is None:
            QMessageBox.information(self, "提示", "请先连接串口后再操作电机控制。")
            return False
        if not getattr(self.worker, "tx_enabled", False) or getattr(self.worker, "_tx_ser", None) is None:
            QMessageBox.information(self, "提示", "请先启用“发送串口(输出)”并连接。")
            return False
        return True

    def _send_motor_cmd(self, cmd: str) -> bool:
        if not self._motor_can_send():
            return False
        text = (cmd or "").strip()
        if not text:
            return False
        self.worker.enqueue_custom_send("tx", text, True)
        return True

    def _parse_number_text(self, text: str, label: str) -> Optional[str]:
        t = (text or "").strip()
        if not t:
            QMessageBox.warning(self, "提示", f"请输入{label}。")
            return None
        try:
            v = float(t)
        except Exception:
            QMessageBox.warning(self, "提示", f"{label}不是有效数字。")
            return None
        if v != v or v in (float("inf"), float("-inf")):
            QMessageBox.warning(self, "提示", f"{label}不是有效数字。")
            return None
        return f"{v:g}"

    def on_motor_mode_tension(self):
        if self._send_motor_cmd("ConMode 0"):
            self.motor_mode = 0
            self._set_motor_mode_lamps(self.motor_mode)

    def on_motor_mode_speed(self):
        if self._send_motor_cmd("ConMode 1"):
            self.motor_mode = 1
            self._set_motor_mode_lamps(self.motor_mode)

    def on_motor_enable(self):
        if self._send_motor_cmd("Enable"):
            self._set_lamp_color(self.motor_enable_lamp, "#5cb85c")

    def on_motor_disable(self):
        if self._send_motor_cmd("Disable"):
            self._set_lamp_color(self.motor_enable_lamp, "#777777")

    def on_motor_forward(self):
        if self._send_motor_cmd("Forward"):
            self._set_lamp_color(self.motor_dir_lamp, "#5cb85c")

    def on_motor_backward(self):
        if self._send_motor_cmd("Backward"):
            self._set_lamp_color(self.motor_dir_lamp, "#f0ad4e")

    def on_motor_speed(self):
        if not self._require_motor_mode(1):
            return
        val = self._parse_number_text(self.motor_speed_edit.text(), "转速(RPM)")
        if val is None:
            return
        self._send_motor_cmd(f"Con {val}")

    def on_motor_tension(self):
        if not self._require_motor_mode(0):
            return
        val = self._parse_number_text(self.motor_tension_edit.text(), "张力(g)")
        if val is None:
            return
        try:
            self._last_tension_setpoint = float(val)
        except Exception:
            pass
        self._send_motor_cmd(f"F {val}")

    def on_motor_pid(self):
        kp = self._parse_number_text(self.motor_kp_edit.text(), "Kp")
        if kp is None:
            return
        ki = self._parse_number_text(self.motor_ki_edit.text(), "Ki")
        if ki is None:
            return
        kd = self._parse_number_text(self.motor_kd_edit.text(), "Kd")
        if kd is None:
            return
        self._send_motor_cmd(f"pid {kp} {ki} {kd}")

    def on_motor_estop(self):
        if not self._motor_can_send():
            return
        # 顺序：设置模式 -> 张力 -> 设置模式 -> 速度 -> 禁用
        # 在发送 F/Con 命令前确保模式已设置。
        self._send_motor_cmd("ConMode 0")
        self._send_motor_cmd("F 0")
        self._send_motor_cmd("ConMode 1")
        self._send_motor_cmd("Con 0")
        self._send_motor_cmd("Disable")
        self.motor_mode = None
        self._set_motor_mode_lamps(None)
        self._set_lamp_color(self.motor_enable_lamp, "#777777")
        self._set_lamp_color(self.motor_dir_lamp, "#777777")

    def _set_serial_widgets_enabled(self, enabled: bool):
        for w in [
            self.port_combo, self.refresh_ports_btn, self.baud_combo,
            self.parity_combo, self.stopbits_combo, self.bytesize_combo, self.timeout_spin,
            self.rs485_mode_combo, self.pre_tx_spin, self.post_tx_spin,
            self.tx_port_combo, self.tx_baud_combo, self.enable_tx_chk, self.mon_rx_chk, self.mon_tx_chk
        ]:
            w.setEnabled(enabled)

    # ---------- 通道表格 ----------
    def add_channel_row(self, default_name: str = "", default_addr: int = 0, default_dtype: str = "float32"):
        row = self.ch_table.rowCount()
        self.ch_table.insertRow(row)

        enabled_chk = QCheckBox()
        enabled_chk.setChecked(True)
        enabled_chk.setStyleSheet("margin-left:12px;")
        self.ch_table.setCellWidget(row, 0, enabled_chk)

        self.ch_table.setItem(row, 1, QTableWidgetItem(default_name or f"CH{row+1}"))
        self.ch_table.setItem(row, 2, QTableWidgetItem(str(default_addr)))

        dtype_combo = QComboBox()
        dtype_combo.addItems(list(DTYPE_INFO.keys()))
        dtype_combo.setCurrentText(default_dtype if default_dtype in DTYPE_INFO else "float32")
        self.ch_table.setCellWidget(row, 3, dtype_combo)

        byte_combo = QComboBox()
        byte_combo.addItems(["big", "little"])
        byte_combo.setCurrentText("big")
        self.ch_table.setCellWidget(row, 4, byte_combo)

        word_combo = QComboBox()
        word_combo.addItems(["big", "little"])
        word_combo.setCurrentText("big")
        self.ch_table.setCellWidget(row, 5, word_combo)

        self.ch_table.setItem(row, 6, QTableWidgetItem("-0.01"))
        self.ch_table.setItem(row, 7, QTableWidgetItem("N"))

        self._refresh_friction_channel_options()
    def delete_selected_rows(self):
        rows = sorted({idx.row() for idx in self.ch_table.selectedIndexes()}, reverse=True)
        for r in rows:
            self.ch_table.removeRow(r)

        self._refresh_friction_channel_options()
    def gather_channels(self) -> List[ChannelConfig]:
        channels: List[ChannelConfig] = []
        seen_names = set()
        unit_map: Dict[str, str] = {}
        for r in range(self.ch_table.rowCount()):
            enabled_widget = self.ch_table.cellWidget(r, 0)
            enabled = bool(enabled_widget.isChecked()) if enabled_widget else True
            name = (self.ch_table.item(r, 1).text() if self.ch_table.item(r, 1) else "").strip() or f"CH{r+1}"
            if name in seen_names:
                i = 2
                base = name
                while f"{base}_{i}" in seen_names:
                    i += 1
                name = f"{base}_{i}"
            seen_names.add(name)

            try:
                address = int((self.ch_table.item(r, 2).text() if self.ch_table.item(r, 2) else "0").strip())
            except Exception:
                address = 0

            dtype_combo = self.ch_table.cellWidget(r, 3)
            dtype = dtype_combo.currentText() if dtype_combo else "float32"

            byte_combo = self.ch_table.cellWidget(r, 4)
            byte_order = byte_combo.currentText() if byte_combo else "big"

            word_combo = self.ch_table.cellWidget(r, 5)
            word_order = word_combo.currentText() if word_combo else "big"

            try:
                scale = float((self.ch_table.item(r, 6).text() if self.ch_table.item(r, 6) else "1.0").strip())
            except Exception:
                scale = 1.0
            try:
                unit = (self.ch_table.item(r, 7).text() if self.ch_table.item(r, 7) else "").strip()
            except Exception:
                unit = ""
            unit_map[name] = unit

            channels.append(ChannelConfig(enabled=enabled, name=name, address=address, dtype=dtype,
                                         byte_order=byte_order, word_order=word_order, scale=scale))
        self._last_unit_map = unit_map
        return channels


    # ---------- 绘图/数据 ----------
    def clear_data(self):
        # 重置环形缓冲区（大小跟随当前最大点数）
        try:
            size = int(self.max_points_spin.value())
        except Exception:
            size = int(getattr(self, '_buf_size', 100) or 100)

        self.channel_names.clear()
        self._alloc_ring_buffers(size, [], keep_last=False)

        # 重置绘图时间基准（相对秒）
        self._t0_mono_ts = None
        self._last_sample_rel_ts = None
        self._last_sample_mono_ts = None

        # 重置暂停补偿
        self._mono_pause_accum = 0.0
        self._mono_pause_start = None
        self._quality_gap_pending = []
        self._quality_gap_hold_mode = False
        self._quality_gap_start_mono = None
        self._quality_gap_triggered = False
        self._last_valid_row = None

        self.plot.clear()
        self.plot.addLegend()
        self.curves.clear()
        try:
            if hasattr(self, "friction_curve") and self.friction_curve is not None:
                self.friction_curve.setData([], [])
            if hasattr(self, "mu_curve") and self.mu_curve is not None:
                self.mu_curve.setData([], [])
        except Exception:
            pass
        self.set_status("已清空数据")

    def init_curves(self, channel_names: List[str]):
        self.plot.clear()
        self.plot.addLegend()
        self.curves.clear()

        # 1-7 通道配色：红 橙 黄 绿 青 蓝 紫（更高区分度）
        palette = [
            (220, 0, 0),      # 红
            (255, 140, 0),    # 橙
            (255, 200, 0),    # 黄（白底下略深）
            (0, 170, 0),      # 绿
            (0, 170, 170),    # 青
            (0, 0, 220),      # 蓝
            (140, 0, 200),    # 紫
        ]
        width = 2  # 线宽稍微粗一点

        for i, name in enumerate(channel_names):
            color = palette[i % len(palette)]
            pen = pg.mkPen(color=color, width=width)
            self.curves[name] = self.plot.plot([], [], name=name, pen=pen)
            item = self.curves.get(name)
            if item is not None:
                # 每条曲线的性能提示（跨版本安全）
                try:
                    item.setClipToView(True)
                except Exception:
                    pass
                try:
                    item.setDownsampling(auto=True, mode='peak')
                except Exception:
                    try:
                        item.setDownsampling(auto=True, method='peak')
                    except Exception:
                        pass
                # 部分版本支持跳过有限性检查以提速
                try:
                    item.setSkipFiniteCheck(True)
                except Exception:
                    pass



    @Slot(float, dict)
    def on_data_ready(self, ts: float, row: dict):
        mono_now = time.monotonic()
        try:
            wall_ts = float(ts) if ts is not None else time.time()
        except Exception:
            wall_ts = time.time()

        # 首帧懒初始化曲线与缓冲区（兼容未点击开始采集时的数据）
        if not self.channel_names:
            self.channel_names = list(row.keys())
            self.init_curves(self.channel_names)
            try:
                size = int(self.max_points_spin.value())
            except Exception:
                size = int(getattr(self, '_buf_size', 100) or 100)
            self._alloc_ring_buffers(size, list(self.channel_names), keep_last=False)

        # 保持缓冲区大小与 UI 同步
        try:
            want = int(self.max_points_spin.value())
        except Exception:
            want = int(getattr(self, '_buf_size', 0) or 0)
        if want and want != int(getattr(self, '_buf_size', 0) or 0):
            self._resize_ring_buffers(want)

        self._process_quality_sample(mono_now, wall_ts, row)


    def update_plot(self):
        """更新绘图曲线（缓冲 + 刷新率驱动）。"""
        count = int(getattr(self, "_buf_count", 0) or 0)
        if count <= 0 or not self.channel_names:
            return

        last_seq = int(getattr(self, "_last_plotted_seq", -1))
        cur_seq = int(getattr(self, "_plot_seq", 0) or 0)
        new_data = (last_seq != cur_seq)

        size = int(getattr(self, "_buf_size", 0) or 0)
        if size <= 0 or self._ts_buf is None:
            return
        idx = int(getattr(self, "_buf_idx", 0) or 0) % size
        full = (count >= size)
        # 平滑 X 滚动（实时）：用单调时间驱动右边界。
        # 当未采集（停止/暂停）时，冻结滚动且不要
        # 强制更新 XRange（便于用户平移/缩放最后一帧）。
        scroll_live = bool(getattr(self, 'is_acquiring', False)) and (not bool(getattr(self, 'is_paused', False)))

        try:
            if scroll_live and self._t0_mono_ts is not None:
                pause_accum = float(getattr(self, '_mono_pause_accum', 0.0) or 0.0)
                now_rel = float(time.monotonic() - float(self._t0_mono_ts) - pause_accum)
            else:
                now_rel = float(self._last_sample_rel_ts) if self._last_sample_rel_ts is not None else 0.0
        except Exception:
            now_rel = float(self._last_sample_rel_ts) if self._last_sample_rel_ts is not None else 0.0

        # 可见窗口宽度（秒）：max_points * poll_interval
        try:
            poll_ms = int(self.poll_spin.value()) if hasattr(self, 'poll_spin') else 20
        except Exception:
            poll_ms = 20
        poll_s = max(0.001, float(poll_ms) / 1000.0)
        npts = int(min(count, size))
        span = max(0.02, max(1, npts - 1) * poll_s)
        x_left = now_rel - span
        x_right = now_rel

        # 快速路径：没有新样本时避免重新上传曲线数据。
        # 只保持 X 平滑滚动，同时减少 CPU/GPU 开销。
        if not new_data:
            if scroll_live:
                try:
                    self.plot.setXRange(x_left, x_right, padding=0.0)
                except Exception:
                    pass
                try:
                    self.friction_plot.setXRange(x_left, x_right, padding=0.0)
                except Exception:
                    pass
                try:
                    self.mu_plot.setXRange(x_left, x_right, padding=0.0)
                except Exception:
                    pass
            return

        # 仅在上传新曲线数据时准备有序 X 视图。
        xs = None
        if new_data:
            if np is not None:
                if not full:
                    xs = self._ts_buf[:count]
                else:
                    first = size - idx
                    self._plot_x[:first] = self._ts_buf[idx:]
                    self._plot_x[first:] = self._ts_buf[:idx]
                    xs = self._plot_x
            else:
                if not full:
                    xs = list(self._ts_buf[:count])
                else:
                    xs = list(self._ts_buf[idx:]) + list(self._ts_buf[:idx])
                if not xs:
                    xs = None

        # 防止一次更新中多次重绘（对 Windows 有帮助）。
        self.plot.setUpdatesEnabled(False)
        try:
            global_y_min = None
            global_y_max = None

            if new_data and xs is not None and np is not None:
                for name in self.channel_names:
                    buf = self._val_buf_by_channel.get(name)
                    if buf is None:
                        continue

                    if not full:
                        ys = buf[:count]
                    else:
                        ytmp = self._plot_y_by_channel.get(name)
                        if ytmp is None or getattr(ytmp, 'shape', (0,))[0] != size:
                            ytmp = np.empty(size, dtype=float)
                            self._plot_y_by_channel[name] = ytmp
                        first = size - idx
                        ytmp[:first] = buf[idx:]
                        ytmp[first:] = buf[:idx]
                        ys = ytmp
                    ys_use = ys

                    curve = self.curves.get(name)
                    if curve is not None:
                        # 缓冲区很大时限制发送到渲染器的点数。
                        xs_use, ys_use = xs, ys
                        try:
                            max_disp = int(getattr(self, '_max_display_points', 0) or 0)
                        except Exception:
                            max_disp = 0
                        if max_disp:
                            try:
                                n = int(len(xs_use))
                            except Exception:
                                n = 0
                            if n > max_disp:
                                step = max(1, int(n // max_disp))
                                if step > 1:
                                    try:
                                        xs_use = xs_use[::step]
                                        ys_use = ys_use[::step]
                                    except Exception:
                                        # 兜底：不降采样
                                        xs_use, ys_use = xs, ys

                        # 支持时优先跳过有限性检查
                        try:
                            curve.setData(xs_use, ys_use, connect='finite', skipFiniteCheck=True)
                        except Exception:
                            curve.setData(xs_use, ys_use, connect='finite')

                    if self.autoscale_chk.isChecked():
                        finite = np.isfinite(ys_use)
                        if finite.any():
                            y_min = float(np.nanmin(ys_use))
                            y_max = float(np.nanmax(ys_use))
                            global_y_min = y_min if global_y_min is None else min(global_y_min, y_min)
                            global_y_max = y_max if global_y_max is None else max(global_y_max, y_max)
            elif new_data and xs is not None:
                # 兜底（无 numpy）
                for name in self.channel_names:
                    buf = self._val_buf_by_channel.get(name, [])
                    if not full:
                        ys_raw = list(buf[:count])
                    else:
                        ys_raw = list(buf[idx:]) + list(buf[:idx])

                    xs2, ys2 = [], []
                    for t, v in zip(xs, ys_raw):
                        if v is None:
                            continue
                        xs2.append(t)
                        ys2.append(v)

                    curve = self.curves.get(name)
                    if curve is not None:
                        curve.setData(xs2, ys2)

                    if self.autoscale_chk.isChecked() and ys2:
                        y_min, y_max = min(ys2), max(ys2)
                        global_y_min = y_min if global_y_min is None else min(global_y_min, y_min)
                        global_y_max = y_max if global_y_max is None else max(global_y_max, y_max)

            now = time.monotonic()
            # 平滑滚动：保持以“现在”为右边界的固定可见窗口。
            # 仅在采集中执行；停止/暂停后冻结并让
            # 用户检查/平移，不被计时器覆盖。
            if scroll_live:
                try:
                    self.plot.setXRange(x_left, x_right, padding=0.0)
                except Exception:
                    pass

            # Y 轴范围带滞回更新以减少抖动/闪烁
            if self.autoscale_chk.isChecked() and global_y_min is not None and global_y_max is not None:
                if (now - float(getattr(self, "_last_yrange_update", 0.0))) >= 1.0:
                    if global_y_min == global_y_max:
                        pad = 1.0 if global_y_min == 0 else abs(global_y_min) * 0.05
                        new_min = global_y_min - pad
                        new_max = global_y_max + pad
                    else:
                        span = global_y_max - global_y_min
                        pad = span * 0.08
                        new_min = global_y_min - pad
                        new_max = global_y_max + pad

                    apply = True
                    try:
                        cur_min, cur_max = self.plot.viewRange()[1]
                        cur_span = cur_max - cur_min
                        if cur_span > 0:
                            margin = cur_span * 0.05
                            # 如果新范围大多落在当前范围内，则跳过更新。
                            if (new_min >= (cur_min + margin)) and (new_max <= (cur_max - margin)):
                                apply = False
                    except Exception:
                        pass

                    if apply:
                        self.plot.setYRange(new_min, new_max, padding=0.0)
                        self._last_yrange_update = now
        finally:
            self.plot.setUpdatesEnabled(True)
            # 更新所有曲线后只请求一次重绘。
            self.plot.update()

        try:
            self._update_friction_plots(xs, idx, full, count, scroll_live, x_left, x_right)
        except Exception:
            pass

        self._last_plotted_seq = int(getattr(self, "_plot_seq", 0) or 0)


    # ---------- 连接/采集 ----------
    def toggle_connect(self):
        if self.is_connected:
            self.disconnect_serial()
        else:
            self.connect_serial()

    def connect_serial(self):
        port = self.port_combo.currentData()
        if not port:
            QMessageBox.warning(self, "提示", "未选择有效串口。请刷新串口后选择。")
            return

        try:
            baud = int(self.baud_combo.currentText())
        except Exception:
            baud = 9600

        parity = self.parity_combo.currentText()
        stopbits = int(self.stopbits_combo.currentText())
        bytesize = int(self.bytesize_combo.currentText())
        timeout_s = float(self.timeout_spin.value())

        tx_enabled = bool(self.enable_tx_chk.isChecked())
        tx_port = self.tx_port_combo.currentData() if tx_enabled else ""
        try:
            tx_baud = int(self.tx_baud_combo.currentText())
        except Exception:
            tx_baud = 115200

        tx_interval_ms = int(self.tx_interval_spin.value()) if hasattr(self, 'tx_interval_spin') else 20
        if tx_enabled and not tx_port:
            QMessageBox.warning(self, "提示", "已启用发送串口，但未选择发送串口。")
            self._set_serial_widgets_enabled(True)
            self.connect_btn.setEnabled(True)
            return

        rs485_cfg = Rs485CtrlConfig(
            mode=self.rs485_mode_combo.currentText(),
            tx_level_high=True,
            rx_level_high=False,
            pre_tx_ms=int(self.pre_tx_spin.value()),
            post_tx_ms=int(self.post_tx_spin.value()),
        )

        self.worker = ModbusRtuWorker(
            port=port, baudrate=baud, parity=parity, stopbits=stopbits,
            bytesize=bytesize, timeout_s=timeout_s, rs485_cfg=rs485_cfg,
            tx_enabled=tx_enabled, tx_port=tx_port, tx_baudrate=tx_baud, tx_interval_ms=tx_interval_ms
        )
        self.worker.data_ready.connect(self.on_data_ready)
        self.worker.status.connect(self.set_status)
        self.worker.connected.connect(self.on_connected_state)
        self.worker.acquiring.connect(self.on_acquiring_state)
        self.worker.log_line.connect(self.append_monitor)
        self.worker.frame.connect(self.on_frame)
        try:
            self.worker.set_tx_tap_enabled(bool(self.mon_tx_chk.isChecked()))
        except Exception:
            pass
        self.worker.start()

        self.set_status("正在连接...")
        self._set_serial_widgets_enabled(False)
        self.connect_btn.setEnabled(False)

    def disconnect_serial(self):
        if self.worker:
            try:
                self.worker.set_acquiring(False)
                self.worker.stop_thread()
                self.worker.wait(1500)
            except Exception:
                pass
            self.worker = None
        try:
            self._stop_data_logger()
        except Exception:
            pass

        self.is_connected = False
        self.is_acquiring = False
        self.is_paused = False
        self.connect_btn.setText("连接串口")
        self.connect_btn.setEnabled(True)
        self.acquire_btn.setText("开始采集")
        self.acquire_btn.setEnabled(False)
        self.pause_btn.setText("暂停")
        self.pause_btn.setEnabled(False)
        self._set_serial_widgets_enabled(True)
        self.set_status("已断开连接")
        self.update_custom_send_ports()
        try:
            self._update_plot_timer_running()
        except Exception:
            pass

    @Slot(bool)
    def on_connected_state(self, ok: bool):
        self.is_connected = bool(ok)
        if ok:
            self.connect_btn.setText("断开串口")
            self.connect_btn.setEnabled(True)
            self.acquire_btn.setEnabled(True)
            self.pause_btn.setEnabled(False)
            self.pause_btn.setText("暂停")
            self.is_paused = False
            self.set_status("已连接（未采集）")
            self.update_custom_send_ports()
        else:
            self.disconnect_serial()
            return

        try:
            self._update_plot_timer_running()
        except Exception:
            pass

    def toggle_acquire(self):
        if not self.is_connected or not self.worker:
            return
        if self.is_acquiring:
            self.stop_acquire()
        else:
            self.start_acquire()

    def start_acquire(self):
        if not self.worker:
            return

        unit_id = int(self.unit_spin.value())
        func_code = int(self.func_combo.currentData())
        poll_ms = int(self.poll_spin.value())
        address_base_1 = bool(self.addr_base1_chk.isChecked())

        channels = self.gather_channels()
        enabled_channels = [c for c in channels if c.enabled]
        if not enabled_channels:
            QMessageBox.warning(self, "提示", "至少启用一个通道。")
            return

        self.worker.update_runtime(unit_id, func_code, poll_ms, address_base_1, enabled_channels, tx_interval_ms=int(self.tx_interval_spin.value()))

        # 开始时重置绘图数据
        self.clear_data()
        # 重置时间轴的暂停补偿
        self._mono_pause_accum = 0.0
        self._mono_pause_start = None
        self._quality_gap_pending = []
        self._quality_gap_hold_mode = False
        self._quality_gap_start_mono = None
        self._quality_gap_triggered = False
        self._last_valid_row = None
        self.channel_names = [c.name for c in enabled_channels]
        self._log_units = [self._last_unit_map.get(c.name, "") for c in enabled_channels]
        self.init_curves(self.channel_names)
        # 按当前最大点数分配环形缓冲区
        try:
            size = int(self.max_points_spin.value())
        except Exception:
            size = int(getattr(self, '_buf_size', 100) or 100)
        self._alloc_ring_buffers(size, list(self.channel_names), keep_last=False)
        self._start_data_logger(self.channel_names, self._log_units)

        self.worker.set_acquiring(True)
        self.is_acquiring = True
        self.is_paused = False
        self.pause_btn.setText("暂停")
        self.pause_btn.setEnabled(True)
        self.acquire_btn.setText("停止采集")
        self.set_status("采集中...（请看通讯监视窗口 TX/RX）")
        try:
            self._update_plot_timer_running()
        except Exception:
            pass

    def stop_acquire(self):
        if self.worker:
            self.worker.set_acquiring(False)
        self.is_acquiring = False
        self.is_paused = False
        self._mono_pause_start = None
        self._quality_gap_pending = []
        self._quality_gap_hold_mode = False
        self._quality_gap_start_mono = None
        self._quality_gap_triggered = False
        self._last_valid_row = None
        self.pause_btn.setText("暂停")
        self.pause_btn.setEnabled(False)
        self.acquire_btn.setText("开始采集")
        self.set_status("已连接（未采集）")
        self._stop_data_logger()
        try:
            self._update_plot_timer_running()
        except Exception:
            pass

    def toggle_pause(self):
        """
        暂停/继续采集：
        - 暂停：停止发送查询（保持连接），不清空数据，不绘图追加
        - 继续：恢复发送查询，继续在原数据后追加
        """
        if not self.is_connected or not self.worker:
            return
        # 只有在“已开始采集（acquire_btn 处于停止采集状态）”时才允许暂停
        if self.acquire_btn.text() != "停止采集":
            return

        if not self.is_paused:
            # 暂停
            self.worker.set_acquiring(False)
            self.is_acquiring = False
            self.is_paused = True
            self._mono_pause_start = time.monotonic()
            self.pause_btn.setText("继续")
            self.set_status("已暂停（保持连接）")
        else:
            # 继续
            try:
                if getattr(self, '_mono_pause_start', None) is not None:
                    self._mono_pause_accum = float(getattr(self, '_mono_pause_accum', 0.0) or 0.0) + max(0.0, float(time.monotonic() - float(self._mono_pause_start)))
            except Exception:
                pass
            self._mono_pause_start = None
            self.worker.set_acquiring(True)
            self.is_acquiring = True
            self.is_paused = False
            self.pause_btn.setText("暂停")
            self.set_status("采集中...（请看通讯监视窗口 TX/RX）")

        try:
            self._update_plot_timer_running()
        except Exception:
            pass


    @Slot(bool)
    def on_acquiring_state(self, on: bool):
        self.is_acquiring = bool(on)


    # ---------- 数据记录 ----------
    def _start_data_logger(self, channel_names: List[str], channel_units: Optional[List[str]] = None):
        try:
            if not channel_names:
                return
            log_names = list(channel_names)
            log_units = list(channel_units or [])
            if self._quality_flag_name not in log_names:
                log_names.append(self._quality_flag_name)
                log_units.append("")
            path = self._data_logger.start_session(log_names, log_units)
            self._log_db_path = path
            self._log_channels = list(log_names)
            self._log_units = list(log_units)
        except Exception:
            self._log_db_path = ""
            self._log_channels = []
            self._log_units = []

    def _stop_data_logger(self):
        try:
            if self._data_logger:
                self._data_logger.stop()
        except Exception:
            pass

    def _db_has_data(self, path: str) -> bool:
        if not path:
            return False
        if not os.path.isfile(path):
            return False
        try:
            conn = sqlite3.connect(path)
            try:
                cur = conn.execute("SELECT 1 FROM data LIMIT 1")
                return cur.fetchone() is not None
            finally:
                conn.close()
        except Exception:
            return False

    def _format_export_time(self, wall_ts, rel_ts) -> str:
        try:
            if wall_ts is not None and math.isfinite(float(wall_ts)):
                return time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(float(wall_ts)))
        except Exception:
            pass
        try:
            return f"{float(rel_ts):.3f}s"
        except Exception:
            return ""

    def _unit_label(self, unit: str) -> str:
        u = (unit or "").strip()
        if not u:
            return ""
        ul = u.lower()
        if ul == "g":
            return "g【克】"
        if ul == "n":
            return "N【牛】"
        if u in ("无量纲", "-"):
            return "无量纲"
        return f"{u}【单位】"

    def _split_sheet_name(self, base_title: str, idx: int) -> str:
        base = self._safe_sheet_name(base_title)
        name = f"{base}_{idx}" if idx > 1 else base
        if len(name) > 31:
            name = name[:31]
        return name

    def _create_sheet_with_header_cells(self, wb, base_title: str, idx: int, headers):
        ws = wb.create_sheet(title=self._split_sheet_name(base_title, idx))
        for col, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=h)
        return ws

    def _create_sheet_with_header_append(self, wb, base_title: str, idx: int, headers):
        ws = wb.create_sheet(self._split_sheet_name(base_title, idx))
        ws.append(headers)
        return ws

    def _export_xlsx_from_ring(self, path: str, xs, ys_map, xs_wall, qf_vals=None):
        wb = Workbook()
        max_rows = 1048576

        headers = ["Time"]
        units = list(getattr(self, "_log_units", []))
        for idx, name in enumerate(self.channel_names):
            unit = units[idx] if idx < len(units) else ""
            unit_label = self._unit_label(unit)
            headers.append(f"{name}({unit_label})" if unit_label else name)
        headers += ["摩擦力(N【牛】)", "摩擦系数", self._quality_flag_label]

        ws_all_idx = 1
        ws_all = wb.active
        ws_all.title = self._split_sheet_name("All", ws_all_idx)
        for i, h in enumerate(headers, start=1):
            ws_all.cell(row=1, column=i, value=h)
        ws_all_row = 2

        nrows = len(xs)
        hi_name = (getattr(self, "_fric_high_name", "") or "").strip()
        lo_name = (getattr(self, "_fric_low_name", "") or "").strip()

        for i, rel_ts in enumerate(xs):
            if ws_all_row > max_rows:
                ws_all_idx += 1
                ws_all = self._create_sheet_with_header_cells(wb, "All", ws_all_idx, headers)
                ws_all_row = 2
            wall_ts = xs_wall[i] if xs_wall and i < len(xs_wall) else None
            t_str = self._format_export_time(wall_ts, rel_ts)
            ws_all.cell(row=ws_all_row, column=1, value=t_str)

            row_vals = []
            col_idx = 2
            for c, name in enumerate(self.channel_names, start=2):
                col = ys_map.get(name, [])
                v = col[i] if i < len(col) else None
                row_vals.append(v)
                ws_all.cell(row=ws_all_row, column=col_idx, value=v)
                col_idx += 1

            high_v = None
            low_v = None
            if hi_name and hi_name in self.channel_names:
                try:
                    high_v = row_vals[self.channel_names.index(hi_name)]
                except Exception:
                    high_v = None
            if lo_name and lo_name in self.channel_names:
                try:
                    low_v = row_vals[self.channel_names.index(lo_name)]
                except Exception:
                    low_v = None
            fric_n, mu = self._calc_fric_mu(high_v, low_v)
            ws_all.cell(row=ws_all_row, column=col_idx, value=fric_n)
            ws_all.cell(row=ws_all_row, column=col_idx + 1, value=mu)
            col_idx += 2

            qf_v = qf_vals[i] if (qf_vals is not None and i < len(qf_vals)) else None
            ws_all.cell(row=ws_all_row, column=col_idx, value=qf_v)

            ws_all_row += 1

        for i, name in enumerate(self.channel_names):
            unit = units[i] if i < len(units) else ""
            unit_label = self._unit_label(unit)
            header = ["Time", f"{name}({unit_label})" if unit_label else name]
            ws_idx = 1
            ws = self._create_sheet_with_header_cells(wb, name, ws_idx, header)
            row_idx = 2
            vals = ys_map.get(name, [])
            for r in range(nrows):
                if row_idx > max_rows:
                    ws_idx += 1
                    ws = self._create_sheet_with_header_cells(wb, name, ws_idx, header)
                    row_idx = 2
                wall_ts = xs_wall[r] if xs_wall and r < len(xs_wall) else None
                t_str = self._format_export_time(wall_ts, xs[r])
                ws.cell(row=row_idx, column=1, value=t_str)
                v = vals[r] if r < len(vals) else None
                ws.cell(row=row_idx, column=2, value=v)
                row_idx += 1

        if qf_vals is not None:
            q_header = ["Time", self._quality_flag_label]
            ws_q_idx = 1
            ws_q = self._create_sheet_with_header_cells(wb, self._quality_flag_name, ws_q_idx, q_header)
            row_idx = 2
            for r in range(nrows):
                if row_idx > max_rows:
                    ws_q_idx += 1
                    ws_q = self._create_sheet_with_header_cells(wb, self._quality_flag_name, ws_q_idx, q_header)
                    row_idx = 2
                wall_ts = xs_wall[r] if xs_wall and r < len(xs_wall) else None
                t_str = self._format_export_time(wall_ts, xs[r])
                ws_q.cell(row=row_idx, column=1, value=t_str)
                qv = qf_vals[r] if r < len(qf_vals) else None
                ws_q.cell(row=row_idx, column=2, value=qv)
                row_idx += 1

        f_header = ["Time", "摩擦力(N【牛】)", self._quality_flag_label]
        mu_header = ["Time", "摩擦系数", self._quality_flag_label]
        ws_f_idx = 1
        ws_mu_idx = 1
        ws_f = self._create_sheet_with_header_cells(wb, "摩擦力", ws_f_idx, f_header)
        ws_mu = self._create_sheet_with_header_cells(wb, "摩擦系数", ws_mu_idx, mu_header)
        ws_f_row = 2
        ws_mu_row = 2
        for i, rel_ts in enumerate(xs):
            if ws_f_row > max_rows:
                ws_f_idx += 1
                ws_f = self._create_sheet_with_header_cells(wb, "摩擦力", ws_f_idx, f_header)
                ws_f_row = 2
            if ws_mu_row > max_rows:
                ws_mu_idx += 1
                ws_mu = self._create_sheet_with_header_cells(wb, "摩擦系数", ws_mu_idx, mu_header)
                ws_mu_row = 2
            wall_ts = xs_wall[i] if xs_wall and i < len(xs_wall) else None
            t_str = self._format_export_time(wall_ts, rel_ts)
            fric_n = None
            mu = None
            if hi_name and lo_name and hi_name in ys_map and lo_name in ys_map:
                try:
                    hv = ys_map.get(hi_name, [])[i]
                    lv = ys_map.get(lo_name, [])[i]
                except Exception:
                    hv = None
                    lv = None
                fric_n, mu = self._calc_fric_mu(hv, lv)
            ws_f.cell(row=ws_f_row, column=1, value=t_str)
            ws_f.cell(row=ws_f_row, column=2, value=fric_n)
            qf_v = qf_vals[i] if (qf_vals is not None and i < len(qf_vals)) else None
            ws_f.cell(row=ws_f_row, column=3, value=qf_v)
            ws_mu.cell(row=ws_mu_row, column=1, value=t_str)
            ws_mu.cell(row=ws_mu_row, column=2, value=mu)
            ws_mu.cell(row=ws_mu_row, column=3, value=qf_v)
            ws_f_row += 1
            ws_mu_row += 1

        for ws in wb.worksheets:
            self._autosize_sheet(ws)

        wb.save(path)

    def _export_xlsx_from_db(self, db_path: str, path: str, progress_cb=None, progress_ctx=None, phase_cb=None):
        try:
            if self._data_logger:
                self._data_logger.flush(wait=True, timeout=3.0)
        except Exception:
            pass

        conn = sqlite3.connect(db_path)
        total_rows = None
        if progress_cb:
            try:
                total_rows = conn.execute("SELECT COUNT(*) FROM data").fetchone()[0]
                if total_rows is None:
                    total_rows = 0
            except Exception:
                total_rows = None
            try:
                if total_rows is not None:
                    total_rows = max(1, int(total_rows))
                    progress_cb(progress_ctx, 0, total_rows)
            except Exception:
                pass
        try:
            try:
                cur = conn.execute("SELECT idx, name, unit FROM channels ORDER BY idx")
                channel_rows = cur.fetchall()
                channel_names = [row[1] for row in channel_rows]
                channel_units = [row[2] if len(row) > 2 else "" for row in channel_rows]
            except Exception:
                cur = conn.execute("SELECT idx, name FROM channels ORDER BY idx")
                channel_rows = cur.fetchall()
                channel_names = [row[1] for row in channel_rows]
                channel_units = ["" for _ in channel_rows]
            n_ch = len(channel_names)
            col_names = [f"ch{i}" for i in range(n_ch)]
            cols_sql = ", ".join(["ts"] + col_names) if col_names else "ts"
            query = f"SELECT {cols_sql} FROM data ORDER BY id"

            max_rows = 1048576
            wb = Workbook(write_only=True)

            headers = ["Time"]
            for idx, name in enumerate(channel_names):
                if name == self._quality_flag_name:
                    continue
                unit = channel_units[idx] if idx < len(channel_units) else ""
                unit_label = self._unit_label(unit)
                headers.append(f"{name}({unit_label})" if unit_label else name)
            headers += ["摩擦力(N【牛】)", "摩擦系数", self._quality_flag_label]

            ws_all_idx = 1
            ws_all = self._create_sheet_with_header_append(wb, "All", ws_all_idx, headers)
            ws_all_rows = 1
            f_header = ["Time", "摩擦力(N【牛】)", self._quality_flag_label]
            mu_header = ["Time", "摩擦系数", self._quality_flag_label]
            ws_f_idx = 1
            ws_mu_idx = 1
            ws_f = self._create_sheet_with_header_append(wb, "摩擦力", ws_f_idx, f_header)
            ws_mu = self._create_sheet_with_header_append(wb, "摩擦系数", ws_mu_idx, mu_header)
            ws_f_rows = 1
            ws_mu_rows = 1

            hi_name = (getattr(self, "_fric_high_name", "") or "").strip()
            lo_name = (getattr(self, "_fric_low_name", "") or "").strip()
            name_to_idx = {name: i for i, name in enumerate(channel_names)}
            hi_idx = name_to_idx.get(hi_name, None)
            lo_idx = name_to_idx.get(lo_name, None)

            cur = conn.execute(query)
            done_rows = 0
            while True:
                rows = cur.fetchmany(1000)
                if not rows:
                    break
                for row in rows:
                    ts_val = row[0]
                    vals = list(row[1:]) if n_ch > 0 else []
                    t_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(float(ts_val)))
                    high_v = vals[hi_idx] if (hi_idx is not None and hi_idx < len(vals)) else None
                    low_v = vals[lo_idx] if (lo_idx is not None and lo_idx < len(vals)) else None
                    row_out = [t_str]
                    qf_val = None
                    for idx, name in enumerate(channel_names):
                        v = vals[idx] if idx < len(vals) else None
                        if name == self._quality_flag_name:
                            qf_val = v
                            continue
                        row_out.append(v)

                    fric_n, mu = self._calc_fric_mu(high_v, low_v)
                    row_out += [fric_n, mu, qf_val]

                    if ws_all_rows >= max_rows:
                        ws_all_idx += 1
                        ws_all = self._create_sheet_with_header_append(wb, "All", ws_all_idx, headers)
                        ws_all_rows = 1
                    ws_all.append(row_out)
                    ws_all_rows += 1

                    if ws_f_rows >= max_rows:
                        ws_f_idx += 1
                        ws_f = self._create_sheet_with_header_append(wb, "摩擦力", ws_f_idx, f_header)
                        ws_f_rows = 1
                    ws_f.append([t_str, fric_n, qf_val])
                    ws_f_rows += 1

                    if ws_mu_rows >= max_rows:
                        ws_mu_idx += 1
                        ws_mu = self._create_sheet_with_header_append(wb, "摩擦系数", ws_mu_idx, mu_header)
                        ws_mu_rows = 1
                    ws_mu.append([t_str, mu, qf_val])
                    ws_mu_rows += 1

                done_rows += len(rows)
                if progress_cb and total_rows is not None:
                    try:
                        progress_cb(progress_ctx, done_rows, total_rows)
                    except Exception:
                        pass
            wb.save(path)
        finally:
            conn.close()
        try:
            if phase_cb:
                phase_cb("收尾中")
        except Exception:
            pass
        try:
            self._autosize_workbook(path)
        except Exception:
            pass

    def export_history_db(self):
        data_dir = os.path.join(os.getcwd(), "data_logs")
        if not os.path.isdir(data_dir):
            QMessageBox.information(self, "提示", "未找到 data_logs 目录。请先采集数据生成历史数据库。")
            return

        db_path, _ = QFileDialog.getOpenFileName(
            self,
            "选择历史数据库",
            data_dir,
            "SQLite DB (*.sqlite *.db);;All Files (*.*)"
        )
        if not db_path:
            return
        if not os.path.isfile(db_path):
            QMessageBox.warning(self, "提示", "数据库文件不存在。")
            return

        self.queue_export_db_paths([db_path])

    def open_history_dialog(self):
        data_dir = os.path.join(os.getcwd(), "data_logs")
        if not hasattr(self, "_history_db_dialog") or self._history_db_dialog is None:
            self._history_db_dialog = HistoryDbDialog(self, data_dir, self._export_history_db_path)
        else:
            try:
                self._history_db_dialog._data_dir = data_dir
                self._history_db_dialog.reload()
            except Exception:
                pass
        dlg = self._history_db_dialog
        dlg.show()
        try:
            dlg.raise_()
            dlg.activateWindow()
        except Exception:
            pass

    def _get_export_queue_dialog(self):
        if not hasattr(self, "_export_queue_dialog") or self._export_queue_dialog is None:
            self._export_queue_dialog = ExportQueueDialog(self, self._export_xlsx_from_db)
        return self._export_queue_dialog

    def open_export_queue_dialog(self):
        dlg = self._get_export_queue_dialog()
        dlg.show()
        try:
            dlg.raise_()
            dlg.activateWindow()
        except Exception:
            pass

    def queue_export_db_paths(self, db_paths: List[str]):
        if not db_paths:
            return

        if len(db_paths) == 1:
            db_path = db_paths[0]
            base = os.path.splitext(os.path.basename(db_path))[0]
            default_dir = os.path.dirname(db_path) or os.path.join(os.getcwd(), "data_logs")
            out_path, _ = QFileDialog.getSaveFileName(
                self,
                "导出为 XLSX",
                os.path.join(default_dir, f"{base}.xlsx"),
                "Excel Files (*.xlsx)"
            )
            if not out_path:
                return
            if not out_path.lower().endswith(".xlsx"):
                out_path += ".xlsx"
            tasks = [{"db_path": db_path, "out_path": out_path}]
            dlg = self._get_export_queue_dialog()
            dlg.show()
            try:
                dlg.raise_()
                dlg.activateWindow()
            except Exception:
                pass
            dlg.enqueue_exports(tasks)
            return

        default_dir = os.path.dirname(db_paths[0]) or os.path.join(os.getcwd(), "data_logs")
        zip_path, _ = QFileDialog.getSaveFileName(
            self,
            "导出压缩包",
            os.path.join(default_dir, "history_databases.zip"),
            "Zip Files (*.zip)"
        )
        if not zip_path:
            return
        if not zip_path.lower().endswith(".zip"):
            zip_path += ".zip"
        tasks = [{"db_path": p, "out_path": ""} for p in db_paths]
        dlg = self._get_export_queue_dialog()
        dlg.show()
        try:
            dlg.raise_()
            dlg.activateWindow()
        except Exception:
            pass
        dlg.enqueue_exports(tasks, zip_path=zip_path)

    def _build_history_menu(self):
        if not hasattr(self, "hist_menu") or self.hist_menu is None:
            return
        self.hist_menu.clear()

        act_refresh = self.hist_menu.addAction("刷新历史数据库")
        act_refresh.triggered.connect(self._build_history_menu)

        act_pick = self.hist_menu.addAction("管理数据库")
        act_pick.triggered.connect(self.open_history_dialog)
        act_queue = self.hist_menu.addAction("导出队列")
        act_queue.triggered.connect(self.open_export_queue_dialog)


        self.hist_menu.addSeparator()

        data_dir = os.path.join(os.getcwd(), "data_logs")
        if not os.path.isdir(data_dir):
            act_empty = self.hist_menu.addAction("(未找到 data_logs 目录)")
            act_empty.setEnabled(False)
            return

        db_files = []
        try:
            for name in os.listdir(data_dir):
                if not name.lower().endswith((".sqlite", ".db")):
                    continue
                full = os.path.join(data_dir, name)
                if os.path.isfile(full):
                    try:
                        mtime = os.path.getmtime(full)
                    except Exception:
                        mtime = 0.0
                    db_files.append((mtime, full))
        except Exception:
            db_files = []

        if not db_files:
            act_none = self.hist_menu.addAction("(无历史数据库)")
            act_none.setEnabled(False)
            return

        db_files.sort(key=lambda x: x[0], reverse=True)
        max_items = 10
        for mtime, full in db_files[:max_items]:
            base = os.path.splitext(os.path.basename(full))[0]
            try:
                ts_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(float(mtime)))
            except Exception:
                ts_str = "未知时间"
            label = f"{ts_str}  |  {base}"
            act = self.hist_menu.addAction(label)
            act.triggered.connect(lambda _=False, p=full: self._export_history_db_path(p))

    def _export_history_db_path(self, db_path: str):
        if not db_path or not os.path.isfile(db_path):
            QMessageBox.warning(self, "提示", "数据库文件不存在。")
            return
        self.queue_export_db_paths([db_path])

    def save_xlsx(self):
        db_path = self._log_db_path if getattr(self, "_log_db_path", "") else ""
        use_db = self._db_has_data(db_path)

        xs = []
        ys_map = {}
        xs_wall = []
        if not use_db:
            xs, ys_map, xs_wall, qf_vals = self._snapshot_ring(include_wall=True, include_quality=True)
            if not xs or not self.channel_names:
                QMessageBox.information(self, "提示", "当前没有可保存的数据。请先开始采集。")
                return

        path, _ = QFileDialog.getSaveFileName(self, "保存为 XLSX", "modbus_data.xlsx", "Excel Files (*.xlsx)")
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        try:
            if use_db:
                dlg = self._get_export_queue_dialog()
                dlg.show()
                try:
                    dlg.raise_()
                    dlg.activateWindow()
                except Exception:
                    pass
                tasks = [{"db_path": db_path, "out_path": path}]
                dlg.enqueue_exports(tasks)
            else:
                self._export_xlsx_from_ring(path, xs, ys_map, xs_wall, qf_vals=qf_vals)
                dlg = self._get_export_queue_dialog()
                dlg.show()
                try:
                    dlg.raise_()
                    dlg.activateWindow()
                except Exception:
                    pass
                dlg.show_completed_task(path)
            self.set_status(f"已保存：{path}")
        except Exception as e:
            QMessageBox.critical(self, "保存失败", f"保存 xlsx 失败：\n{e}")

    @staticmethod
    def _safe_sheet_name(name: str) -> str:
        bad = ['\\', '/', '*', '[', ']', ':', '?']
        s = "".join("_" if ch in bad else ch for ch in name).strip() or "Sheet"
        return s[:31]

    @staticmethod
    def _autosize_sheet(ws):
        max_rows = min(ws.max_row, 2)
        for col in range(1, ws.max_column + 1):
            max_len = 0
            for row in range(1, max_rows + 1):
                v = ws.cell(row=row, column=col).value
                if v is None:
                    continue
                s = str(v)
                # 将 CJK 宽字符按宽度 2 计算，便于 Excel 列宽调整
                width = 0
                for ch in s:
                    width += 2 if unicodedata.east_asian_width(ch) in ("W", "F") else 1
                max_len = max(max_len, width)
            ws.column_dimensions[get_column_letter(col)].width = min(max(10, max_len + 2), 50)

    def _autosize_workbook(self, path: str):
        wb = load_workbook(path)
        try:
            for ws in wb.worksheets:
                self._autosize_sheet(ws)
            wb.save(path)
        finally:
            try:
                wb.close()
            except Exception:
                pass

    def closeEvent(self, event):
        try:
            export_dialog = getattr(self, "_export_queue_dialog", None)
            if export_dialog and export_dialog.is_exporting():
                ret = QMessageBox.question(
                    self,
                    "确认退出",
                    "检测到正在导出的项目。\n确认退出将强制结束所有软件相关进程，是否退出？",
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No,
                )
                if ret != QMessageBox.Yes:
                    event.ignore()
                    return
                try:
                    export_dialog.force_stop_exports()
                except Exception:
                    pass
            elif export_dialog:
                try:
                    export_dialog.force_stop_exports()
                except Exception:
                    pass
            try:
                import subprocess
                pid = os.getpid()
                subprocess.Popen(["taskkill", "/PID", str(pid), "/T", "/F"])
            except Exception:
                pass
        except Exception:
            pass
        # 持久化工作区布局（停靠位置/分割器尺寸/窗口几何）
        try:
            self._save_window_layout()
        except Exception:
            pass
        # 确保退出前串口线程已停止
        try:
            self.disconnect_serial()
        except Exception:
            pass
        super().closeEvent(event)


