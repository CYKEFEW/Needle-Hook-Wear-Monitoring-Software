#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Modbus RTU 串口上位机助手（浅色主题 / 多通道绘图 / XLSX 导出 / 通讯监视窗口）

本版（v3）重点解决：
1) RS485 模块 TX 灯不亮：很多“TTL<->RS485模块/转接板”需要 DE/RE 方向控制，
   常见做法是用 USB-UART 的 RTS 或 DTR 引脚去控制 DE/RE。若不控制，主机 write() 可能在TTL侧发送了，
   但收发器一直处于接收态，485总线上无电平翻转，TX灯自然不亮。
   -> 本程序增加“RS485方向控制”选项，默认使用 RTS（TX=高，RX=低），并支持可调前后延时。
2) 增加“通讯监视窗口”：显示每次发送(TX)与接收(RX)的十六进制帧，含超时/CRC错误/异常响应提示。
3) 连接串口/开始采集仍然分按钮；默认波特率 9600。
4) 采集中若本次轮询所有通道均无有效数据（全None），不追加点、不绘图。

依赖：
  pip install PySide6 pyqtgraph pyserial openpyxl
运行：
  python modbus_serial_assistant_light_v8_dualserial.py
"""

import sys
import time
import math
import struct
import threading
import queue
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

# ---- Qt binding (PySide6 preferred; fallback to PyQt5) ----
QT_LIB = "PySide6"
try:
    from PySide6.QtCore import Qt, QThread, Signal, Slot, QTimer, QSettings, QPoint
    from PySide6.QtGui import QTextCursor, QGuiApplication
    from PySide6.QtWidgets import (
        QApplication, QMainWindow, QWidget, QLabel, QComboBox, QPushButton,
        QLineEdit,
        QSpinBox, QDoubleSpinBox, QCheckBox, QHBoxLayout, QVBoxLayout,
        QGridLayout, QGroupBox, QTableWidget, QTableWidgetItem, QMessageBox,
        QFileDialog, QHeaderView, QDockWidget, QTabWidget, QTextEdit, QPlainTextEdit, QSplitter, QSizePolicy
    )
except Exception:
    QT_LIB = "PyQt5"
    from PyQt5.QtCore import Qt, QThread, QTimer, QSettings, QPoint, pyqtSignal as Signal, pyqtSlot as Slot
    from PyQt5.QtGui import QTextCursor, QGuiApplication
    from PyQt5.QtWidgets import (
        QApplication, QMainWindow, QWidget, QLabel, QComboBox, QPushButton,
        QLineEdit,
        QSpinBox, QDoubleSpinBox, QCheckBox, QHBoxLayout, QVBoxLayout,
        QGridLayout, QGroupBox, QTableWidget, QTableWidgetItem, QMessageBox,
        QFileDialog, QHeaderView, QDockWidget, QTabWidget, QTextEdit, QPlainTextEdit, QSplitter, QSizePolicy
    )

import pyqtgraph as pg
from pyqtgraph.graphicsItems.DateAxisItem import DateAxisItem

import serial
from serial.tools import list_ports

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# Optional: numpy speeds up plotting at high rates
try:
    import numpy as np
except Exception:  # pragma: no cover
    np = None


# ---------------- Virtual Serial (Simulation) ----------------

class _VirtualEndpoint:
    """A minimal in-memory byte stream endpoint with blocking read."""

    def __init__(self):
        self._buf = bytearray()
        self._lock = threading.Lock()
        self._cv = threading.Condition(self._lock)

    def feed(self, data: bytes):
        if not data:
            return
        with self._cv:
            self._buf.extend(data)
            self._cv.notify_all()

    def read(self, n: int, timeout: Optional[float]) -> bytes:
        if n <= 0:
            return b""
        t_end = None if timeout is None else (time.time() + float(timeout))
        with self._cv:
            while not self._buf:
                if timeout == 0:
                    return b""
                if t_end is None:
                    self._cv.wait()
                else:
                    remaining = t_end - time.time()
                    if remaining <= 0:
                        return b""
                    self._cv.wait(timeout=remaining)
            out = self._buf[:n]
            del self._buf[:len(out)]
            return bytes(out)

    @property
    def in_waiting(self) -> int:
        with self._lock:
            return len(self._buf)

    def clear(self):
        with self._lock:
            self._buf.clear()


class VirtualSerial:
    """A tiny subset of pyserial.Serial API for in-app simulation ports."""

    def __init__(self, port: str, rx_ep: _VirtualEndpoint, peer_rx_ep: _VirtualEndpoint):
        self.port = port
        self._rx = rx_ep
        self._peer_rx = peer_rx_ep
        self.timeout: Optional[float] = 0.1
        self.write_timeout: Optional[float] = 0.1
        self.rts = False
        self.dtr = False
        self.is_open = True

    def close(self):
        self.is_open = False


    def open(self):
        """Re-open the virtual port after it was closed."""
        self.is_open = True

    def flush(self):
        # in-memory stream: nothing to flush
        return

    def reset_input_buffer(self):
        self._rx.clear()

    @property
    def in_waiting(self) -> int:
        return self._rx.in_waiting

    def write(self, data: bytes) -> int:
        if not self.is_open:
            raise IOError("VirtualSerial is closed")
        if data is None:
            data = b""
        self._peer_rx.feed(bytes(data))
        return len(data)

    def read(self, n: int = 1) -> bytes:
        if not self.is_open:
            return b""
        return self._rx.read(int(n), self.timeout)


@dataclass
class VirtualPortInfo:
    key: str               # e.g. 'SIM:COM10'
    com: str               # e.g. 'COM10'
    baudrate: int
    host: VirtualSerial    # used by main app when selecting the simulated port
    sim: VirtualSerial     # used by simulator UI


class VirtualSerialRegistry:
    def __init__(self):
        self._ports: Dict[str, VirtualPortInfo] = {}
        self._lock = threading.Lock()

    def list_infos(self) -> List[VirtualPortInfo]:
        with self._lock:
            return list(self._ports.values())

    def create(self, com_num: int, baudrate: int = 9600) -> VirtualPortInfo:
        com = f"COM{int(com_num)}"
        key = f"SIM:{com}"
        with self._lock:
            if key in self._ports:
                raise ValueError(f"仿真串口已存在：{com}")
            ep_host = _VirtualEndpoint()
            ep_sim = _VirtualEndpoint()
            host = VirtualSerial(com, rx_ep=ep_host, peer_rx_ep=ep_sim)
            sim = VirtualSerial(com, rx_ep=ep_sim, peer_rx_ep=ep_host)
            info = VirtualPortInfo(key=key, com=com, baudrate=int(baudrate), host=host, sim=sim)
            self._ports[key] = info
            return info

    def remove(self, key: str):
        with self._lock:
            info = self._ports.pop(key, None)
        if info:
            try:
                info.host.close()
            except Exception:
                pass
            try:
                info.sim.close()
            except Exception:
                pass

    def get_host(self, key: str) -> Optional[VirtualSerial]:
        with self._lock:
            info = self._ports.get(key)
            return info.host if info else None

    def get_sim(self, key: str) -> Optional[VirtualSerial]:
        with self._lock:
            info = self._ports.get(key)
            return info.sim if info else None

    def set_baudrate(self, key: str, baudrate: int):
        """Update baudrate metadata for a virtual port (thread-safe).

        Note: Virtual ports are in-memory streams; baudrate is used for UI display and
        for user-facing configuration/estimation only.
        """
        with self._lock:
            info = self._ports.get(key)
            if info:
                info.baudrate = int(baudrate)


SIM_REGISTRY = VirtualSerialRegistry()


# ---------------- Modbus utilities ----------------

def crc16_modbus(data: bytes) -> int:
    """Return Modbus RTU CRC16 (poly 0xA001), little-endian on the wire."""
    crc = 0xFFFF
    for b in data:
        crc ^= b
        for _ in range(8):
            if crc & 0x0001:
                crc = (crc >> 1) ^ 0xA001
            else:
                crc >>= 1
    return crc & 0xFFFF


def hex_bytes(data: bytes) -> str:
    return " ".join(f"{b:02X}" for b in data)


DTYPE_INFO: Dict[str, Tuple[str, int]] = {
    "int16": ("h", 1),
    "uint16": ("H", 1),
    "int32": ("i", 2),
    "uint32": ("I", 2),
    "float32": ("f", 2),
    "float64": ("d", 4),
}


@dataclass
class ChannelConfig:
    enabled: bool
    name: str
    address: int
    dtype: str
    byte_order: str  # "big" or "little" within 16-bit word
    word_order: str  # "big" or "little" across words
    scale: float


@dataclass
class _ChSpec:
    ch: ChannelConfig
    start: int
    end: int
    reg_count: int


@dataclass
class _Block:
    start: int
    end: int
    channels: List[_ChSpec]

    @property
    def count(self) -> int:
        return self.end - self.start + 1


def regs_to_bytes(registers: List[int], byte_order: str, word_order: str) -> bytes:
    words = list(registers)
    if word_order == "little":
        words.reverse()
    out = bytearray()
    for w in words:
        hi = (w >> 8) & 0xFF
        lo = w & 0xFF
        if byte_order == "big":
            out.extend([hi, lo])
        else:
            out.extend([lo, hi])
    return bytes(out)


def decode_registers(registers: List[int], dtype: str, byte_order: str, word_order: str) -> float:
    if dtype not in DTYPE_INFO:
        raise ValueError(f"Unsupported dtype: {dtype}")
    fmt_char, reg_count = DTYPE_INFO[dtype]
    if len(registers) < reg_count:
        raise ValueError(f"Need {reg_count} regs for {dtype}, got {len(registers)}")
    raw = regs_to_bytes(registers[:reg_count], byte_order, word_order)
    # bytes already arranged: unpack as big-endian
    return float(struct.unpack(">" + fmt_char, raw)[0])


def build_blocks(channels: List[ChannelConfig], address_base_1: bool) -> List[_Block]:
    specs: List[_ChSpec] = []
    for ch in channels:
        if not ch.enabled:
            continue
        _, reg_count = DTYPE_INFO.get(ch.dtype, ("", 1))
        start = int(ch.address) - (1 if address_base_1 else 0)
        if start < 0:
            start = 0
        end = start + reg_count - 1
        specs.append(_ChSpec(ch=ch, start=start, end=end, reg_count=reg_count))

    specs.sort(key=lambda s: (s.start, s.end))
    if not specs:
        return []

    blocks: List[_Block] = []
    cur_start, cur_end = specs[0].start, specs[0].end
    cur_list = [specs[0]]
    for sp in specs[1:]:
        if sp.start <= cur_end + 1:
            cur_end = max(cur_end, sp.end)
            cur_list.append(sp)
        else:
            blocks.append(_Block(start=cur_start, end=cur_end, channels=cur_list))
            cur_start, cur_end, cur_list = sp.start, sp.end, [sp]
    blocks.append(_Block(start=cur_start, end=cur_end, channels=cur_list))
    return blocks


# ---------------- RS485 direction control ----------------

class Rs485CtrlMode:
    AUTO = "自动(不控制)"
    RTS = "RTS 控制 DE"
    DTR = "DTR 控制 DE"


@dataclass
class Rs485CtrlConfig:
    mode: str = Rs485CtrlMode.RTS
    tx_level_high: bool = True   # True => set pin True during TX
    rx_level_high: bool = False  # True => set pin True during RX
    pre_tx_ms: int = 0
    post_tx_ms: int = 2


def apply_rs485_rx_level(ser: serial.Serial, cfg: Rs485CtrlConfig):
    if cfg.mode == Rs485CtrlMode.RTS:
        ser.rts = bool(cfg.rx_level_high)
    elif cfg.mode == Rs485CtrlMode.DTR:
        ser.dtr = bool(cfg.rx_level_high)


def apply_rs485_tx_level(ser: serial.Serial, cfg: Rs485CtrlConfig):
    if cfg.mode == Rs485CtrlMode.RTS:
        ser.rts = bool(cfg.tx_level_high)
    elif cfg.mode == Rs485CtrlMode.DTR:
        ser.dtr = bool(cfg.tx_level_high)


# ---------------- Worker thread ----------------

class ModbusRtuWorker(QThread):
    data_ready = Signal(float, dict)      # ts, row dict
    status = Signal(str)
    connected = Signal(bool)
    acquiring = Signal(bool)
    log_line = Signal(str)                # to comm monitor
    # Structured frames for monitor/custom windows.
    # kind: 'TX'/'RX'/'TX_MANUAL'/'RX_MANUAL'
    # data: raw bytes (may be b'' for timeout/error)
    # tag: e.g. 'rx:COM3'/'tx:COM4'
    # note: extra message for empty frames
    frame = Signal(str, bytes, str, str)

    def __init__(
        self,
        port: str,
        baudrate: int,
        parity: str,
        stopbits: int,
        bytesize: int,
        timeout_s: float,
        rs485_cfg: Rs485CtrlConfig,
        tx_enabled: bool,
        tx_port: str,
        tx_baudrate: int,
        tx_interval_ms: int,
        parent=None
    ):
        super().__init__(parent)
        self.port = port
        self.baudrate = baudrate
        self.parity = parity
        self.stopbits = stopbits
        self.bytesize = bytesize
        self.timeout_s = timeout_s
        self.rs485_cfg = rs485_cfg

        self.tx_enabled = bool(tx_enabled)
        self.tx_port = tx_port
        self.tx_baudrate = int(tx_baudrate)
        self.tx_interval_ms = int(tx_interval_ms)

        self._running = True
        self._acquiring = False

        # Manual/custom send queue (processed in worker thread)
        self._custom_send_q: "queue.Queue[Tuple[str, str, bool]]" = queue.Queue()

        self._lock = threading.Lock()
        self.unit_id = 1
        self.func_code = 3
        self.poll_ms = 20
        self.address_base_1 = False
        self.channels: List[ChannelConfig] = []
        self.blocks: List[_Block] = []

        self._ser: Optional[serial.Serial] = None
        self._tx_ser: Optional[serial.Serial] = None

        # TX port async RX tap enabled (for comm monitor). When disabled, we won't
        # read from the TX serial in the background, avoiding consuming bytes.
        self.tx_tap_enabled = True

        # Output sending is decoupled from Modbus polling: we keep only the latest
        # formatted payload and send it at a fixed interval in a separate writer thread.
        self._latest_out_payload: bytes = b""
        self._latest_lock = threading.Lock()
        self._tx_write_q: "queue.Queue[bytes]" = queue.Queue(maxsize=1)
        self._tx_writer_thread: Optional[threading.Thread] = None

    # ---------- custom send (thread-safe enqueue from UI thread) ----------
    def enqueue_custom_send(self, target: str, text: str, add_crlf: bool):
        """Enqueue a manual send task.

        target: "rx" (Modbus serial) or "tx" (output serial)
        text: user input string; supports text or hex bytes (e.g. "01 03 00 00 00 02")
        add_crlf: whether to append \r\n (also applies to hex mode)
        """
        try:
            self._custom_send_q.put_nowait((str(target), str(text), bool(add_crlf)))
        except Exception:
            pass

    @staticmethod
    def _try_parse_hex_bytes(s: str) -> Optional[bytes]:
        """Parse common hex input formats.

        Accepts:
          - "01 03 00 00 00 02"
          - "0x01 0x03 00 00"
          - "010300000002"
        Returns bytes on success, else None.
        """
        t = (s or "").strip()
        if not t:
            return None
        # strip common separators
        t2 = t.replace(",", " ").replace("\t", " ")
        parts = [p for p in t2.split(" ") if p]

        # 1) spaced tokens
        if len(parts) >= 2:
            out = bytearray()
            for p in parts:
                pp = p.lower()
                if pp.startswith("0x"):
                    pp = pp[2:]
                if len(pp) != 2:
                    return None
                try:
                    out.append(int(pp, 16) & 0xFF)
                except Exception:
                    return None
            return bytes(out)

        # 2) single token, maybe continuous hex
        one = parts[0].lower() if parts else ""
        if one.startswith("0x"):
            one = one[2:]
        if len(one) >= 2 and all(ch in "0123456789abcdef" for ch in one) and (len(one) % 2 == 0):
            try:
                return bytes(int(one[i:i+2], 16) & 0xFF for i in range(0, len(one), 2))
            except Exception:
                return None
        return None

    @staticmethod
    def _read_until_idle(
        ser: serial.Serial,
        idle_ms: int = 20,
        max_ms: int = 200,
        max_bytes: int = 4096,
    ) -> bytes:
        """Read bytes until the line stays idle for a short period.

        This is used for "manual send" reply listening. It avoids needing to know
        the expected frame length.
        """
        if ser is None:
            return b""

        t0 = time.time()
        last_rx = t0
        buf = bytearray()
        orig_timeout = getattr(ser, "timeout", None)

        try:
            # non-blocking reads
            try:
                ser.timeout = 0
            except Exception:
                pass

            while True:
                try:
                    chunk = ser.read(1024)
                except Exception:
                    chunk = b""

                if chunk:
                    buf.extend(chunk)
                    last_rx = time.time()
                    if len(buf) >= max_bytes:
                        break
                else:
                    now = time.time()
                    if (now - last_rx) * 1000.0 >= float(idle_ms):
                        break
                    if (now - t0) * 1000.0 >= float(max_ms):
                        break
                    time.sleep(0.002)
        finally:
            try:
                if orig_timeout is not None:
                    ser.timeout = orig_timeout
            except Exception:
                pass

        return bytes(buf)

    def _process_custom_sends(self, max_items: int = 50):
        """Process pending manual sends in the worker thread."""
        processed = 0
        while processed < max_items:
            try:
                target, text, add_crlf = self._custom_send_q.get_nowait()
            except queue.Empty:
                break

            target = (target or "").strip().lower()
            add_crlf = bool(add_crlf)
            text = text or ""

            # choose serial
            ser = None
            tag = "?"
            if target == "rx":
                ser = self._ser
                tag = f"rx:{self.port}"
            elif target == "tx":
                ser = self._tx_ser
                tag = f"tx:{self.tx_port}"

            if ser is None:
                self._emit_frame("TX_MANUAL", b"", tag=tag, note="<target not connected>")
                processed += 1
                continue

            payload = self._try_parse_hex_bytes(text)
            mode = "HEX" if payload is not None else "TEXT"
            if payload is None:
                payload = text.encode("utf-8", errors="replace")

            if add_crlf:
                payload += b"\r\n"

            # optional RS485 direction control (important for many RS485 modules)
            # If enabled, apply RS485 direction control for manual sends.
            # Many RS485 modules need RTS/DTR toggling to actually drive the bus (TX灯才会亮)。
            use_rs485_dir = (self.rs485_cfg.mode != Rs485CtrlMode.AUTO)

            try:
                if use_rs485_dir:
                    try:
                        apply_rs485_tx_level(ser, self.rs485_cfg)
                        if self.rs485_cfg.pre_tx_ms > 0:
                            time.sleep(self.rs485_cfg.pre_tx_ms / 1000.0)
                    except Exception:
                        pass

                ser.write(payload)
                ser.flush()
                self._emit_frame("TX_MANUAL", payload, tag=tag, note=f"[{mode}]")

                if use_rs485_dir:
                    try:
                        if self.rs485_cfg.post_tx_ms > 0:
                            time.sleep(self.rs485_cfg.post_tx_ms / 1000.0)
                        apply_rs485_rx_level(ser, self.rs485_cfg)
                    except Exception:
                        pass

                # listen for reply (best-effort)
                reply = self._read_until_idle(ser, idle_ms=20, max_ms=200)
                if reply:
                    self._emit_frame("RX_MANUAL", reply, tag=tag)
                else:
                    self._emit_frame("RX_MANUAL", b"", tag=tag, note="<no reply>")

            except Exception as e:
                self._emit_frame("TX_MANUAL", b"", tag=tag, note=f"<error> {e}")

            processed += 1

    def stop_thread(self):
        self._running = False

    def set_acquiring(self, on: bool):
        self._acquiring = bool(on)
        self.acquiring.emit(self._acquiring)

    def update_runtime(self, unit_id: int, func_code: int, poll_ms: int, address_base_1: bool, channels: List[ChannelConfig], tx_interval_ms: Optional[int] = None):
        with self._lock:
            self.unit_id = int(unit_id)
            self.func_code = int(func_code)
            self.poll_ms = int(poll_ms)
            if tx_interval_ms is not None:
                self.tx_interval_ms = int(tx_interval_ms)
            self.address_base_1 = bool(address_base_1)
            self.channels = list(channels)
            self.blocks = build_blocks(self.channels, self.address_base_1)

    def _log(self, s: str):
        self.log_line.emit(s)

    def _emit_frame(self, kind: str, data: bytes, tag: str = "", note: str = ""):
        """Emit raw frame data for UI display.

        The UI can render as HEX or decode as text (UTF-8/GBK).
        """
        try:
            self.frame.emit(str(kind), bytes(data or b""), str(tag or ""), str(note or ""))
        except Exception:
            pass
    def set_tx_tap_enabled(self, enabled: bool):
        """Enable/disable async RX tap for TX/output serial.

        仅影响后台窥探式读取（用于通讯监视窗口），不影响业务发送。
        """
        self.tx_tap_enabled = bool(enabled)


    def _pump_async_rx(self, max_bytes: int = 512):
        """Best-effort async RX monitor for the *TX/output* serial.

        用途：让“通讯监视窗口”能同时看到发送串口(tx)收到的 RX 数据。
        不改变业务逻辑，只做窥探式读取（read existing bytes）。
        """
        ser = self._tx_ser
        if ser is None:
            return

        try:
            n = int(getattr(ser, "in_waiting", 0) or 0)
        except Exception:
            n = 0

        if max_bytes is not None and max_bytes>0 and n>max_bytes:
            n = int(max_bytes)
        if n <= 0:
            return

        try:
            n = int(n)
            if max_bytes is not None and int(max_bytes) > 0:
                n = min(n, int(max_bytes))
        except Exception:
            pass

        try:
            # read currently available bytes (non-blocking)
            chunk = ser.read(n)
        except Exception:
            chunk = b""

        if chunk:
            self._emit_frame("RX_TX", chunk, tag=f"tx:{self.tx_port}")

    def _send_recv_readregs(self, start_addr: int, count: int) -> Optional[List[int]]:
        """
        Send Modbus RTU request (03/04), return list of 16-bit registers or None on timeout/error.
        Request: [id][fc][addr_hi][addr_lo][qty_hi][qty_lo][crc_lo][crc_hi]
        Response: [id][fc][byte_count][data...][crc_lo][crc_hi]
        """
        ser = self._ser
        if ser is None:
            return None

        unit = self.unit_id & 0xFF
        fc = self.func_code & 0xFF
        req_wo_crc = bytes([
            unit, fc,
            (start_addr >> 8) & 0xFF, start_addr & 0xFF,
            (count >> 8) & 0xFF, count & 0xFF
        ])
        crc = crc16_modbus(req_wo_crc)
        req = req_wo_crc + bytes([crc & 0xFF, (crc >> 8) & 0xFF])

        # flush stale input to avoid mixing old bytes
        try:
            ser.reset_input_buffer()
        except Exception:
            pass

        # direction control: TX
        if self.rs485_cfg.mode != Rs485CtrlMode.AUTO:
            apply_rs485_tx_level(ser, self.rs485_cfg)
            if self.rs485_cfg.pre_tx_ms > 0:
                time.sleep(self.rs485_cfg.pre_tx_ms / 1000.0)

        # write
        try:
            ser.write(req)
            ser.flush()
            self._emit_frame("TX", req, tag=f"rx:{self.port}")
        except Exception as e:
            self._log(f"TX_ERR: {e}")
            return None

        # direction control: back to RX
        if self.rs485_cfg.mode != Rs485CtrlMode.AUTO:
            if self.rs485_cfg.post_tx_ms > 0:
                time.sleep(self.rs485_cfg.post_tx_ms / 1000.0)
            apply_rs485_rx_level(ser, self.rs485_cfg)

        # read response
        # read first 3 bytes
        hdr = ser.read(3)
        if len(hdr) < 3:
            self._emit_frame("RX", b"", tag=f"rx:{self.port}", note="<timeout/no header>")
            return None

        resp_unit, resp_fc, third = hdr[0], hdr[1], hdr[2]
        if resp_unit != unit:
            # might be noise; read remaining quickly but report
            tail = ser.read(256)
            self._emit_frame("RX", hdr + tail, tag=f"rx:{self.port}", note="<unit mismatch>")
            return None

        # exception response
        if resp_fc == (fc | 0x80):
            exc = ser.read(2)  # exception code + crc?
            more = ser.read(2)
            full = hdr + exc + more
            self._log(f"RX_EXC: {hex_bytes(full)}")
            return None

        if resp_fc != fc:
            tail = ser.read(256)
            self._emit_frame("RX", hdr + tail, tag=f"rx:{self.port}", note="<fc mismatch>")
            return None

        byte_count = third
        expected_data = byte_count + 2  # data + crc
        rest = ser.read(expected_data)
        if len(rest) < expected_data:
            self._emit_frame("RX", hdr + rest, tag=f"rx:{self.port}", note="<timeout/incomplete>")
            return None

        resp = hdr + rest
        self._emit_frame("RX", resp, tag=f"rx:{self.port}")

        # CRC check
        body = resp[:-2]
        crc_rx = resp[-2] | (resp[-1] << 8)
        crc_calc = crc16_modbus(body)
        if crc_rx != crc_calc:
            self._log(f"CRC_ERR: rx={crc_rx:04X} calc={crc_calc:04X}")
            return None

        if byte_count != count * 2:
            # still parse what we got, but warn
            self._log(f"LEN_WARN: byte_count={byte_count}, expected={count*2}")

        data = resp[3:3 + byte_count]
        regs = []
        for i in range(0, len(data), 2):
            if i + 1 >= len(data):
                break
            regs.append((data[i] << 8) | data[i + 1])
        return regs


    def _open_any_serial(self, port_key: str, baudrate: int, parity: str = 'N', stopbits: int = 1, bytesize: int = 8, timeout: float = 0.2):
        """Open a physical COM port or a simulated port.

        Simulated ports are addressed with key like 'SIM:COM10'.
        """
        if (port_key or '').startswith('SIM:'):
            ser = SIM_REGISTRY.get_host(port_key)
            if ser is None:
                raise ValueError(f"仿真串口不存在或未启动：{port_key}")
            try:
                ser.timeout = float(timeout)
            except Exception:
                pass
            # VirtualSerial may be closed by a previous disconnect; reopen it.
            try:
                if hasattr(ser, 'open'):
                    ser.open()
                else:
                    ser.is_open = True
            except Exception:
                pass
            return ser

        return serial.Serial(
            port=port_key,
            baudrate=int(baudrate),
            parity=parity,
            stopbits=int(stopbits),
            bytesize=int(bytesize),
            timeout=float(timeout),
            write_timeout=0.05,
        )

    def _tx_writer_loop(self):
        """Write TX/output serial in a background thread (so it never blocks Modbus polling)."""
        while self._running:
            ser = self._tx_ser
            if ser is None:
                time.sleep(0.05)
                continue
            try:
                payload = self._tx_write_q.get(timeout=0.2)
            except queue.Empty:
                continue
            if not payload:
                continue
            try:
                # RS485 direction control (if user enabled)
                if self.rs485_cfg.mode != Rs485CtrlMode.AUTO:
                    try:
                        apply_rs485_tx_level(ser, self.rs485_cfg)
                        if self.rs485_cfg.pre_tx_ms > 0:
                            time.sleep(self.rs485_cfg.pre_tx_ms / 1000.0)
                    except Exception:
                        pass

                ser.write(payload)
                # do not flush aggressively (avoid blocking). virtual serial flush is no-op.
                try:
                    ser.flush()
                except Exception:
                    pass

                if self.rs485_cfg.mode != Rs485CtrlMode.AUTO:
                    try:
                        if self.rs485_cfg.post_tx_ms > 0:
                            time.sleep(self.rs485_cfg.post_tx_ms / 1000.0)
                        apply_rs485_rx_level(ser, self.rs485_cfg)
                    except Exception:
                        pass

                self._emit_frame('TX_TX', payload, tag=f'tx:{self.tx_port}')
            except Exception as e:
                self._log(f'TX_OUT_ERR: {e}')

    def _queue_tx_payload_latest(self, payload: bytes):
        """Keep only the latest payload in the TX queue."""
        if payload is None:
            return
        try:
            while True:
                self._tx_write_q.get_nowait()
        except Exception:
            pass
        try:
            self._tx_write_q.put_nowait(bytes(payload))
        except Exception:
            pass

    def run(self):
        # open serial
        try:
            ser = self._open_any_serial(
                port_key=self.port,
                baudrate=self.baudrate,
                parity=self.parity,
                stopbits=self.stopbits,
                bytesize=self.bytesize,
                timeout=self.timeout_s,
            )
            self._ser = ser

            # open TX (output) serial if enabled
            if self.tx_enabled:
                if not self.tx_port:
                    raise ValueError('已启用发送串口，但未选择发送串口')
                tx_ser = self._open_any_serial(
                    port_key=self.tx_port,
                    baudrate=self.tx_baudrate,
                    parity='N',
                    stopbits=1,
                    bytesize=8,
                    timeout=0.05,
                )
                self._tx_ser = tx_ser

                # start background writer
                self._tx_writer_thread = threading.Thread(target=self._tx_writer_loop, daemon=True)
                self._tx_writer_thread.start()

        except Exception as e:
            self.status.emit(f'连接失败：{e}')
            self.connected.emit(False)
            return

        # default RS485 RX level to avoid blocking bus
        try:
            if self.rs485_cfg.mode != Rs485CtrlMode.AUTO:
                apply_rs485_rx_level(ser, self.rs485_cfg)
        except Exception:
            pass

        self.connected.emit(True)
        self.status.emit(f'已连接 {self.port} @ {self.baudrate}（未采集）')
        self._log(f'INFO: connected {self.port} @ {self.baudrate}, rs485_mode={self.rs485_cfg.mode}')
        if self.tx_enabled and self._tx_ser is not None:
            self._log(f'INFO: tx_port={self.tx_port} @ {self.tx_baudrate} (fixed interval sender)')

        last_warn_ts = 0.0

        # fixed schedules
        next_poll = time.monotonic()
        next_tx = time.monotonic()

        try:
            while self._running:
                # Always process manual/custom send tasks (even when not acquiring)
                self._process_custom_sends()

                # Async RX tap on TX/output port (if enabled)
                if self.tx_tap_enabled:
                    self._pump_async_rx(max_bytes=512)

                if not self._acquiring:
                    time.sleep(0.03)
                    # keep schedules fresh
                    next_poll = time.monotonic()
                    next_tx = time.monotonic()
                    continue

                with self._lock:
                    poll_ms = int(self.poll_ms)
                    tx_iv_ms = int(self.tx_interval_ms)
                    channels = list(self.channels)
                    blocks = list(self.blocks)

                if not blocks or not channels:
                    time.sleep(0.02)
                    continue

                now = time.monotonic()

                # 1) Modbus polling at fixed interval (independent of TX sending)
                if now >= next_poll:
                    next_poll += max(0.02, poll_ms / 1000.0)

                    ts = time.time()
                    row: Dict[str, Optional[float]] = {ch.name: None for ch in channels if ch.enabled}
                    any_success = False

                    for blk in blocks:
                        if not self._running or not self._acquiring:
                            break
                        # allow low-latency manual sends between blocks
                        self._process_custom_sends()
                        regs = self._send_recv_readregs(blk.start, blk.count)
                        if regs is None or len(regs) < blk.count:
                            continue

                        for sp in blk.channels:
                            try:
                                offset = sp.start - blk.start
                                slice_regs = regs[offset: offset + sp.reg_count]
                                val = decode_registers(slice_regs, sp.ch.dtype, sp.ch.byte_order, sp.ch.word_order)
                                val *= float(sp.ch.scale)
                                row[sp.ch.name] = val
                                any_success = True
                            except Exception:
                                row[sp.ch.name] = None

                    if any_success:
                        self.data_ready.emit(ts, row)

                        # build and store latest output line; actual sending is periodic
                        if self.tx_enabled and self._tx_ser is not None:
                            try:
                                parts = ['Data']
                                for ch in channels:
                                    if not ch.enabled:
                                        continue
                                    v = row.get(ch.name, None)
                                    parts.append('nan' if v is None else (f'{v:.6g}'))
                                line = (' '.join(parts) + '\r\n').encode('ascii', errors='replace')
                                with self._latest_lock:
                                    self._latest_out_payload = line
                            except Exception:
                                pass
                    else:
                        tnow = time.time()
                        if tnow - last_warn_ts > 2.0:
                            self.status.emit('采集中：未收到有效响应（请看通讯监视窗口）')
                            last_warn_ts = tnow

                # 2) Output sending at fixed interval (send latest received data)
                if self.tx_enabled and self._tx_ser is not None and tx_iv_ms > 0:
                    now2 = time.monotonic()
                    if now2 >= next_tx:
                        next_tx += max(0.005, tx_iv_ms / 1000.0)
                        with self._latest_lock:
                            payload = bytes(self._latest_out_payload or b'')
                        if payload:
                            self._queue_tx_payload_latest(payload)

                # Sleep until next event (avoid busy loop)
                soon = min(next_poll, next_tx) if (self.tx_enabled and self._tx_ser is not None) else next_poll
                dt = soon - time.monotonic()
                if dt > 0.05:
                    dt = 0.05
                if dt > 0:
                    time.sleep(dt)

        finally:
            try:
                if self._ser:
                    self._ser.close()
            except Exception:
                pass
            self._ser = None

            try:
                if self._tx_ser:
                    self._tx_ser.close()
            except Exception:
                pass
            self._tx_ser = None

            self.connected.emit(False)
            self.acquiring.emit(False)
            self.status.emit('已断开连接')
            self._log('INFO: disconnected')


# ---------------- UI ----------------


class SerialSimManagerWindow(QMainWindow):
    """串口仿真器（程序内虚拟串口）。

    说明：
    - 这些“仿真串口”只存在于本程序内部，不会在系统中创建真实 COM 口。
    - 创建后，会出现在主界面的串口下拉框中：例如 “COM10  —  仿真串口”。
    """

    ports_changed = Signal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle('串口仿真')
        self.resize(760, 560)

        cw = QWidget()
        self.setCentralWidget(cw)
        v = QVBoxLayout(cw)

        top = QHBoxLayout()
        top.addWidget(QLabel('COM号'))
        self.com_spin = QSpinBox()
        self.com_spin.setRange(1, 256)
        self.com_spin.setValue(10)
        top.addWidget(self.com_spin)

        top.addWidget(QLabel('波特率'))
        self.baud_combo = QComboBox()
        for b in [9600, 19200, 38400, 57600, 115200, 230400]:
            self.baud_combo.addItem(str(b), b)
        self.baud_combo.setCurrentText('115200')
        top.addWidget(self.baud_combo)

        self.btn_create = QPushButton('创建仿真串口')
        self.btn_remove = QPushButton('移除当前')
        top.addWidget(self.btn_create)
        top.addWidget(self.btn_remove)
        top.addStretch(1)
        v.addLayout(top)

        self.tabs = QTabWidget()
        v.addWidget(self.tabs, 1)

        self.btn_create.clicked.connect(lambda *_: self.create_port())
        self.btn_remove.clicked.connect(lambda *_: self.remove_current())

        self.rebuild_tabs()

    def closeEvent(self, ev):
        # 关闭时隐藏窗口，不销毁（保持仿真逻辑继续运行）
        try:
            ev.ignore()
        except Exception:
            pass
        self.hide()

    def rebuild_tabs(self):
        self.tabs.clear()
        for info in SIM_REGISTRY.list_infos():
            w = SimPortWidget(info, parent=self)
            self.tabs.addTab(w, info.com)

    def create_port(self):
        com_num = int(self.com_spin.value())
        baud = int(self.baud_combo.currentData() or 9600)
        try:
            info = SIM_REGISTRY.create(com_num=com_num, baudrate=baud)
        except Exception as e:
            QMessageBox.warning(self, '创建失败', str(e))
            return
        w = SimPortWidget(info, parent=self)
        self.tabs.addTab(w, info.com)
        self.tabs.setCurrentWidget(w)
        self.ports_changed.emit()

    def remove_current(self):
        w = self.tabs.currentWidget()
        if w is None:
            return
        key = getattr(w, 'port_key', '')
        if not key:
            return
        SIM_REGISTRY.remove(key)
        self.rebuild_tabs()
        self.ports_changed.emit()


class SimPortWidget(QWidget):
    """单个仿真串口的配置页。"""

    def __init__(self, info: VirtualPortInfo, parent=None):
        super().__init__(parent)
        self.info = info
        self.port_key = info.key
        self.ser = info.sim

        self._rx_buf = bytearray()

        # log entries are stored as structured data so the user can switch
        # between HEX / text(UTF-8/GBK) display at any time.
        self._log_entries: List[dict] = []
        self._log_dirty = False
        self._log_timer = QTimer(self)
        self._log_timer.setInterval(50)
        self._log_timer.setSingleShot(True)
        self._log_timer.timeout.connect(self._flush_log_render)

        root = QVBoxLayout(self)
        root.addWidget(QLabel(f"{info.com}  (key: {info.key})"))

        top_row = QHBoxLayout()
        top_row.addWidget(QLabel('波特率'))
        self.port_baud_combo = QComboBox()
        for b in [9600, 19200, 38400, 57600, 115200, 230400, 460800, 921600]:
            self.port_baud_combo.addItem(str(b), b)
        try:
            self.port_baud_combo.setCurrentText(str(int(getattr(info, 'baudrate', 115200))))
        except Exception:
            self.port_baud_combo.setCurrentText('115200')
        self.port_baud_combo.currentIndexChanged.connect(self.on_port_baud_changed)
        top_row.addWidget(self.port_baud_combo)
        top_row.addStretch(1)
        root.addLayout(top_row)

        # 连续发送
        g1 = QGroupBox('连续发送模式')
        l1 = QGridLayout(g1)
        self.cont_enable = QCheckBox('启用')
        self.cont_interval = QSpinBox()
        self.cont_interval.setRange(5, 100000)
        self.cont_interval.setValue(100)
        self.cont_interval.setSuffix(' ms')
        self.cont_source = QComboBox()
        self.cont_source.addItem('随机数', 'random')
        self.cont_source.addItem('固定文本', 'fixed')
        l1.addWidget(self.cont_enable, 0, 0)
        l1.addWidget(QLabel('间隔'), 0, 1)
        l1.addWidget(self.cont_interval, 0, 2)
        l1.addWidget(QLabel('发送内容'), 0, 3)
        l1.addWidget(self.cont_source, 0, 4)
        root.addWidget(g1)

        # 随机数配置
        g2 = QGroupBox('随机数生成')
        l2 = QGridLayout(g2)
        self.rand_channels = QSpinBox()
        self.rand_channels.setRange(1, 200)
        self.rand_channels.setValue(2)
        self.rand_min = QDoubleSpinBox()
        self.rand_min.setRange(-1e9, 1e9)
        self.rand_min.setValue(-1.0)
        self.rand_max = QDoubleSpinBox()
        self.rand_max.setRange(-1e9, 1e9)
        self.rand_max.setValue(1.0)
        self.rand_mode = QComboBox()
        self.rand_mode.addItem('纯文本(逗号分隔)', 'text')
        self.rand_mode.addItem('Modbus模式(03响应帧)', 'modbus')
        self.rand_unit = QSpinBox()
        self.rand_unit.setRange(1, 247)
        self.rand_unit.setValue(1)
        self.rand_crlf = QCheckBox('末尾\r\n(文本)')
        self.rand_crlf.setChecked(True)
        self.enc_combo = QComboBox()
        self.enc_combo.addItem('UTF-8', 'utf-8')
        self.enc_combo.addItem('GBK', 'gbk')
        self.btn_rand_once = QPushButton('随机单次发送')

        l2.addWidget(QLabel('通道数'), 0, 0)
        l2.addWidget(self.rand_channels, 0, 1)
        l2.addWidget(QLabel('下限'), 0, 2)
        l2.addWidget(self.rand_min, 0, 3)
        l2.addWidget(QLabel('上限'), 0, 4)
        l2.addWidget(self.rand_max, 0, 5)

        l2.addWidget(QLabel('输出模式'), 1, 0)
        l2.addWidget(self.rand_mode, 1, 1, 1, 2)
        l2.addWidget(QLabel('Modbus地址'), 1, 3)
        l2.addWidget(self.rand_unit, 1, 4)
        l2.addWidget(self.rand_crlf, 1, 5)

        l2.addWidget(QLabel('文本编码'), 2, 0)
        l2.addWidget(self.enc_combo, 2, 1)
        l2.addWidget(self.btn_rand_once, 2, 5)

        root.addWidget(g2)

        # 固定文本
        g3 = QGroupBox('固定信号')
        l3 = QGridLayout(g3)
        self.fixed_cont = QLineEdit()
        self.fixed_cont.setPlaceholderText('固定信号(连续发送/触发发送时使用)')
        self.fixed_cont_crlf = QCheckBox('末尾\r\n')
        self.fixed_cont_crlf.setChecked(True)

        self.fixed_once = QLineEdit()
        self.fixed_once.setPlaceholderText('固定信号(单次发送)')
        self.fixed_once_crlf = QCheckBox('末尾\r\n')
        self.fixed_once_crlf.setChecked(True)
        self.btn_fixed_once = QPushButton('单次发送')

        l3.addWidget(QLabel('固定(连续/触发)'), 0, 0)
        l3.addWidget(self.fixed_cont, 0, 1, 1, 3)
        l3.addWidget(self.fixed_cont_crlf, 0, 4)

        l3.addWidget(QLabel('固定(单次)'), 1, 0)
        l3.addWidget(self.fixed_once, 1, 1, 1, 3)
        l3.addWidget(self.fixed_once_crlf, 1, 4)
        l3.addWidget(self.btn_fixed_once, 1, 5)

        root.addWidget(g3)

        # 触发发送
        g4 = QGroupBox('触发发送模式')
        l4 = QGridLayout(g4)
        self.tr_enable = QCheckBox('启用触发')
        self.tr_type = QComboBox()
        self.tr_type.addItem('Modbus解析模式(识别03/04读请求)', 'modbus')
        self.tr_type.addItem('指令模式(以\\r\\n或\\n结尾)', 'cmd')
        self.tr_cmd = QLineEdit()
        self.tr_cmd.setPlaceholderText('触发指令(不需要输入\\r\\n)')
        self.tr_source = QComboBox()
        self.tr_source.addItem('随机数', 'random')
        self.tr_source.addItem('固定文本', 'fixed')

        l4.addWidget(self.tr_enable, 0, 0)
        l4.addWidget(QLabel('触发类型'), 0, 1)
        l4.addWidget(self.tr_type, 0, 2, 1, 2)
        l4.addWidget(QLabel('触发后发送'), 0, 4)
        l4.addWidget(self.tr_source, 0, 5)
        l4.addWidget(QLabel('触发指令'), 1, 0)
        l4.addWidget(self.tr_cmd, 1, 1, 1, 5)
        root.addWidget(g4)

        # 日志显示模式
        log_mode_row = QHBoxLayout()
        log_mode_row.addWidget(QLabel('日志显示'))
        self.log_mode_combo = QComboBox()
        self.log_mode_combo.addItem('HEX', 'hex')
        self.log_mode_combo.addItem('文本(UTF-8)', 'utf-8')
        self.log_mode_combo.addItem('文本(GBK)', 'gbk')
        self.log_mode_combo.setCurrentIndex(1)
        self.log_mode_combo.currentIndexChanged.connect(lambda *_: self._schedule_log_render(full=True))
        log_mode_row.addWidget(self.log_mode_combo)
        log_mode_row.addStretch(1)
        root.addLayout(log_mode_row)

        self.log = QPlainTextEdit()
        self.log.setReadOnly(True)
        self.log.setUndoRedoEnabled(False)
        try:
            self.log.setMaximumBlockCount(900)
        except Exception:
            pass
        self.log.setMinimumHeight(140)
        root.addWidget(QLabel('日志(仅本窗口)'))
        root.addWidget(self.log, 1)

        # timers
        self._rx_timer = QTimer(self)
        self._rx_timer.setInterval(10)
        self._rx_timer.timeout.connect(self._on_rx_timer)
        self._rx_timer.start()

        self._cont_timer = QTimer(self)
        self._cont_timer.timeout.connect(self._on_cont_timer)

        # signals
        self.cont_enable.toggled.connect(self._update_cont_timer)
        self.cont_interval.valueChanged.connect(self._update_cont_timer)
        self.btn_fixed_once.clicked.connect(lambda *_: self.send_fixed_once())
        self.btn_rand_once.clicked.connect(lambda *_: self.send_random_once())
        self.tr_type.currentIndexChanged.connect(self._update_tr_ui)
        self._update_tr_ui()
        self._update_cont_timer()
    def _log(self, s: str):
        self._append_log_text(str(s))

    def _append_log_text(self, s: str):
        self._log_entries.append({'t': 'text', 'text': str(s)})
        # cap memory
        if len(self._log_entries) > 2000:
            overflow = len(self._log_entries) - 2000
            del self._log_entries[:overflow]
            try:
                self._log_render_idx = max(0, int(getattr(self, '_log_render_idx', 0)) - overflow)
            except Exception:
                self._log_render_idx = 0
        self._schedule_log_render(full=False)

    def _append_log_bytes(self, prefix: str, data: bytes):
        self._log_entries.append({'t': 'bytes', 'prefix': str(prefix), 'data': bytes(data or b'')})
        if len(self._log_entries) > 2000:
            overflow = len(self._log_entries) - 2000
            del self._log_entries[:overflow]
            try:
                self._log_render_idx = max(0, int(getattr(self, '_log_render_idx', 0)) - overflow)
            except Exception:
                self._log_render_idx = 0
        self._schedule_log_render(full=False)

    def _schedule_log_render(self, full: bool = False):
        if full:
            setattr(self, '_log_force_full', True)
        self._log_dirty = True
        if not self._log_timer.isActive():
            self._log_timer.start()

    def _flush_log_render(self):
        if not self._log_dirty:
            return
        self._log_dirty = False
        force_full = bool(getattr(self, '_log_force_full', False))
        try:
            self.render_log(force_full=force_full)
        finally:
            self._log_force_full = False

    def _decode_for_log(self, data: bytes, mode: str) -> str:
        if mode == 'hex':
            return hex_bytes(data)
        enc = 'utf-8' if mode == 'utf-8' else 'gbk'
        try:
            return data.decode(enc, errors='replace')
        except Exception:
            return data.decode('utf-8', errors='replace')

    def _format_log_entry(self, e: dict, mode: str) -> str:
        if e.get('t') == 'text':
            return str(e.get('text', ''))
        b = e.get('data', b'') or b''
        pfx = str(e.get('prefix', ''))
        show = b if len(b) <= 128 else b[:128]
        payload = self._decode_for_log(show, mode)
        if len(b) > 128:
            payload = payload + ' ...'
        return f"{pfx}({len(b)}): {payload}"

    def render_log(self, force_full: bool = False):
        mode = 'utf-8'
        if hasattr(self, 'log_mode_combo'):
            mode = self.log_mode_combo.currentData() or 'utf-8'

        last_mode = getattr(self, '_log_render_mode', None)
        render_idx = int(getattr(self, '_log_render_idx', 0) or 0)

        if force_full or last_mode != mode or render_idx <= 0:
            max_lines = 900
            entries = self._log_entries[-max_lines:]
            out_lines = [self._format_log_entry(e, mode) for e in entries]
            try:
                self.log.blockSignals(True)
                self.log.setPlainText('\n'.join(out_lines))
                self.log.blockSignals(False)
                self.log.moveCursor(QTextCursor.End)
            except Exception:
                pass
            self._log_render_mode = mode
            self._log_render_idx = len(self._log_entries)
            return

        if render_idx < len(self._log_entries):
            new_entries = self._log_entries[render_idx:]
            out_lines = [self._format_log_entry(e, mode) for e in new_entries]
            if out_lines:
                try:
                    self.log.blockSignals(True)
                    self.log.appendPlainText('\n'.join(out_lines))
                    self.log.blockSignals(False)
                    self.log.moveCursor(QTextCursor.End)
                except Exception:
                    pass
        self._log_render_mode = mode
        self._log_render_idx = len(self._log_entries)

    def on_port_baud_changed(self):
        try:
            baud = int(self.port_baud_combo.currentData() or 9600)
        except Exception:
            baud = 9600
        try:
            SIM_REGISTRY.set_baudrate(self.port_key, baud)
            try:
                self.info.baudrate = baud
            except Exception:
                pass
            self._append_log_text(f"INFO: 波特率已设置为 {baud}")
        except Exception as e:
            self._append_log_text(f"BAUD_ERR: {e}")

    def _update_tr_ui(self):
        self.tr_cmd.setEnabled(self.tr_type.currentData() == 'cmd')

    def _update_cont_timer(self):
        if not self.cont_enable.isChecked():
            self._cont_timer.stop()
            return
        iv = max(5, int(self.cont_interval.value()))
        self._cont_timer.setInterval(iv)
        if not self._cont_timer.isActive():
            self._cont_timer.start()

    def _encode_text(self, s: str, add_crlf: bool) -> bytes:
        if add_crlf:
            s = s + '\r\n'
        enc = str(self.enc_combo.currentData() or 'utf-8')
        try:
            return s.encode(enc, errors='replace')
        except Exception:
            return s.encode('utf-8', errors='replace')

    def _rand_values(self, n: int):
        import random
        lo = float(self.rand_min.value())
        hi = float(self.rand_max.value())
        if hi < lo:
            lo, hi = hi, lo
        return [random.uniform(lo, hi) for _ in range(n)]

    def _build_random_payload(self) -> bytes:
        mode = self.rand_mode.currentData()
        n = int(self.rand_channels.value())
        vals = self._rand_values(n)

        if mode == 'modbus':
            # 生成一个“03响应帧”（可用于连续发送；也可用于指令触发发送）
            unit = int(self.rand_unit.value()) & 0xFF
            fc = 0x03
            regs = []
            for v in vals:
                iv = int(round(v))
                iv = max(-32768, min(32767, iv))
                regs.append(iv & 0xFFFF)
            data = bytearray([unit, fc, len(regs) * 2])
            for r in regs:
                data.append((r >> 8) & 0xFF)
                data.append(r & 0xFF)
            crc = crc16_modbus(bytes(data))
            data.append(crc & 0xFF)
            data.append((crc >> 8) & 0xFF)
            return bytes(data)

        # text
        s = ','.join([f'{v:.6g}' for v in vals])
        return self._encode_text(s, add_crlf=bool(self.rand_crlf.isChecked()))

    def _build_fixed_payload(self, s: str, add_crlf: bool) -> bytes:
        return self._encode_text(s or '', add_crlf=add_crlf)
    def _send_bytes(self, b: bytes, prefix: str = 'TX'):
        if not b:
            return
        try:
            self.ser.write(b)
            self._append_log_bytes(prefix, b)
        except Exception as e:
            self._append_log_text(f'{prefix}_ERR: {e}')

    def send_fixed_once(self):
        b = self._build_fixed_payload(self.fixed_once.text(), bool(self.fixed_once_crlf.isChecked()))
        self._send_bytes(b, prefix='TX_ONCE')

    def send_random_once(self):
        b = self._build_random_payload()
        self._send_bytes(b, prefix='TX_RAND')

    def _on_cont_timer(self):
        if not self.cont_enable.isChecked():
            return
        src = self.cont_source.currentData()
        if src == 'random':
            b = self._build_random_payload()
        else:
            b = self._build_fixed_payload(self.fixed_cont.text(), bool(self.fixed_cont_crlf.isChecked()))
        self._send_bytes(b, prefix='TX_CONT')

    def _consume_modbus_read_req(self):
        """Consume one Modbus RTU read request frame (03/04)."""
        buf = self._rx_buf
        while len(buf) >= 8:
            unit = buf[0]
            fc = buf[1]
            if fc not in (0x03, 0x04):
                del buf[0]
                continue
            frame = bytes(buf[:8])
            body = frame[:6]
            crc_rx = frame[6] | (frame[7] << 8)
            if crc16_modbus(body) != crc_rx:
                del buf[0]
                continue
            start = (frame[2] << 8) | frame[3]
            count = (frame[4] << 8) | frame[5]
            del buf[:8]
            return unit, fc, start, count
        return None

    def _handle_trigger_modbus(self):
        req = self._consume_modbus_read_req()
        if not req:
            return
        unit, fc, start, count = req
        cfg_unit = int(self.rand_unit.value()) & 0xFF
        if unit != cfg_unit:
            return

        # 按请求数量返回随机 16-bit 有符号寄存器（四舍五入）
        try:
            import random
            lo = float(self.rand_min.value())
            hi = float(self.rand_max.value())
            if hi < lo:
                lo, hi = hi, lo
            regs = []
            for _ in range(int(count)):
                iv = int(round(random.uniform(lo, hi)))
                iv = max(-32768, min(32767, iv))
                regs.append(iv & 0xFFFF)

            resp = bytearray([unit, fc, len(regs) * 2])
            for r in regs:
                resp.append((r >> 8) & 0xFF)
                resp.append(r & 0xFF)
            crc = crc16_modbus(bytes(resp))
            resp.append(crc & 0xFF)
            resp.append((crc >> 8) & 0xFF)

            self.ser.write(bytes(resp))
            self._log(f'RX_TRIG(modbus): start={start} count={count} -> sent {len(resp)} bytes')
        except Exception as e:
            self._log(f'TRIG_MODBUS_ERR: {e}')

    def _handle_trigger_cmd(self, line: str):
        cmd = (self.tr_cmd.text() or '').strip()
        if not cmd:
            return
        if line.strip() != cmd:
            return
        src = self.tr_source.currentData()
        if src == 'random':
            b = self._build_random_payload()
        else:
            b = self._build_fixed_payload(self.fixed_cont.text(), bool(self.fixed_cont_crlf.isChecked()))
        self._send_bytes(b, prefix='TX_TRIG')

    def _on_rx_timer(self):
        # read bytes from host->sim direction
        try:
            n = int(getattr(self.ser, 'in_waiting', 0) or 0)
        except Exception:
            n = 0
        if n > 0:
            try:
                chunk = self.ser.read(n)
            except Exception:
                chunk = b''
            if chunk:
                self._rx_buf.extend(chunk)
                self._append_log_bytes('RX', chunk)

        if not self.tr_enable.isChecked():
            return

        t = self.tr_type.currentData()
        if t == 'modbus':
            self._handle_trigger_modbus()
        else:
            # 指令模式：按换行解析
            while True:
                idx = self._rx_buf.find(b'\n')
                if idx < 0:
                    break
                line_bytes = bytes(self._rx_buf[:idx])
                del self._rx_buf[:idx+1]
                if line_bytes.endswith(b'\r'):
                    line_bytes = line_bytes[:-1]
                try:
                    line = line_bytes.decode('utf-8', errors='ignore')
                except Exception:
                    line = ''
                if line:
                    self._handle_trigger_cmd(line)


class MainWindow(QMainWindow):

    def __init__(self):
        super().__init__()
        try:
            self.setObjectName('MainWindow')
        except Exception:
            pass
        self.setWindowTitle("Modbus RTU 上位机助手（浅色主题 / 通讯监视）")
        self.resize(1260, 820)

        self.worker: Optional[ModbusRtuWorker] = None
        self.is_connected = False
        self.is_acquiring = False
        self.is_paused = False

        self.channel_names: List[str] = []
        self.curves: Dict[str, pg.PlotDataItem] = {}

        # ---- Plot ring buffer (size = 当前窗口最大点数) ----
        self._buf_size = 100
        self._buf_count = 0
        self._buf_idx = 0  # next write index
        self._ts_buf = None  # numpy array or list
        self._val_buf_by_channel: Dict[str, object] = {}  # name -> np.ndarray or list
        self._plot_x = None  # contiguous x for plotting (numpy)
        self._plot_y_by_channel: Dict[str, object] = {}  # name -> np.ndarray or list
        self._fric_buf = None
        self._mu_buf = None
        self._fric_plot_y = None
        self._mu_plot_y = None
        self._plot_seq = 0

        self._last_plotted_seq = -1
        # Smooth scrolling time base (relative seconds).
        self._t0_mono_ts = None           # time.monotonic() at first sample
        self._last_sample_rel_ts = None   # last sample relative seconds
        self._last_sample_mono_ts = None  # time.monotonic() at last sample

        # Pause-compensated monotonic timeline (so resume does not jump).
        self._mono_pause_accum = 0.0
        self._mono_pause_start = None
        self._settings = QSettings("ModbusAssistant", "ModbusAssistant")
        self._fric_high_name = ""
        self._fric_low_name = ""
        self._wrap_angle_deg = 0.0
        self._wrap_angle_rad = 0.0

        root = QWidget()
        self.setCentralWidget(root)
        layout = QHBoxLayout(root)
        layout.setContentsMargins(0, 0, 0, 0)

        # Use a splitter so left/right sizes stay stable even when widgets change enabled/text states.
        self.main_splitter = QSplitter(Qt.Horizontal)
        layout.addWidget(self.main_splitter, 1)

        left_widget = QWidget()
        left = QVBoxLayout(left_widget)
        left.setContentsMargins(8, 8, 8, 8)


        self.main_splitter.addWidget(left_widget)
        self.main_splitter.setStretchFactor(0, 1)
        try:
            self.main_splitter.setCollapsible(0, False)
        except Exception:
            pass

        # ---- Serial group ----
        serial_box = QGroupBox("串口配置")
        left.addWidget(serial_box)
        sg = QGridLayout(serial_box)

        self.port_combo = QComboBox()
        self.refresh_ports_btn = QPushButton("刷新串口")
        self.refresh_ports_btn.clicked.connect(lambda *_: self.refresh_ports())

        self.baud_combo = QComboBox()
        for b in [9600, 19200, 38400, 57600, 115200, 230400, 460800, 921600]:
            self.baud_combo.addItem(str(b), b)
        self.baud_combo.setCurrentText("38400")  # 默认 9600

        self.parity_combo = QComboBox()
        self.parity_combo.addItems(["N", "E", "O"])

        self.stopbits_combo = QComboBox()
        self.stopbits_combo.addItems(["1", "2"])

        self.bytesize_combo = QComboBox()
        self.bytesize_combo.addItems(["7", "8"])
        self.bytesize_combo.setCurrentText("8")

        self.timeout_spin = QDoubleSpinBox()
        self.timeout_spin.setDecimals(2)
        self.timeout_spin.setRange(0.05, 10.0)
        self.timeout_spin.setSingleStep(0.05)
        self.timeout_spin.setValue(0.5)

        # RS485 direction control
        self.rs485_mode_combo = QComboBox()
        self.rs485_mode_combo.addItems([Rs485CtrlMode.RTS, Rs485CtrlMode.DTR, Rs485CtrlMode.AUTO])
        self.rs485_mode_combo.setCurrentText(Rs485CtrlMode.RTS)  # 默认尝试用RTS解决TX不亮

        self.pre_tx_spin = QSpinBox()
        self.pre_tx_spin.setRange(0, 200)
        self.pre_tx_spin.setValue(0)
        self.pre_tx_spin.setSuffix(" ms")

        self.post_tx_spin = QSpinBox()
        self.post_tx_spin.setRange(0, 200)
        self.post_tx_spin.setValue(2)
        self.post_tx_spin.setSuffix(" ms")

        # Monitor enable for comm monitor (rx/tx)
        self.mon_rx_chk = QCheckBox("监听")
        self.mon_rx_chk.setChecked(True)
        self.mon_tx_chk = QCheckBox("监听")
        self.mon_tx_chk.setChecked(True)

        sg.addWidget(QLabel("接收串口(Modbus)"), 0, 0)
        sg.addWidget(self.port_combo, 0, 1)
        sg.addWidget(self.refresh_ports_btn, 0, 2)

        sg.addWidget(self.mon_rx_chk, 0, 3)

        self.tx_port_combo = QComboBox()
        self.tx_baud_combo = QComboBox()
        for b in [9600, 19200, 38400, 57600, 115200, 230400, 460800, 921600]:
            self.tx_baud_combo.addItem(str(b), b)
        self.tx_baud_combo.setCurrentText("115200")

        self.tx_interval_spin = QSpinBox()
        self.tx_interval_spin.setRange(5, 100000)
        self.tx_interval_spin.setValue(20)  # 默认发送间隔：20ms（115200下更稳妥，且与轮询周期匹配）
        self.tx_interval_spin.setSuffix(" ms")
        self.enable_tx_chk = QCheckBox("启用数据发送串口")
        self.enable_tx_chk.setChecked(True)
        sg.addWidget(QLabel("发送串口(输出)"), 1, 0)
        sg.addWidget(self.tx_port_combo, 1, 1)
        sg.addWidget(self.enable_tx_chk, 1, 2)

        sg.addWidget(self.mon_tx_chk, 1, 3)
        sg.addWidget(QLabel("发送波特率"), 2, 0)
        sg.addWidget(self.tx_baud_combo, 2, 1)
        sg.addWidget(QLabel("发送间隔"), 2, 2)
        sg.addWidget(self.tx_interval_spin, 2, 3)

        sg.addWidget(QLabel("接收波特率"), 3, 0)
        sg.addWidget(self.baud_combo, 3, 1)

        sg.addWidget(QLabel("接收校验"), 4, 0)
        sg.addWidget(self.parity_combo, 4, 1)

        sg.addWidget(QLabel("接收停止位"), 5, 0)
        sg.addWidget(self.stopbits_combo, 5, 1)

        sg.addWidget(QLabel("接收数据位"), 6, 0)
        sg.addWidget(self.bytesize_combo, 6, 1)

        sg.addWidget(QLabel("接收超时(s)"), 7, 0)
        sg.addWidget(self.timeout_spin, 7, 1)

        sg.addWidget(QLabel("RS485方向控制"), 8, 0)
        sg.addWidget(self.rs485_mode_combo, 8, 1)
        sg.addWidget(QLabel("TX前延时"), 9, 0)
        sg.addWidget(self.pre_tx_spin, 9, 1)
        sg.addWidget(QLabel("TX后延时"), 10, 0)
        sg.addWidget(self.post_tx_spin, 10, 1)

        # Connect / Acquire buttons
        btns = QHBoxLayout()
        self.connect_btn = QPushButton("连接串口")
        self.connect_btn.clicked.connect(lambda *_: self.toggle_connect())

        self.acquire_btn = QPushButton("开始采集")
        self.acquire_btn.clicked.connect(lambda *_: self.toggle_acquire())
        self.acquire_btn.setEnabled(False)

        self.pause_btn = QPushButton("暂停")
        self.pause_btn.clicked.connect(lambda *_: self.toggle_pause())
        self.pause_btn.setEnabled(False)

        btns.addWidget(self.connect_btn)
        btns.addWidget(self.acquire_btn)
        btns.addWidget(self.pause_btn)
        left.addLayout(btns)

        # ---- Modbus group ----
        modbus_box = QGroupBox("Modbus 配置")
        left.addWidget(modbus_box)
        mg = QGridLayout(modbus_box)

        self.unit_spin = QSpinBox()
        self.unit_spin.setRange(1, 247)
        self.unit_spin.setValue(1)

        self.func_combo = QComboBox()
        self.func_combo.addItem("03 读保持寄存器", 3)
        self.func_combo.addItem("04 读输入寄存器", 4)

        self.poll_spin = QSpinBox()
        self.poll_spin.setRange(20, 100000)
        self.poll_spin.setValue(20)
        self.poll_spin.setSuffix(" ms")

        self.addr_base1_chk = QCheckBox("地址从 1 开始（自动 -1）")
        self.addr_base1_chk.setChecked(False)  # 你的示例帧 start=0x0000

        mg.addWidget(QLabel("从站地址(站号)"), 0, 0)
        mg.addWidget(self.unit_spin, 0, 1)
        mg.addWidget(QLabel("功能码"), 1, 0)
        mg.addWidget(self.func_combo, 1, 1)
        mg.addWidget(QLabel("轮询周期"), 2, 0)
        mg.addWidget(self.poll_spin, 2, 1)
        mg.addWidget(self.addr_base1_chk, 3, 0, 1, 2)

        # ---- Plot group ----
        plot_box = QGroupBox("绘图设置")
        left.addWidget(plot_box)
        pgd = QGridLayout(plot_box)

        self.max_points_spin = QSpinBox()
        self.max_points_spin.setRange(10, 200000)
        self.max_points_spin.setValue(100)

        self.autoscale_chk = QCheckBox("Y轴自动缩放")
        self.autoscale_chk.setChecked(True)

        # plot update is throttled; mark dirty when settings change
        self.max_points_spin.valueChanged.connect(self._on_max_points_changed)
        self.autoscale_chk.toggled.connect(self._mark_plot_dirty)

        self.plot_fps_spin = QSpinBox()
        self.plot_fps_spin.setRange(1, 240)
        self.plot_fps_spin.setValue(60)
        self.plot_fps_spin.setSuffix(" Hz")
        self.plot_fps_spin.valueChanged.connect(self._on_plot_fps_changed)

        self.clear_btn = QPushButton("清空数据")
        self.clear_btn.clicked.connect(lambda *_: self.clear_data())

        self.save_btn = QPushButton("保存为 XLSX")
        self.save_btn.clicked.connect(lambda *_: self.save_xlsx())

        pgd.addWidget(QLabel("当前窗口最大点数"), 0, 0)
        pgd.addWidget(self.max_points_spin, 0, 1)
        pgd.addWidget(QLabel("绘图刷新率"), 1, 0)
        pgd.addWidget(self.plot_fps_spin, 1, 1)
        pgd.addWidget(self.autoscale_chk, 2, 0, 1, 2)
        pgd.addWidget(self.clear_btn, 3, 0)
        pgd.addWidget(self.save_btn, 3, 1)

        # ---- Channel group ----
        ch_box = QGroupBox("通道配置（多通道）")
        left.addWidget(ch_box, 1)
        cl = QVBoxLayout(ch_box)

        self.ch_table = QTableWidget(0, 7)
        self.ch_table.setHorizontalHeaderLabels([
            "启用", "名称", "地址", "数据类型", "字节序(Word内)", "字顺序(Word间)", "缩放系数"
        ])
        self.ch_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.ch_table.horizontalHeader().setStretchLastSection(True)
        self.ch_table.setAlternatingRowColors(True)
        self.ch_table.itemChanged.connect(lambda *_: self._refresh_friction_channel_options())
        cl.addWidget(self.ch_table)

        btn_row = QHBoxLayout()
        self.add_ch_btn = QPushButton("添加通道")
        self.del_ch_btn = QPushButton("删除选中")
        self.add_ch_btn.clicked.connect(lambda *_: self.add_channel_row())
        self.del_ch_btn.clicked.connect(lambda *_: self.delete_selected_rows())
        btn_row.addWidget(self.add_ch_btn)
        btn_row.addWidget(self.del_ch_btn)
        btn_row.addStretch(1)
        cl.addLayout(btn_row)

        # ---- Plot area ----
        # Use a numeric time axis (seconds) for smooth scrolling.
        # DateAxisItem snaps ticks to "nice" boundaries (often 1s), which can
        # look like the x-axis only moves once per second.
        self.plot = pg.PlotWidget()
        # ---- Plot performance options ----
        # 1) Clip drawing to view to avoid rendering off-screen segments
        # 2) Enable auto-downsampling (peak mode) when data is dense
        # 3) Keep a renderer-side point budget as a safe fallback
        try:
            pi = self.plot.getPlotItem()
            try:
                pi.setClipToView(True)
            except Exception:
                pass
            try:
                # Different pyqtgraph versions use "mode" or "method"
                pi.setDownsampling(auto=True, mode="peak")
            except Exception:
                try:
                    pi.setDownsampling(auto=True, method="peak")
                except Exception:
                    pass
        except Exception:
            pass
        # Fallback point budget sent to renderer per curve (0=disable)
        self._max_display_points = 6000

        self.plot.setLabel("bottom", "时间", units="s")
        self.plot.setLabel("left", "张力")
        self.plot.addLegend()
        self.plot.showGrid(x=True, y=True, alpha=0.25)

        # ---- Friction force plot ----
        self.friction_plot = pg.PlotWidget()
        try:
            pi = self.friction_plot.getPlotItem()
            try:
                pi.setClipToView(True)
            except Exception:
                pass
            try:
                pi.setDownsampling(auto=True, mode="peak")
            except Exception:
                try:
                    pi.setDownsampling(auto=True, method="peak")
                except Exception:
                    pass
        except Exception:
            pass
        self.friction_plot.setLabel("bottom", "时间", units="s")
        self.friction_plot.setLabel("left", "摩擦力")
        self.friction_plot.addLegend()
        self.friction_plot.showGrid(x=True, y=True, alpha=0.25)
        self.friction_curve = self.friction_plot.plot([], [], name="摩擦力", pen=pg.mkPen(color=(255, 140, 0), width=2))

        # ---- Friction coefficient plot ----
        self.mu_plot = pg.PlotWidget()
        try:
            pi = self.mu_plot.getPlotItem()
            try:
                pi.setClipToView(True)
            except Exception:
                pass
            try:
                pi.setDownsampling(auto=True, mode="peak")
            except Exception:
                try:
                    pi.setDownsampling(auto=True, method="peak")
                except Exception:
                    pass
        except Exception:
            pass
        self.mu_plot.setLabel("bottom", "时间", units="s")
        self.mu_plot.setLabel("left", "摩擦系数")
        self.mu_plot.addLegend()
        self.mu_plot.showGrid(x=True, y=True, alpha=0.25)
        self.mu_curve = self.mu_plot.plot([], [], name="摩擦系数", pen=pg.mkPen(color=(0, 120, 220), width=2))

        # ---- Plot window (dock) ----
        self.plot_tabs = QTabWidget()
        tension_tab = QWidget()
        t_layout = QVBoxLayout(tension_tab)
        t_layout.setContentsMargins(0, 0, 0, 0)
        t_layout.addWidget(self.plot, 1)
        self.status_label = QLabel("状态：未连接")
        t_layout.addWidget(self.status_label)
        self.plot_tabs.addTab(tension_tab, "张力")

        self.friction_tab = QWidget()
        f_layout = QVBoxLayout(self.friction_tab)
        f_layout.setContentsMargins(0, 0, 0, 0)
        cfg = QGridLayout()
        self.fric_high_combo = QComboBox()
        self.fric_low_combo = QComboBox()
        self.fric_swap_btn = QPushButton("互换")
        self.wrap_angle_spin = QDoubleSpinBox()
        self.wrap_angle_spin.setDecimals(2)
        self.wrap_angle_spin.setRange(0.0, 360.0)
        self.wrap_angle_spin.setSingleStep(1.0)
        self.wrap_angle_spin.setValue(0.0)
        self.wrap_angle_spin.setSuffix(" °")
        cfg.addWidget(QLabel("高张力侧"), 0, 0)
        cfg.addWidget(self.fric_high_combo, 0, 1)
        cfg.addWidget(QLabel("低张力侧"), 0, 2)
        cfg.addWidget(self.fric_low_combo, 0, 3)
        cfg.addWidget(self.fric_swap_btn, 0, 4)
        cfg.addWidget(QLabel("包角"), 1, 0)
        cfg.addWidget(self.wrap_angle_spin, 1, 1)
        cfg.setColumnStretch(5, 1)
        f_layout.addLayout(cfg)
        f_layout.addWidget(self.friction_plot, 1)

        self.mu_tab = QWidget()
        mu_layout = QVBoxLayout(self.mu_tab)
        mu_layout.setContentsMargins(0, 0, 0, 0)
        mu_layout.addWidget(self.mu_plot, 1)

        self.plot_dock = QDockWidget("绘图窗口", self)
        try:
            self.plot_dock.setObjectName("dock_plot")
        except Exception:
            pass
        self.plot_dock.setAllowedAreas(Qt.BottomDockWidgetArea | Qt.TopDockWidgetArea | Qt.LeftDockWidgetArea | Qt.RightDockWidgetArea)
        plot_container = QWidget()
        plot_layout = QVBoxLayout(plot_container)
        plot_layout.setContentsMargins(6, 6, 6, 6)
        plot_layout.addWidget(self.plot_tabs, 1)
        self.plot_dock.setWidget(plot_container)
        self.addDockWidget(Qt.RightDockWidgetArea, self.plot_dock)

        self.fric_high_combo.currentIndexChanged.connect(self._on_friction_config_changed)
        self.fric_low_combo.currentIndexChanged.connect(self._on_friction_config_changed)
        self.wrap_angle_spin.valueChanged.connect(self._on_friction_config_changed)
        self.fric_swap_btn.clicked.connect(self._swap_friction_channels)

        # ---- Comm Monitor Dock ----
        self.monitor_dock = QDockWidget("通讯监视窗口", self)
        try:
            self.monitor_dock.setObjectName('dock_monitor')
        except Exception:
            pass
        self.monitor_dock.setAllowedAreas(Qt.BottomDockWidgetArea | Qt.TopDockWidgetArea)
        self.monitor_text = QPlainTextEdit()
        self.monitor_text.setReadOnly(True)
        self.monitor_text.setUndoRedoEnabled(False)
        try:
            self.monitor_text.setMaximumBlockCount(2000)
        except Exception:
            pass

        mon_container = QWidget()
        mon_layout = QVBoxLayout(mon_container)
        mon_layout.setContentsMargins(6, 6, 6, 6)
        # display mode
        mon_mode_row = QHBoxLayout()
        mon_mode_row.addWidget(QLabel("显示模式"))
        self.monitor_mode_combo = QComboBox()
        self.monitor_mode_combo.addItem("HEX", "hex")
        self.monitor_mode_combo.addItem("文本(UTF-8)", "utf-8")
        self.monitor_mode_combo.addItem("文本(GBK)", "gbk")
        self.monitor_mode_combo.setCurrentIndex(1)
        self.monitor_mode_combo.currentIndexChanged.connect(lambda *_: self.schedule_monitor_render(full=True))
        mon_mode_row.addWidget(self.monitor_mode_combo)
        mon_mode_row.addStretch(1)
        mon_layout.addLayout(mon_mode_row)

        mon_layout.addWidget(self.monitor_text, 1)

        mon_btns = QHBoxLayout()
        self.mon_clear_btn = QPushButton("清空监视")
        self.mon_clear_btn.clicked.connect(lambda *_: self.clear_monitor())
        self.mon_save_btn = QPushButton("保存日志")
        self.mon_save_btn.clicked.connect(lambda *_: self.save_monitor_log())
        mon_btns.addWidget(self.mon_clear_btn)
        mon_btns.addWidget(self.mon_save_btn)
        mon_btns.addStretch(1)
        mon_layout.addLayout(mon_btns)

        self.monitor_dock.setWidget(mon_container)
        self.addDockWidget(Qt.BottomDockWidgetArea, self.monitor_dock)
        self.monitor_dock.visibilityChanged.connect(lambda vis: vis and (self.schedule_monitor_render(full=True) or True))

        # structured log storage for re-rendering in HEX/TEXT modes
        self._monitor_entries: List[dict] = []
        self._manual_entries: List[dict] = []
        self._motor_mon_entries: List[dict] = []

        # ---- UI throttle timers (avoid stutter at high frame rates) ----
        self._monitor_dirty = False
        self._manual_dirty = False
        self._monitor_timer = QTimer(self)
        self._monitor_timer.setInterval(50)
        self._monitor_timer.setSingleShot(True)
        self._monitor_timer.timeout.connect(self._flush_monitor_render)

        self._manual_timer = QTimer(self)
        self._manual_timer.setInterval(50)
        self._manual_timer.setSingleShot(True)
        self._manual_timer.timeout.connect(self._flush_manual_render)

        self._motor_mon_dirty = False
        self._motor_mon_timer = QTimer(self)
        self._motor_mon_timer.setInterval(50)
        self._motor_mon_timer.setSingleShot(True)
        self._motor_mon_timer.timeout.connect(self._flush_motor_monitor_render)

        # ---- Plot throttle (refresh-rate driven) ----
        self._plot_dirty = False
        self._plot_timer = QTimer(self)
        self._plot_timer.setInterval(16)  # will be updated by _on_plot_fps_changed()
        self._plot_timer.timeout.connect(self._flush_plot)
        # NOTE: plot timer starts only while acquiring (see _update_plot_timer_running)
        self._on_plot_fps_changed()

        self._last_xrange_update = 0.0
        self._last_yrange_update = 0.0

        # ---- Custom Send Dock (default hidden, show on right) ----
        self.custom_send_dock = QDockWidget("自定义串口发送", self)
        try:
            self.custom_send_dock.setObjectName('dock_custom_send')
        except Exception:
            pass
        self.custom_send_dock.setAllowedAreas(Qt.RightDockWidgetArea | Qt.LeftDockWidgetArea)

        send_container = QWidget()
        send_layout = QVBoxLayout(send_container)
        send_layout.setContentsMargins(8, 8, 8, 8)

        send_grid = QGridLayout()
        self.custom_send_port_combo = QComboBox()
        self.custom_send_line = QLineEdit()
        self.custom_send_line.setPlaceholderText("输入要发送的内容（文本或HEX：01 03 00 00 00 02）")
        self.custom_send_crlf_chk = QCheckBox("末尾添加 \\r\\n")
        self.custom_send_crlf_chk.setChecked(True)
        self.custom_send_btn = QPushButton("发送")

        send_grid.addWidget(QLabel("发送串口"), 0, 0)
        send_grid.addWidget(self.custom_send_port_combo, 0, 1)
        send_grid.addWidget(QLabel("发送内容"), 1, 0)
        send_grid.addWidget(self.custom_send_line, 1, 1)
        send_layout.addLayout(send_grid)

        # display mode
        mode_row = QHBoxLayout()
        mode_row.addWidget(QLabel("日志显示"))
        self.custom_send_mode_combo = QComboBox()
        self.custom_send_mode_combo.addItem("HEX", "hex")
        self.custom_send_mode_combo.addItem("文本(UTF-8)", "utf-8")
        self.custom_send_mode_combo.addItem("文本(GBK)", "gbk")
        self.custom_send_mode_combo.setCurrentIndex(1)
        self.custom_send_mode_combo.currentIndexChanged.connect(lambda *_: self.schedule_custom_send_render(full=True))
        mode_row.addWidget(self.custom_send_mode_combo)
        mode_row.addStretch(1)
        send_layout.addLayout(mode_row)

        send_row = QHBoxLayout()
        send_row.addWidget(self.custom_send_crlf_chk)
        send_row.addStretch(1)
        send_row.addWidget(self.custom_send_btn)
        send_layout.addLayout(send_row)

        tip = QLabel("提示：Enter 直接发送；可输入文本，或输入十六进制字节（如：01 03 00 00 00 02 / 010300000002）。")
        tip.setWordWrap(True)
        send_layout.addWidget(tip)

        # reply/listen area (filtered manual TX/RX for the selected opened port)
        send_layout.addWidget(QLabel("回复 / 日志"))
        self.custom_send_log = QPlainTextEdit()
        self.custom_send_log.setReadOnly(True)
        self.custom_send_log.setUndoRedoEnabled(False)
        try:
            self.custom_send_log.setMaximumBlockCount(1200)
        except Exception:
            pass
        self.custom_send_log.setPlaceholderText("这里会显示自定义发送对应的 TX/RX。")
        send_layout.addWidget(self.custom_send_log, 1)

        log_btns = QHBoxLayout()
        self.custom_send_clear_btn = QPushButton("清空")
        self.custom_send_clear_btn.clicked.connect(lambda *_: self.clear_custom_send_log())
        log_btns.addWidget(self.custom_send_clear_btn)
        log_btns.addStretch(1)
        send_layout.addLayout(log_btns)

        send_layout.addStretch(1)

        self.custom_send_dock.setWidget(send_container)
        self.addDockWidget(Qt.RightDockWidgetArea, self.custom_send_dock)
        self.custom_send_dock.hide()  # default closed

        # ---- Motor Control Dock (default hidden) ----
        self.motor_dock = QDockWidget("电机控制", self)
        try:
            self.motor_dock.setObjectName('dock_motor')
        except Exception:
            pass
        self.motor_dock.setAllowedAreas(Qt.RightDockWidgetArea | Qt.LeftDockWidgetArea)

        motor_container = QWidget()
        motor_layout = QVBoxLayout(motor_container)
        motor_layout.setContentsMargins(8, 8, 8, 8)

        # Enable/Disable
        en_row = QHBoxLayout()
        self.motor_enable_btn = QPushButton("使能")
        self.motor_disable_btn = QPushButton("去使能")
        self.motor_enable_lamp = QLabel()
        self.motor_enable_lamp.setFixedSize(14, 14)
        self._set_lamp_color(self.motor_enable_lamp, "#777777")
        en_row.addWidget(self.motor_enable_btn)
        en_row.addWidget(self.motor_disable_btn)
        en_row.addWidget(QLabel("使能灯"))
        en_row.addWidget(self.motor_enable_lamp)
        en_row.addStretch(1)
        motor_layout.addLayout(en_row)

        # Direction
        dir_row = QHBoxLayout()
        self.motor_forward_btn = QPushButton("正转")
        self.motor_backward_btn = QPushButton("反转")
        self.motor_dir_lamp = QLabel()
        self.motor_dir_lamp.setFixedSize(14, 14)
        self._set_lamp_color(self.motor_dir_lamp, "#777777")
        dir_row.addWidget(self.motor_forward_btn)
        dir_row.addWidget(self.motor_backward_btn)
        dir_row.addWidget(QLabel("转向灯"))
        dir_row.addWidget(self.motor_dir_lamp)
        dir_row.addStretch(1)
        motor_layout.addLayout(dir_row)
        # Mode selection
        mode_row = QHBoxLayout()
        self.motor_mode_tension_btn = QPushButton("张力模式")
        self.motor_mode_speed_btn = QPushButton("速度模式")
        self.motor_mode_tension_lamp = QLabel()
        self.motor_mode_tension_lamp.setFixedSize(14, 14)
        self._set_lamp_color(self.motor_mode_tension_lamp, "#777777")
        self.motor_mode_speed_lamp = QLabel()
        self.motor_mode_speed_lamp.setFixedSize(14, 14)
        self._set_lamp_color(self.motor_mode_speed_lamp, "#777777")
        mode_row.addWidget(self.motor_mode_tension_btn)
        mode_row.addWidget(self.motor_mode_speed_btn)
        mode_row.addWidget(QLabel("张力模式灯"))
        mode_row.addWidget(self.motor_mode_tension_lamp)
        mode_row.addWidget(QLabel("速度模式灯"))
        mode_row.addWidget(self.motor_mode_speed_lamp)
        mode_row.addStretch(1)
        motor_layout.addLayout(mode_row)



        # Speed control
        speed_row = QHBoxLayout()
        speed_row.addWidget(QLabel("转速(RPM)"))
        self.motor_speed_edit = QLineEdit()
        self.motor_speed_edit.setPlaceholderText("例如 100")
        self.motor_speed_btn = QPushButton("转速控制")
        speed_row.addWidget(self.motor_speed_edit, 1)
        speed_row.addWidget(self.motor_speed_btn)
        motor_layout.addLayout(speed_row)

        # Tension control
        tension_row = QHBoxLayout()
        tension_row.addWidget(QLabel("张力(g)"))
        self.motor_tension_edit = QLineEdit()
        self.motor_tension_edit.setPlaceholderText("例如 1")
        self.motor_tension_btn = QPushButton("张力控制")
        tension_row.addWidget(self.motor_tension_edit, 1)
        tension_row.addWidget(self.motor_tension_btn)
        motor_layout.addLayout(tension_row)

        # PID control
        pid_row = QHBoxLayout()
        pid_row.addWidget(QLabel("Kp"))
        self.motor_kp_edit = QLineEdit()
        self.motor_kp_edit.setPlaceholderText("Kp:例如0.1")
        pid_row.addWidget(self.motor_kp_edit)
        pid_row.addWidget(QLabel("Ki"))
        self.motor_ki_edit = QLineEdit()
        self.motor_ki_edit.setPlaceholderText("Ki:例如0.02")
        pid_row.addWidget(self.motor_ki_edit)
        pid_row.addWidget(QLabel("Kd"))
        self.motor_kd_edit = QLineEdit()
        self.motor_kd_edit.setPlaceholderText("Kd:例如0.005")
        pid_row.addWidget(self.motor_kd_edit)
        self.motor_pid_btn = QPushButton("PID设置")
        pid_row.addWidget(self.motor_pid_btn)
        motor_layout.addLayout(pid_row)

        # Emergency stop
        self.motor_estop_btn = QPushButton("急停")
        try:
            self.motor_estop_btn.setStyleSheet("background:#d9534f;color:white;font-weight:bold;")
        except Exception:
            pass
        try:
            self.motor_estop_btn.setFixedSize(160, 120)
        except Exception:
            pass
        estop_row = QHBoxLayout()
        estop_row.addStretch(1)
        estop_row.addWidget(self.motor_estop_btn)
        estop_row.addStretch(1)
        motor_layout.addLayout(estop_row)

        # TX monitor for motor control
        motor_mon_group = QGroupBox("发送串口监视")
        motor_mon_layout = QVBoxLayout(motor_mon_group)
        motor_mon_layout.setContentsMargins(6, 6, 6, 6)
        motor_mon_mode_row = QHBoxLayout()
        motor_mon_mode_row.addWidget(QLabel("显示方式"))
        self.motor_mon_mode_combo = QComboBox()
        self.motor_mon_mode_combo.addItem("文本(UTF-8)", "utf-8")
        self.motor_mon_mode_combo.addItem("HEX", "hex")
        self.motor_mon_mode_combo.addItem("文本(GBK)", "gbk")
        self.motor_mon_mode_combo.setCurrentIndex(0)
        self.motor_mon_mode_combo.currentIndexChanged.connect(lambda *_: self.schedule_motor_monitor_render(full=True))
        motor_mon_mode_row.addWidget(self.motor_mon_mode_combo)
        motor_mon_mode_row.addStretch(1)
        self.motor_mon_clear_btn = QPushButton("清空")
        self.motor_mon_clear_btn.clicked.connect(self.clear_motor_monitor)
        motor_mon_mode_row.addWidget(self.motor_mon_clear_btn)
        motor_mon_layout.addLayout(motor_mon_mode_row)
        self.motor_mon_text = QPlainTextEdit()
        self.motor_mon_text.setReadOnly(True)
        self.motor_mon_text.setUndoRedoEnabled(False)
        try:
            self.motor_mon_text.setMaximumBlockCount(2000)
        except Exception:
            pass
        self.motor_mon_text.setPlaceholderText("这里显示发送串口返回的数据。")
        motor_mon_layout.addWidget(self.motor_mon_text, 1)
        motor_layout.addWidget(motor_mon_group, 1)


        self.motor_dock.setWidget(motor_container)
        self.addDockWidget(Qt.RightDockWidgetArea, self.motor_dock)
        self.motor_dock.hide()  # default hidden
        self.motor_dock.visibilityChanged.connect(lambda vis: vis and (self.schedule_motor_monitor_render(full=True) or True))

        self.motor_enable_btn.clicked.connect(lambda *_: self.on_motor_enable())
        self.motor_disable_btn.clicked.connect(lambda *_: self.on_motor_disable())
        self.motor_forward_btn.clicked.connect(lambda *_: self.on_motor_forward())
        self.motor_backward_btn.clicked.connect(lambda *_: self.on_motor_backward())
        self.motor_speed_btn.clicked.connect(lambda *_: self.on_motor_speed())
        self.motor_tension_btn.clicked.connect(lambda *_: self.on_motor_tension())
        self.motor_pid_btn.clicked.connect(lambda *_: self.on_motor_pid())
        self.motor_estop_btn.clicked.connect(lambda *_: self.on_motor_estop())
        self.motor_mode_tension_btn.clicked.connect(lambda *_: self.on_motor_mode_tension())
        self.motor_mode_speed_btn.clicked.connect(lambda *_: self.on_motor_mode_speed())
        self.motor_mode = None

        self.custom_send_btn.clicked.connect(lambda *_: self.send_custom_serial())
        self.custom_send_line.returnPressed.connect(self.send_custom_serial)
        self.custom_send_dock.visibilityChanged.connect(lambda vis: vis and (self.update_custom_send_ports() or True) and (self.schedule_custom_send_render(full=True) or True))
        self.custom_send_port_combo.currentIndexChanged.connect(lambda *_: self.schedule_custom_send_render(full=True))

        # ---- Workspace menu (show/hide panels) ----
        ws_menu = self.menuBar().addMenu("工作区")

        act_serial = ws_menu.addAction("串口配置")
        act_serial.setCheckable(True)
        act_serial.setChecked(True)
        act_serial.toggled.connect(serial_box.setVisible)

        act_modbus = ws_menu.addAction("Modbus 配置")
        act_modbus.setCheckable(True)
        act_modbus.setChecked(True)
        act_modbus.toggled.connect(modbus_box.setVisible)

        act_plotset = ws_menu.addAction("绘图设置")
        act_plotset.setCheckable(True)
        act_plotset.setChecked(True)
        act_plotset.toggled.connect(plot_box.setVisible)

        act_channels = ws_menu.addAction("通道配置")
        act_channels.setCheckable(True)
        act_channels.setChecked(True)
        act_channels.toggled.connect(ch_box.setVisible)

        ws_menu.addSeparator()
        # Dock 自带的切换 Action
        ws_menu.addAction(self.monitor_dock.toggleViewAction())
        ws_menu.addAction(self.custom_send_dock.toggleViewAction())

        # Monitor enable switches
        self.mon_rx_chk.toggled.connect(self.schedule_monitor_render)
        self.mon_tx_chk.toggled.connect(self.on_tx_monitor_toggled)

        # ---- Control menu ----
        ctrl_menu = self.menuBar().addMenu("控制")
        act_motor = ctrl_menu.addAction("电机控制")
        act_motor.triggered.connect(self.open_motor_control)


        # ---- Plot window menu ----
        plot_menu = self.menuBar().addMenu("绘图窗口")
        self.act_plot_window = plot_menu.addAction("显示绘图窗口")
        self.act_plot_window.setCheckable(True)
        self.act_plot_window.setChecked(True)
        self.act_plot_window.toggled.connect(lambda on: self.plot_dock.setVisible(on))
        self.plot_dock.visibilityChanged.connect(lambda vis: self.act_plot_window.setChecked(vis))

        self.act_friction_plot = plot_menu.addAction("摩擦力绘图窗口")
        self.act_friction_plot.setCheckable(True)
        self.act_friction_plot.setChecked(False)
        self.act_friction_plot.toggled.connect(lambda on: self._set_plot_tab_visible(self.friction_tab, "摩擦力", on))

        self.act_mu_plot = plot_menu.addAction("摩擦系数绘图窗口")
        self.act_mu_plot.setCheckable(True)
        self.act_mu_plot.setChecked(False)
        self.act_mu_plot.toggled.connect(lambda on: self._set_plot_tab_visible(self.mu_tab, "摩擦系数", on))

        # ---- Serial simulator menu ----
        sim_menu = self.menuBar().addMenu('串口仿真')
        act_sim = sim_menu.addAction('打开仿真串口界面')
        self.sim_manager = SerialSimManagerWindow(self)
        self.sim_manager.hide()
        act_sim.triggered.connect(self.open_simulator)
        self.sim_manager.ports_changed.connect(self.refresh_ports)


        # Init
        self.refresh_ports()
        # 默认两通道：两个 int16（01 03 00 00 00 02 ...）
        self.add_channel_row(default_name="CH1", default_addr=0, default_dtype="int16")
        self.add_channel_row(default_name="CH2", default_addr=1, default_dtype="int16")
        self._refresh_friction_channel_options()

        # Keep layout stable (no left-panel width jitter) and restore last workspace state.
        self._apply_stable_widget_sizing()

        # Restore last workspace state only after the window is actually shown.
        # Doing it too early (during __init__) may lead to incomplete first layout on some systems.
        self._restored_once = False

    def showEvent(self, e):
        super().showEvent(e)
        if getattr(self, "_restored_once", False):
            return
        self._restored_once = True
        QTimer.singleShot(0, self._restore_after_show)
    def _restore_after_show(self):
        self._restore_window_layout()

        # Fix: saved window position may be partially off-screen (negative Y, etc.)
        self._ensure_frame_on_screen()
        QTimer.singleShot(50, self._ensure_frame_on_screen)
        QTimer.singleShot(180, self._ensure_frame_on_screen)

        # Force a couple of layout/paint passes.
        # On some systems (Windows + high DPI, and/or OpenGL-backed widgets), the very first
        # frame may not fully lay out/paint until a resizeEvent happens (e.g. user drags border).
        # We emulate that once, without changing the visible size.
        self._force_first_layout_pass()
        QTimer.singleShot(0, self._force_first_layout_pass)

    def _force_first_layout_pass(self):
        # Activate central layout
        try:
            cw = self.centralWidget()
            if cw:
                try:
                    cw.updateGeometry()
                except Exception:
                    pass
                if cw.layout():
                    cw.layout().activate()
        except Exception:
            pass

        # Stabilize splitter geometry
        try:
            if hasattr(self, "main_splitter") and self.main_splitter is not None:
                try:
                    self.main_splitter.updateGeometry()
                except Exception:
                    pass
                # Re-apply current sizes to force internal recompute
                self.main_splitter.setSizes(self.main_splitter.sizes())
                # Guard against a bad restored splitter state (e.g. one side nearly collapsed).
                try:
                    sizes = self.main_splitter.sizes()
                    if isinstance(sizes, (list, tuple)) and len(sizes) >= 2:
                        total = int(sizes[0]) + int(sizes[1])
                        if total > 0 and (int(sizes[0]) < 180 or int(sizes[1]) < 180):
                            left = max(360, int(total * 0.35))
                            right = max(420, total - left)
                            self.main_splitter.setSizes([left, right])
                except Exception:
                    pass

        except Exception:
            pass

        # Flush pending events once (first-show only)
        try:
            QApplication.processEvents()
        except Exception:
            pass

        # Nudge window size (triggers resizeEvent like a manual border drag)
        try:
            is_max = getattr(self, "isMaximized", lambda: False)()
            is_full = getattr(self, "isFullScreen", lambda: False)()
            if not is_max and not is_full:
                w, h = int(self.width()), int(self.height())
                self.resize(w + 1, h + 1)
                self.resize(w, h)
        except Exception:
            pass

        # Help plotting widget settle (pyqtgraph / OpenGL)
        try:
            if hasattr(self, "plot") and self.plot is not None:
                try:
                    self.plot.updateGeometry()
                except Exception:
                    pass
                self.plot.update()

                try:
                    self.plot.repaint()
                except Exception:
                    pass
        except Exception:
            pass

        try:
            self.update()
            self.repaint()
        except Exception:
            pass

        # Ensure the restored window frame is fully visible on the current screen
        # (fixes cases where QSettings restored a negative Y, etc.)
        self._ensure_frame_on_screen()

    def _ensure_frame_on_screen(self):
        """Keep the window frame inside the current screen's available area.

        This fixes cases where a previously saved QSettings window position is restored with
        a negative Y (top clipped) or otherwise partially off-screen. We clamp using the
        *frameGeometry* (including title bar) rather than the client geometry.
        """
        try:
            # Don't interfere with maximized/fullscreen states
            try:
                if getattr(self, "isMaximized", lambda: False)() or getattr(self, "isFullScreen", lambda: False)():
                    return
            except Exception:
                pass

            screen = None
            try:
                screen = self.screen()
            except Exception:
                screen = None
            if screen is None:
                try:
                    screen = QApplication.primaryScreen()
                except Exception:
                    screen = None
            if screen is None:
                return

            avail = screen.availableGeometry()
            fg = self.frameGeometry()

            x, y = int(fg.x()), int(fg.y())
            w, h = int(fg.width()), int(fg.height())

            # If it's totally outside, re-center
            if (x + w) < (avail.left() + 20) or x > (avail.left() + avail.width() - 20) or (y + h) < (avail.top() + 20) or y > (avail.top() + avail.height() - 20):
                new_x = int(avail.left() + max(0, (avail.width() - w) // 2))
                new_y = int(avail.top() + max(0, (avail.height() - h) // 2))
                self.move(new_x, new_y)
                return

            # Clamp into available rect
            max_x = int(avail.left() + max(0, avail.width() - w))
            max_y = int(avail.top() + max(0, avail.height() - h))
            new_x = min(max(x, int(avail.left())), max_x)
            new_y = min(max(y, int(avail.top())), max_y)

            if new_x != x or new_y != y:
                self.move(new_x, new_y)
        except Exception:
            pass

    def open_simulator(self):
        """打开串口仿真界面（可同时仿真多个串口）。"""
        if not hasattr(self, 'sim_manager') or self.sim_manager is None:
            self.sim_manager = SerialSimManagerWindow(self)
        self.sim_manager.show()
        try:
            self.sim_manager.raise_()
            self.sim_manager.activateWindow()
        except Exception:
            pass


    def open_motor_control(self):
        if not hasattr(self, 'motor_dock') or self.motor_dock is None:
            return
        self.motor_dock.show()
        try:
            self.motor_dock.raise_()
            self.motor_dock.activateWindow()
        except Exception:
            pass

    def _apply_stable_widget_sizing(self):
        """Prevent left-panel width jitter caused by long/short combo texts.

        This keeps the current workspace layout stable when clicking connect/refresh/etc.
        """
        combos = [
            getattr(self, 'port_combo', None),
            getattr(self, 'tx_port_combo', None),
            getattr(self, 'custom_send_port_combo', None),
        ]
        for cb in combos:
            if cb is None:
                continue
            # Prefer policy that does NOT resize to current text.
            pol = None
            try:
                pol = QComboBox.SizeAdjustPolicy.AdjustToMinimumContentsLengthWithIcon
            except Exception:
                pol = getattr(QComboBox, 'AdjustToMinimumContentsLengthWithIcon', None)
                if pol is None:
                    pol = getattr(QComboBox, 'AdjustToMinimumContentsLength', None)
            try:
                if pol is not None:
                    cb.setSizeAdjustPolicy(pol)
            except Exception:
                pass
            try:
                cb.setMinimumContentsLength(28)
            except Exception:
                pass
            try:
                cb.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            except Exception:
                pass

        try:
            self.status_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        except Exception:
            pass

    def _apply_safe_geometry(self, x, y, w, h):
        """Clamp geometry to current screen's available area.

        This avoids Qt warnings like:
        QWindowsWindow::setGeometry: Unable to set geometry ...
        when restoring a window size larger than the current available screen
        area (taskbar, DPI scaling, monitor change, etc.).
        """
        try:
            x = int(x); y = int(y); w = int(w); h = int(h)
        except Exception:
            return

        # If maximized/fullscreen, don't fight window manager.
        try:
            if getattr(self, 'isMaximized', lambda: False)() or getattr(self, 'isFullScreen', lambda: False)():
                return
        except Exception:
            pass

        try:
            cx = x + max(0, w // 2)
            cy = y + max(0, h // 2)
            try:
                center = QPoint(cx, cy)
            except Exception:
                center = None

            screen = None
            try:
                if center is not None:
                    screen = QGuiApplication.screenAt(center)
            except Exception:
                screen = None
            if screen is None:
                try:
                    screen = QGuiApplication.primaryScreen()
                except Exception:
                    screen = None

            if screen is None:
                try:
                    self.setGeometry(x, y, w, h)
                except Exception:
                    pass
                return

            avail = screen.availableGeometry()

            try:
                minw = int(self.minimumWidth()) if int(self.minimumWidth()) > 0 else 640
            except Exception:
                minw = 640
            try:
                minh = int(self.minimumHeight()) if int(self.minimumHeight()) > 0 else 480
            except Exception:
                minh = 480

            w = max(minw, min(w, int(avail.width())))
            h = max(minh, min(h, int(avail.height())))

            x = min(max(x, int(avail.left())), int(avail.right()) - w + 1)
            y = min(max(y, int(avail.top())), int(avail.bottom()) - h + 1)

            self.setGeometry(x, y, w, h)
        except Exception:
            try:
                self.setGeometry(x, y, w, h)
            except Exception:
                pass

    def _ensure_window_on_screen(self):
        """Clamp current window geometry to be visible on some screen."""
        try:
            if getattr(self, 'isMaximized', lambda: False)() or getattr(self, 'isFullScreen', lambda: False)():
                return
        except Exception:
            pass
        try:
            g = self.geometry()
            self._apply_safe_geometry(g.x(), g.y(), g.width(), g.height())
        except Exception:
            pass

    def _save_window_layout(self):
        s = getattr(self, '_settings', None)
        if s is None:
            return
        try:
            s.setValue('main/geometry', self.saveGeometry())
            s.setValue('main/state', self.saveState())
            try:
                s.setValue('main/wstate', int(self.windowState()))
            except Exception:
                pass
            try:
                g = self.normalGeometry() if getattr(self, 'isMaximized', lambda: False)() else self.geometry()
                s.setValue('main/rect', [int(g.x()), int(g.y()), int(g.width()), int(g.height())])
            except Exception:
                pass
            if hasattr(self, 'main_splitter'):
                s.setValue('main/splitter', self.main_splitter.saveState())
        except Exception:
            pass

    def _restore_window_layout(self):
        s = getattr(self, '_settings', None)
        if s is None:
            return
        had_split = False
        try:
            had_split = bool(s.value('main/splitter'))
        except Exception:
            had_split = False

        # --- Restore main window rect (preferred) ---
        rect = None
        try:
            rect = s.value('main/rect')
        except Exception:
            rect = None
        if rect and isinstance(rect, (list, tuple)) and len(rect) >= 4:
            try:
                x, y, w, h = rect[:4]
                self._apply_safe_geometry(x, y, w, h)
            except Exception:
                pass
        else:
            # Backward-compat: fall back to Qt's saveGeometry/restoreGeometry
            try:
                geom = s.value('main/geometry')
                if geom:
                    self.restoreGeometry(geom)
            except Exception:
                pass
            # Clamp in case the restored geometry doesn't fit current screen/DPI
            self._ensure_window_on_screen()

        # --- Restore docks/tool state ---
        try:
            state = s.value('main/state')
            if state:
                self.restoreState(state)
        except Exception:
            pass

        # --- Restore splitter sizes ---
        try:
            sp = s.value('main/splitter')
            if sp and hasattr(self, 'main_splitter'):
                self.main_splitter.restoreState(sp)
        except Exception:
            pass

        # Restore window maximized state (after geometry)
        try:
            ws = s.value('main/wstate')
            ws_i = int(ws) if ws is not None else 0
            if ws_i & int(Qt.WindowMaximized):
                self.setWindowState(self.windowState() | Qt.WindowMaximized)
        except Exception:
            pass

        # Clamp again after restoreState (dock/min-size may change)
        self._ensure_window_on_screen()

        # First run fallback: give the left panel a reasonable width.
        if not had_split:
            try:
                if hasattr(self, 'main_splitter'):
                    try:
                        if int(self.main_splitter.count()) >= 2:
                            self.main_splitter.setSizes([420, 1000])
                        else:
                            self.main_splitter.setSizes([1000])
                    except Exception:
                        pass
            except Exception:
                pass


    # ---------- UI throttling helpers ----------
    def schedule_monitor_render(self, full: bool = False):
        """Throttle comm-monitor UI updates.

        full=True forces a full re-render (e.g. display mode changed / panel shown).
        """
        if full:
            setattr(self, "_monitor_force_full", True)
        self._monitor_dirty = True
        try:
            if not self._monitor_timer.isActive():
                self._monitor_timer.start()
        except Exception:
            self.render_monitor(force_full=bool(full))

    def schedule_custom_send_render(self, full: bool = False):
        if full:
            setattr(self, "_manual_force_full", True)
        self._manual_dirty = True
        try:
            if not self._manual_timer.isActive():
                self._manual_timer.start()
        except Exception:
            self.render_custom_send_log(force_full=bool(full))


    def schedule_motor_monitor_render(self, full: bool = False):
        if full:
            setattr(self, "_motor_mon_force_full", True)
        self._motor_mon_dirty = True
        try:
            if not self._motor_mon_timer.isActive():
                self._motor_mon_timer.start()
        except Exception:
            self.render_motor_monitor(force_full=bool(full))
    def _flush_monitor_render(self):
        if not getattr(self, "_monitor_dirty", False):
            return
        # if panel hidden, defer render to when it becomes visible
        if hasattr(self, "monitor_dock") and not self.monitor_dock.isVisible():
            self._monitor_dirty = False
            self._monitor_force_full = True
            return

        self._monitor_dirty = False
        force_full = bool(getattr(self, "_monitor_force_full", False))
        try:
            self.render_monitor(force_full=force_full)
        finally:
            self._monitor_force_full = False

        if getattr(self, "_monitor_dirty", False):
            try:
                if not self._monitor_timer.isActive():
                    self._monitor_timer.start()
            except Exception:
                pass


    def _flush_motor_monitor_render(self):
        if not getattr(self, "_motor_mon_dirty", False):
            return
        if hasattr(self, "motor_dock") and not self.motor_dock.isVisible():
            self._motor_mon_dirty = False
            self._motor_mon_force_full = True
            return

        self._motor_mon_dirty = False
        force_full = bool(getattr(self, "_motor_mon_force_full", False))
        try:
            self.render_motor_monitor(force_full=force_full)
        finally:
            self._motor_mon_force_full = False

        if getattr(self, "_motor_mon_dirty", False):
            try:
                if not self._motor_mon_timer.isActive():
                    self._motor_mon_timer.start()
            except Exception:
                pass
    def _flush_manual_render(self):
        if not getattr(self, "_manual_dirty", False):
            return
        if hasattr(self, "custom_send_dock") and not self.custom_send_dock.isVisible():
            self._manual_dirty = False
            self._manual_force_full = True
            return

        self._manual_dirty = False
        force_full = bool(getattr(self, "_manual_force_full", False))
        try:
            self.render_custom_send_log(force_full=force_full)
        finally:
            self._manual_force_full = False

        if getattr(self, "_manual_dirty", False):
            try:
                if not self._manual_timer.isActive():
                    self._manual_timer.start()
            except Exception:
                pass

    def _mark_plot_dirty(self, *args, **kwargs):
        self._plot_dirty = True

    def _set_plot_tab_visible(self, tab_widget: QWidget, title: str, visible: bool):
        if not hasattr(self, "plot_tabs"):
            return
        idx = self.plot_tabs.indexOf(tab_widget)
        if visible and idx < 0:
            self.plot_tabs.addTab(tab_widget, title)
            try:
                self.plot_tabs.setCurrentWidget(tab_widget)
            except Exception:
                pass
            try:
                self.plot_dock.show()
                self.plot_dock.raise_()
                self.plot_dock.activateWindow()
            except Exception:
                pass
        elif (not visible) and idx >= 0:
            try:
                self.plot_tabs.removeTab(idx)
            except Exception:
                pass

    def _refresh_friction_channel_options(self):
        if not hasattr(self, "fric_high_combo") or not hasattr(self, "fric_low_combo"):
            return
        names = []
        try:
            rows = int(self.ch_table.rowCount()) if hasattr(self, "ch_table") else 0
        except Exception:
            rows = 0
        for r in range(rows):
            name = None
            try:
                item = self.ch_table.item(r, 1) if hasattr(self, "ch_table") else None
                name = (item.text() if item else "").strip()
            except Exception:
                name = ""
            if not name:
                name = f"CH{r+1}"
            if name and name not in names:
                names.append(name)

        cur_high = self.fric_high_combo.currentText() if hasattr(self, "fric_high_combo") else ""
        cur_low = self.fric_low_combo.currentText() if hasattr(self, "fric_low_combo") else ""

        for combo in [self.fric_high_combo, self.fric_low_combo]:
            try:
                combo.blockSignals(True)
                combo.clear()
                if not names:
                    combo.addItem("(无通道)")
                    combo.setEnabled(False)
                else:
                    for n in names:
                        combo.addItem(n)
                    combo.setEnabled(True)
            except Exception:
                pass
            finally:
                try:
                    combo.blockSignals(False)
                except Exception:
                    pass

        # Restore selection if possible
        try:
            if cur_high and cur_high in names:
                self.fric_high_combo.setCurrentText(cur_high)
            if cur_low and cur_low in names:
                self.fric_low_combo.setCurrentText(cur_low)
        except Exception:
            pass

        self._on_friction_config_changed()

    def _swap_friction_channels(self):
        try:
            if not self.fric_high_combo.isEnabled() or not self.fric_low_combo.isEnabled():
                return
            hi = self.fric_high_combo.currentIndex()
            lo = self.fric_low_combo.currentIndex()
            if hi < 0 or lo < 0:
                return
            self.fric_high_combo.setCurrentIndex(lo)
            self.fric_low_combo.setCurrentIndex(hi)
        except Exception:
            pass
        self._on_friction_config_changed()

    def _on_friction_config_changed(self, *args):
        try:
            self._fric_high_name = (self.fric_high_combo.currentText() or "").strip()
            self._fric_low_name = (self.fric_low_combo.currentText() or "").strip()
        except Exception:
            self._fric_high_name = ""
            self._fric_low_name = ""
        try:
            self._wrap_angle_deg = float(self.wrap_angle_spin.value()) if hasattr(self, "wrap_angle_spin") else 0.0
        except Exception:
            self._wrap_angle_deg = 0.0
        try:
            self._wrap_angle_rad = math.radians(float(self._wrap_angle_deg)) if float(self._wrap_angle_deg) > 0 else 0.0
        except Exception:
            self._wrap_angle_rad = 0.0

        self._recalc_friction_buffers()
        try:
            self._plot_seq = int(getattr(self, "_plot_seq", 0) or 0) + 1
        except Exception:
            pass
        self._plot_dirty = True
        try:
            self.update_plot()
        except Exception:
            pass

    def _calc_fric_mu(self, high_v, low_v):
        try:
            if high_v is None or low_v is None:
                return None, None
            high = float(high_v)
            low = float(low_v)
        except Exception:
            return None, None
        try:
            if not math.isfinite(high) or not math.isfinite(low):
                return None, None
        except Exception:
            pass
        fric = high - low
        mu = None
        theta = float(getattr(self, "_wrap_angle_rad", 0.0) or 0.0)
        if theta > 0 and low > 0 and high > 0:
            try:
                ratio = high / low
                if ratio > 0:
                    mu = math.log(ratio) / theta
            except Exception:
                mu = None
        return fric, mu

    def _update_friction_buffers_at_index(self, idx: int, row: dict):
        if self._fric_buf is None or self._mu_buf is None:
            return
        high_name = (getattr(self, "_fric_high_name", "") or "").strip()
        low_name = (getattr(self, "_fric_low_name", "") or "").strip()
        if (not high_name) or (not low_name):
            fric, mu = None, None
        else:
            fric, mu = self._calc_fric_mu(row.get(high_name), row.get(low_name))
        if np is not None:
            try:
                self._fric_buf[idx] = (np.nan if fric is None else float(fric))
                self._mu_buf[idx] = (np.nan if mu is None else float(mu))
            except Exception:
                pass
        else:
            self._fric_buf[idx] = fric
            self._mu_buf[idx] = mu

    def _recalc_friction_buffers(self):
        size = int(getattr(self, "_buf_size", 0) or 0)
        if size <= 0 or self._fric_buf is None or self._mu_buf is None:
            return
        high_name = (getattr(self, "_fric_high_name", "") or "").strip()
        low_name = (getattr(self, "_fric_low_name", "") or "").strip()
        high_buf = self._val_buf_by_channel.get(high_name) if high_name else None
        low_buf = self._val_buf_by_channel.get(low_name) if low_name else None
        for i in range(size):
            if high_buf is None or low_buf is None:
                fric, mu = None, None
            else:
                try:
                    hv = high_buf[i]
                    lv = low_buf[i]
                except Exception:
                    hv = None
                    lv = None
                fric, mu = self._calc_fric_mu(hv, lv)
            if np is not None:
                try:
                    self._fric_buf[i] = (np.nan if fric is None else float(fric))
                    self._mu_buf[i] = (np.nan if mu is None else float(mu))
                except Exception:
                    pass
            else:
                self._fric_buf[i] = fric
                self._mu_buf[i] = mu

    def _update_friction_plots(self, xs, idx: int, full: bool, count: int, scroll_live: bool, x_left: float, x_right: float):
        if xs is None or self._fric_buf is None or self._mu_buf is None:
            return
        size = int(getattr(self, "_buf_size", 0) or 0)
        if size <= 0:
            return

        # X range sync
        if scroll_live:
            try:
                self.friction_plot.setXRange(x_left, x_right, padding=0.0)
            except Exception:
                pass
            try:
                self.mu_plot.setXRange(x_left, x_right, padding=0.0)
            except Exception:
                pass

        if np is not None:
            if full:
                first = size - idx
                if self._fric_plot_y is None or getattr(self._fric_plot_y, "shape", (0,))[0] != size:
                    self._fric_plot_y = np.empty(size, dtype=float)
                if self._mu_plot_y is None or getattr(self._mu_plot_y, "shape", (0,))[0] != size:
                    self._mu_plot_y = np.empty(size, dtype=float)
                self._fric_plot_y[:first] = self._fric_buf[idx:]
                self._fric_plot_y[first:] = self._fric_buf[:idx]
                self._mu_plot_y[:first] = self._mu_buf[idx:]
                self._mu_plot_y[first:] = self._mu_buf[:idx]
                ys_fric = self._fric_plot_y
                ys_mu = self._mu_plot_y
            else:
                ys_fric = self._fric_buf[:count]
                ys_mu = self._mu_buf[:count]
            try:
                self.friction_curve.setData(xs, ys_fric, connect="finite", skipFiniteCheck=True)
            except Exception:
                try:
                    self.friction_curve.setData(xs, ys_fric, connect="finite")
                except Exception:
                    pass
            try:
                self.mu_curve.setData(xs, ys_mu, connect="finite", skipFiniteCheck=True)
            except Exception:
                try:
                    self.mu_curve.setData(xs, ys_mu, connect="finite")
                except Exception:
                    pass
        else:
            if full:
                fric_raw = list(self._fric_buf[idx:]) + list(self._fric_buf[:idx])
                mu_raw = list(self._mu_buf[idx:]) + list(self._mu_buf[:idx])
            else:
                fric_raw = list(self._fric_buf[:count])
                mu_raw = list(self._mu_buf[:count])
            xs_f, ys_f = [], []
            xs_m, ys_m = [], []
            for t, v in zip(xs, fric_raw):
                if v is None:
                    continue
                xs_f.append(t)
                ys_f.append(v)
            for t, v in zip(xs, mu_raw):
                if v is None:
                    continue
                xs_m.append(t)
                ys_m.append(v)
            try:
                self.friction_curve.setData(xs_f, ys_f)
            except Exception:
                pass
            try:
                self.mu_curve.setData(xs_m, ys_m)
            except Exception:
                pass

        # Autoscale handling for derived plots
        try:
            auto = bool(self.autoscale_chk.isChecked()) if hasattr(self, "autoscale_chk") else True
        except Exception:
            auto = True
        try:
            self.friction_plot.enableAutoRange(axis="y", enable=auto)
            self.mu_plot.enableAutoRange(axis="y", enable=auto)
        except Exception:
            pass
    def _flush_plot(self):
        # Plot refresh is driven by timer (Hz). We always call update_plot(),
        # which will only re-upload curve data when new samples arrive, but
        # will keep X scrolling smoothly at the requested refresh rate.
        self.update_plot()


    def _on_plot_fps_changed(self, *args):
        """Apply plot refresh rate (Hz) to the plot timer."""
        try:
            hz = int(self.plot_fps_spin.value()) if hasattr(self, 'plot_fps_spin') else 60
        except Exception:
            hz = 60
        hz = max(1, min(240, hz))
        interval_ms = max(1, int(round(1000.0 / float(hz))))
        try:
            self._plot_timer.setInterval(interval_ms)
        except Exception:
            pass

        # apply start/stop based on current state
        try:
            self._update_plot_timer_running()
        except Exception:
            pass

    def _update_plot_timer_running(self):
        """Start/stop plot timer based on acquisition state (save CPU, freeze scrolling)."""
        try:
            live = bool(getattr(self, "is_acquiring", False)) and (not bool(getattr(self, "is_paused", False)))
        except Exception:
            live = False

        if live:
            try:
                if not self._plot_timer.isActive():
                    self._plot_timer.start()
            except Exception:
                pass
        else:
            try:
                if self._plot_timer.isActive():
                    self._plot_timer.stop()
            except Exception:
                pass


    def _on_max_points_changed(self, *args):
        """Resize ring buffers when max points changes."""
        try:
            new_size = int(self.max_points_spin.value())
        except Exception:
            return
        self._resize_ring_buffers(new_size)
        self._mark_plot_dirty()

    def _resize_ring_buffers(self, new_size: int):
        new_size = int(max(10, new_size))
        old_size = int(getattr(self, '_buf_size', 0) or 0)
        if new_size <= 0 or new_size == old_size:
            return
        xs, ys_map = self._snapshot_ring()
        self._alloc_ring_buffers(new_size, list(self.channel_names), keep_last=True, xs=xs, ys_map=ys_map)

    def _alloc_ring_buffers(self, size: int, channel_names: list, keep_last: bool = False, xs=None, ys_map=None):
        """Allocate ring buffers.

        When keep_last=True, copies the last min(len(xs), size) samples into the new buffer.
        """
        size = int(max(10, size))
        self._buf_size = size
        self._buf_count = 0
        self._buf_idx = 0
        self._plot_seq = 0
        self._last_plotted_seq = -1

        if np is not None:
            self._ts_buf = np.full(size, np.nan, dtype=float)
            self._plot_x = np.empty(size, dtype=float)
        else:
            self._ts_buf = [None] * size
            self._plot_x = None

        if np is not None:
            self._fric_buf = np.full(size, np.nan, dtype=float)
            self._mu_buf = np.full(size, np.nan, dtype=float)
            self._fric_plot_y = np.empty(size, dtype=float)
            self._mu_plot_y = np.empty(size, dtype=float)
        else:
            self._fric_buf = [None] * size
            self._mu_buf = [None] * size
            self._fric_plot_y = None
            self._mu_plot_y = None

        self._val_buf_by_channel = {}
        self._plot_y_by_channel = {}
        for name in channel_names:
            if np is not None:
                self._val_buf_by_channel[name] = np.full(size, np.nan, dtype=float)
                self._plot_y_by_channel[name] = np.empty(size, dtype=float)
            else:
                self._val_buf_by_channel[name] = [None] * size

        if keep_last and xs:
            try:
                k = min(len(xs), size)
            except Exception:
                k = 0
            if k > 0:
                tail_x = xs[-k:]
                if np is not None:
                    self._ts_buf[:k] = np.asarray(tail_x, dtype=float)
                else:
                    self._ts_buf[:k] = list(tail_x)

                for name in channel_names:
                    ys = (ys_map or {}).get(name, [])
                    tail_y = ys[-k:] if ys else [None] * k
                    if np is not None:
                        arr = np.asarray([(np.nan if v is None else float(v)) for v in tail_y], dtype=float)
                        self._val_buf_by_channel[name][:k] = arr
                    else:
                        self._val_buf_by_channel[name][:k] = list(tail_y)

                # recompute friction buffers for preserved samples
                high_name = (getattr(self, "_fric_high_name", "") or "").strip()
                low_name = (getattr(self, "_fric_low_name", "") or "").strip()
                if high_name and low_name:
                    tail_high = (ys_map or {}).get(high_name, [])
                    tail_low = (ys_map or {}).get(low_name, [])
                    tail_high = tail_high[-k:] if tail_high else [None] * k
                    tail_low = tail_low[-k:] if tail_low else [None] * k
                    for j in range(k):
                        fric, mu = self._calc_fric_mu(tail_high[j], tail_low[j])
                        if np is not None:
                            try:
                                self._fric_buf[j] = (np.nan if fric is None else float(fric))
                                self._mu_buf[j] = (np.nan if mu is None else float(mu))
                            except Exception:
                                pass
                        else:
                            self._fric_buf[j] = fric
                            self._mu_buf[j] = mu
                else:
                    if np is not None:
                        try:
                            self._fric_buf[:k] = np.nan
                            self._mu_buf[:k] = np.nan
                        except Exception:
                            pass
                    else:
                        self._fric_buf[:k] = [None] * k
                        self._mu_buf[:k] = [None] * k
                self._buf_count = k
                self._buf_idx = k % size

    def _snapshot_ring(self):
        """Snapshot ring buffer into time-ordered python lists (for resize/export)."""
        count = int(getattr(self, '_buf_count', 0) or 0)
        size = int(getattr(self, '_buf_size', 0) or 0)
        if count <= 0 or size <= 0 or self._ts_buf is None:
            return [], {}
        idx = int(getattr(self, '_buf_idx', 0) or 0)

        if count < size:
            # not wrapped
            if np is not None:
                xs = [float(x) for x in self._ts_buf[:count]]
            else:
                xs = list(self._ts_buf[:count])
        else:
            # wrapped: oldest at idx
            if np is not None:
                xs = [float(x) for x in self._ts_buf[idx:]] + [float(x) for x in self._ts_buf[:idx]]
            else:
                xs = list(self._ts_buf[idx:]) + list(self._ts_buf[:idx])

        ys_map = {}
        for name in list(self.channel_names):
            buf = self._val_buf_by_channel.get(name)
            if buf is None:
                continue
            if np is not None:
                if count < size:
                    arr = buf[:count]
                else:
                    arr = np.concatenate((buf[idx:], buf[:idx]))
                ys = [None if (not np.isfinite(v)) else float(v) for v in arr]
            else:
                if count < size:
                    ys = list(buf[:count])
                else:
                    ys = list(buf[idx:]) + list(buf[:idx])
            ys_map[name] = ys
        return xs, ys_map

    # ---------- monitor ----------
    def append_monitor(self, s: str):
        """Append a plain text info line (non-frame).

        NOTE: UI rendering is throttled to avoid stutter at high frame rates.
        """
        self._monitor_entries.append({"kind": "INFO", "data": b"", "tag": "", "note": str(s)})
        self.schedule_monitor_render()

    def _custom_send_current_tag(self) -> str:
        """Return the current filter tag (e.g. 'rx:COM3' / 'tx:COM4') for custom-send dock."""
        if not self.is_connected or self.worker is None:
            return ""
        if not hasattr(self, "custom_send_port_combo"):
            return ""
        target = self.custom_send_port_combo.currentData()
        if target == "rx":
            p = getattr(self.worker, "port", "") or ""
            return f"rx:{p}" if p else ""
        if target == "tx":
            p = getattr(self.worker, "tx_port", "") or ""
            return f"tx:{p}" if p else ""
        return ""
    def on_tx_monitor_toggled(self, checked: bool):
        """Enable/disable TX port async RX monitoring tap.

        仅影响通讯监视窗口对发送串口的 RX 监听（窥探式读取）。
        """
        if self.is_connected and self.worker is not None:
            try:
                self.worker.set_tx_tap_enabled(bool(checked))
            except Exception:
                pass
        self.schedule_monitor_render()


    def clear_monitor(self):
        self._monitor_entries.clear()
        self.monitor_text.clear()
        self._monitor_render_idx = 0
        self._monitor_render_mode = None

    def clear_custom_send_log(self):
        self._manual_entries.clear()
        self.custom_send_log.clear()
        self._manual_render_idx = 0
        self._manual_render_mode = None
        self._manual_render_tag = None


    def clear_motor_monitor(self):
        self._motor_mon_entries.clear()
        if hasattr(self, "motor_mon_text"):
            self.motor_mon_text.clear()
        self._motor_mon_render_idx = 0
        self._motor_mon_render_mode = None
    def _decode_bytes(self, data: bytes, mode: str) -> str:
        if mode == "hex":
            return hex_bytes(data)
        enc = "utf-8" if mode == "utf-8" else "gbk"
        try:
            return data.decode(enc, errors="replace")
        except Exception:
            return data.decode("utf-8", errors="replace")

    def _format_entry(self, e: dict, mode: str) -> str:
        kind = e.get("kind", "")
        tag = e.get("tag", "")
        note = e.get("note", "")
        data = e.get("data", b"") or b""
        prefix = kind
        if tag:
            prefix += f"[{tag}]"
        if data:
            payload = self._decode_bytes(data, mode)
            if note:
                return f"{prefix}: {payload}  {note}"
            return f"{prefix}: {payload}"
        # no data
        return f"{prefix}: {note}" if prefix else str(note)

    def render_monitor(self, force_full: bool = False):
        mode = "hex"
        if hasattr(self, "monitor_mode_combo"):
            mode = self.monitor_mode_combo.currentData() or "hex"

        # Track mode/index for delta-append rendering
        last_mode = getattr(self, "_monitor_render_mode", None)
        render_idx = int(getattr(self, "_monitor_render_idx", 0) or 0)

        # If mode changed or forced, rebuild last N lines (rare path)
        if force_full or last_mode != mode or render_idx <= 0:
            max_lines = 2000
            entries = self._monitor_entries[-max_lines:]
            try:
                self.monitor_text.blockSignals(True)
                self.monitor_text.setPlainText("\n".join(self._format_entry(e, mode) for e in entries))
                self.monitor_text.blockSignals(False)
                self.monitor_text.moveCursor(QTextCursor.End)
            except Exception:
                pass
            self._monitor_render_mode = mode
            self._monitor_render_idx = len(self._monitor_entries)
            return

        # Delta append new lines
        if render_idx < len(self._monitor_entries):
            new_entries = self._monitor_entries[render_idx:]
            # Batch append to reduce UI updates
            lines = [self._format_entry(e, mode) for e in new_entries]
            if lines:
                try:
                    self.monitor_text.blockSignals(True)
                    self.monitor_text.appendPlainText("\n".join(lines))
                    self.monitor_text.blockSignals(False)
                    self.monitor_text.moveCursor(QTextCursor.End)
                except Exception:
                    pass
        self._monitor_render_mode = mode
        self._monitor_render_idx = len(self._monitor_entries)

    def render_custom_send_log(self, force_full: bool = False):
        if not hasattr(self, "custom_send_mode_combo"):
            return
        mode = self.custom_send_mode_combo.currentData() or "hex"
        tag_filter = self._custom_send_current_tag()

        last_mode = getattr(self, "_manual_render_mode", None)
        last_tag = getattr(self, "_manual_render_tag", None)
        render_idx = int(getattr(self, "_manual_render_idx", 0) or 0)

        if force_full or last_mode != mode or last_tag != tag_filter or render_idx <= 0:
            max_lines = 1200
            entries = [e for e in self._manual_entries if (not tag_filter or e.get("tag") == tag_filter)]
            entries = entries[-max_lines:]
            try:
                self.custom_send_log.blockSignals(True)
                self.custom_send_log.setPlainText("\n".join(self._format_entry(e, mode) for e in entries))
                self.custom_send_log.blockSignals(False)
                self.custom_send_log.moveCursor(QTextCursor.End)
            except Exception:
                pass
            self._manual_render_mode = mode
            self._manual_render_tag = tag_filter
            self._manual_render_idx = len(self._manual_entries)
            return

        if render_idx < len(self._manual_entries):
            new_entries = self._manual_entries[render_idx:]
            lines = []
            for e in new_entries:
                if (not tag_filter) or (e.get("tag") == tag_filter):
                    lines.append(self._format_entry(e, mode))
            if lines:
                try:
                    self.custom_send_log.blockSignals(True)
                    self.custom_send_log.appendPlainText("\n".join(lines))
                    self.custom_send_log.blockSignals(False)
                    self.custom_send_log.moveCursor(QTextCursor.End)
                except Exception:
                    pass
        self._manual_render_mode = mode
        self._manual_render_tag = tag_filter
        self._manual_render_idx = len(self._manual_entries)


    def render_motor_monitor(self, force_full: bool = False):
        mode = "utf-8"
        if hasattr(self, "motor_mon_mode_combo"):
            mode = self.motor_mon_mode_combo.currentData() or "utf-8"
        last_mode = getattr(self, "_motor_mon_render_mode", None)
        render_idx = int(getattr(self, "_motor_mon_render_idx", 0) or 0)
        max_lines = 2000

        if force_full or last_mode != mode or render_idx > len(self._motor_mon_entries):
            entries = self._motor_mon_entries[-max_lines:]
            try:
                self.motor_mon_text.blockSignals(True)
                self.motor_mon_text.setPlainText("\n".join(self._format_entry(e, mode) for e in entries))
                self.motor_mon_text.blockSignals(False)
                self.motor_mon_text.moveCursor(QTextCursor.End)
            except Exception:
                pass
            self._motor_mon_render_mode = mode
            self._motor_mon_render_idx = len(self._motor_mon_entries)
            return

        if render_idx < len(self._motor_mon_entries):
            new_entries = self._motor_mon_entries[render_idx:]
            lines_out = [self._format_entry(e, mode) for e in new_entries]
            if lines_out:
                try:
                    self.motor_mon_text.blockSignals(True)
                    self.motor_mon_text.appendPlainText("\n".join(lines_out))
                    self.motor_mon_text.blockSignals(False)
                    self.motor_mon_text.moveCursor(QTextCursor.End)
                except Exception:
                    pass
        self._motor_mon_render_mode = mode
        self._motor_mon_render_idx = len(self._motor_mon_entries)
    @Slot(str, bytes, str, str)
    def on_frame(self, kind: str, data: bytes, tag: str, note: str):
        e = {"kind": str(kind), "data": bytes(data or b""), "tag": str(tag or ""), "note": str(note or "")}

        # Always keep manual TX/RX entries for the custom-send dock
        is_manual = str(kind).startswith('TX_MANUAL') or str(kind).startswith('RX_MANUAL')
        if is_manual:
            self._manual_entries.append(e)
            if self.custom_send_dock.isVisible():
                self.schedule_custom_send_render()

            # cap memory for logs
            if len(self._manual_entries) > 6000:
                overflow = len(self._manual_entries) - 6000
                del self._manual_entries[:overflow]
                try:
                    self._manual_render_idx = max(0, int(getattr(self, '_manual_render_idx', 0)) - overflow)
                except Exception:
                    self._manual_render_idx = 0

        # Comm monitor filtering per-port (rx/tx listen checkboxes)
        t = e.get('tag', '') or ''
        allow = True
        if t.startswith('rx:') and hasattr(self, 'mon_rx_chk'):
            allow = bool(self.mon_rx_chk.isChecked())
        elif t.startswith('tx:') and hasattr(self, 'mon_tx_chk'):
            allow = bool(self.mon_tx_chk.isChecked())

        if allow:
            self._monitor_entries.append(e)
            self.schedule_monitor_render()

            if len(self._monitor_entries) > 12000:
                overflow = len(self._monitor_entries) - 12000
                del self._monitor_entries[:overflow]
                try:
                    self._monitor_render_idx = max(0, int(getattr(self, '_monitor_render_idx', 0)) - overflow)
                except Exception:
                    self._monitor_render_idx = 0



        # Motor TX monitor: only RX frames from tx port
        t2 = e.get("tag", "") or ""
        k2 = str(kind)
        if t2.startswith("tx:") and k2.startswith("RX"):
            self._motor_mon_entries.append(e)
            self.schedule_motor_monitor_render()
            if len(self._motor_mon_entries) > 8000:
                overflow = len(self._motor_mon_entries) - 8000
                del self._motor_mon_entries[:overflow]
                try:
                    self._motor_mon_render_idx = max(0, int(getattr(self, "_motor_mon_render_idx", 0)) - overflow)
                except Exception:
                    self._motor_mon_render_idx = 0
    def save_monitor_log(self):
        path, _ = QFileDialog.getSaveFileName(self, "保存通讯日志", "comm_log.txt", "Text Files (*.txt)")
        if not path:
            return
        if not path.lower().endswith(".txt"):
            path += ".txt"
        try:
            with open(path, "w", encoding="utf-8") as f:
                f.write(self.monitor_text.toPlainText())
            self.set_status(f"通讯日志已保存：{path}")
        except Exception as e:
            QMessageBox.critical(self, "保存失败", f"保存通讯日志失败：\n{e}")

    # ---------- status ----------
    def set_status(self, msg: str):
        self.status_label.setText(f"状态：{msg}")

    # ---------- ports ----------
    def refresh_ports(self):
        """Refresh physical COM ports + in-app simulated ports."""
        # Preserve current selections to avoid UI jump.
        cur_rx = self.port_combo.currentData()
        cur_tx = self.tx_port_combo.currentData()

        self.port_combo.clear()
        self.tx_port_combo.clear()

        items = []

        # Physical ports
        try:
            ports = list(list_ports.comports())
        except Exception:
            ports = []

        for p in ports:
            desc = p.description or ""
            manu = getattr(p, "manufacturer", "") or ""
            prod = getattr(p, "product", "") or ""
            hwid = getattr(p, "hwid", "") or ""
            extra = " | ".join([x for x in [desc, manu, prod] if x and x.lower() != "n/a"])
            label = f"{p.device}"
            if extra:
                label += f"  —  {extra}"
            elif hwid:
                label += f"  —  {hwid}"
            items.append((label, p.device))

        # Simulated ports
        for info in SIM_REGISTRY.list_infos():
            label = f"{info.com}  —  仿真串口"
            items.append((label, info.key))

        if not items:
            self.port_combo.addItem("(未发现串口)", "")
            self.tx_port_combo.addItem("(未发现串口)", "")
        else:
            for label, key in items:
                self.port_combo.addItem(label, key)
                self.tx_port_combo.addItem(label, key)
        # Restore selections if possible
        try:
            if cur_rx:
                idx = self.port_combo.findData(cur_rx)
                if idx >= 0:
                    self.port_combo.setCurrentIndex(idx)
            if cur_tx:
                idx2 = self.tx_port_combo.findData(cur_tx)
                if idx2 >= 0:
                    self.tx_port_combo.setCurrentIndex(idx2)
        except Exception:
            pass

        # keep custom-send dock port list in sync
        self.update_custom_send_ports()
    def update_custom_send_ports(self):
        """Only show ports that are already opened by this program."""
        if not hasattr(self, "custom_send_port_combo"):
            return

        combo = self.custom_send_port_combo
        combo.blockSignals(True)
        combo.clear()

        items = []
        if self.is_connected and self.worker is not None:
            # rx/modbus port is always connected when worker emits connected(True)
            rx_port = getattr(self.worker, "port", "")
            if rx_port:
                items.append((f"接收串口(Modbus)：{rx_port}", "rx"))

            # tx/output port may be enabled
            if getattr(self.worker, "tx_enabled", False) and getattr(self.worker, "_tx_ser", None) is not None:
                tx_port = getattr(self.worker, "tx_port", "")
                if tx_port and tx_port != rx_port:
                    items.append((f"发送串口(输出)：{tx_port}", "tx"))
                elif tx_port:
                    items.append((f"发送串口(输出)：{tx_port}", "tx"))

        if not items:
            combo.addItem("(未连接串口)", "")
            if hasattr(self, "custom_send_line"):
                self.custom_send_line.setEnabled(False)
            if hasattr(self, "custom_send_btn"):
                self.custom_send_btn.setEnabled(False)
        else:
            for text, data in items:
                combo.addItem(text, data)
            if hasattr(self, "custom_send_line"):
                self.custom_send_line.setEnabled(True)
            if hasattr(self, "custom_send_btn"):
                self.custom_send_btn.setEnabled(True)

        combo.blockSignals(False)

    def send_custom_serial(self):
        if not self.is_connected or self.worker is None:
            QMessageBox.information(self, "提示", "请先连接串口后再发送。")
            return

        target = self.custom_send_port_combo.currentData()
        if not target:
            QMessageBox.warning(self, "提示", "当前没有可发送的已连接串口。")
            return

        text = (self.custom_send_line.text() if self.custom_send_line else "").strip()
        if not text:
            return

        add_crlf = bool(self.custom_send_crlf_chk.isChecked())
        self.worker.enqueue_custom_send(str(target), text, add_crlf)
        # don't clear input (方便连续修改/重复发送)
        self.custom_send_line.setFocus()

    # ---------- motor control ----------
    def _set_lamp_color(self, label: QLabel, color: str):
        if label is None:
            return
        try:
            label.setStyleSheet(f"background:{color};border-radius:7px;border:1px solid #444;")
        except Exception:
            pass


    def _set_motor_mode_lamps(self, mode: Optional[int]):
        if mode == 0:
            self._set_lamp_color(self.motor_mode_tension_lamp, "#5cb85c")
            self._set_lamp_color(self.motor_mode_speed_lamp, "#777777")
        elif mode == 1:
            self._set_lamp_color(self.motor_mode_tension_lamp, "#777777")
            self._set_lamp_color(self.motor_mode_speed_lamp, "#5cb85c")
        else:
            self._set_lamp_color(self.motor_mode_tension_lamp, "#777777")
            self._set_lamp_color(self.motor_mode_speed_lamp, "#777777")

    def _require_motor_mode(self, required: Optional[int] = None) -> bool:
        if getattr(self, "motor_mode", None) is None:
            QMessageBox.warning(self, "提示", "请先选择控制模式。")
            return False
        if required is not None and int(self.motor_mode) != int(required):
            mode_name = "张力模式" if int(required) == 0 else "速度模式"
            QMessageBox.warning(self, "提示", f"请先切换到{mode_name}。")
            return False
        return True

    def _motor_can_send(self) -> bool:
        if not self.is_connected or self.worker is None:
            QMessageBox.information(self, "提示", "请先连接串口后再操作电机控制。")
            return False
        if not getattr(self.worker, "tx_enabled", False) or getattr(self.worker, "_tx_ser", None) is None:
            QMessageBox.information(self, "提示", "请先启用“发送串口(输出)”并连接。")
            return False
        return True

    def _send_motor_cmd(self, cmd: str) -> bool:
        if not self._motor_can_send():
            return False
        text = (cmd or "").strip()
        if not text:
            return False
        self.worker.enqueue_custom_send("tx", text, True)
        return True

    def _parse_number_text(self, text: str, label: str) -> Optional[str]:
        t = (text or "").strip()
        if not t:
            QMessageBox.warning(self, "提示", f"请输入{label}。")
            return None
        try:
            v = float(t)
        except Exception:
            QMessageBox.warning(self, "提示", f"{label}不是有效数字。")
            return None
        if v != v or v in (float("inf"), float("-inf")):
            QMessageBox.warning(self, "提示", f"{label}不是有效数字。")
            return None
        return f"{v:g}"

    def on_motor_mode_tension(self):
        if self._send_motor_cmd("ConMode 0"):
            self.motor_mode = 0
            self._set_motor_mode_lamps(self.motor_mode)

    def on_motor_mode_speed(self):
        if self._send_motor_cmd("ConMode 1"):
            self.motor_mode = 1
            self._set_motor_mode_lamps(self.motor_mode)

    def on_motor_enable(self):
        if self._send_motor_cmd("Enable"):
            self._set_lamp_color(self.motor_enable_lamp, "#5cb85c")

    def on_motor_disable(self):
        if self._send_motor_cmd("Disable"):
            self._set_lamp_color(self.motor_enable_lamp, "#777777")

    def on_motor_forward(self):
        if self._send_motor_cmd("Forward"):
            self._set_lamp_color(self.motor_dir_lamp, "#5cb85c")

    def on_motor_backward(self):
        if self._send_motor_cmd("Backward"):
            self._set_lamp_color(self.motor_dir_lamp, "#f0ad4e")

    def on_motor_speed(self):
        if not self._require_motor_mode(1):
            return
        val = self._parse_number_text(self.motor_speed_edit.text(), "转速(RPM)")
        if val is None:
            return
        self._send_motor_cmd(f"Con {val}")

    def on_motor_tension(self):
        if not self._require_motor_mode(0):
            return
        val = self._parse_number_text(self.motor_tension_edit.text(), "张力(g)")
        if val is None:
            return
        self._send_motor_cmd(f"F {val}")

    def on_motor_pid(self):
        kp = self._parse_number_text(self.motor_kp_edit.text(), "Kp")
        if kp is None:
            return
        ki = self._parse_number_text(self.motor_ki_edit.text(), "Ki")
        if ki is None:
            return
        kd = self._parse_number_text(self.motor_kd_edit.text(), "Kd")
        if kd is None:
            return
        self._send_motor_cmd(f"pid {kp} {ki} {kd}")

    def on_motor_estop(self):
        if not self._motor_can_send():
            return
        # Order: tension -> speed -> disable
        self._send_motor_cmd("F 0")
        self._send_motor_cmd("Con 0")
        self._send_motor_cmd("Disable")
        self.motor_mode = None
        self._set_motor_mode_lamps(None)
        self._set_lamp_color(self.motor_enable_lamp, "#777777")
        self._set_lamp_color(self.motor_dir_lamp, "#777777")

    def _set_serial_widgets_enabled(self, enabled: bool):
        for w in [
            self.port_combo, self.refresh_ports_btn, self.baud_combo,
            self.parity_combo, self.stopbits_combo, self.bytesize_combo, self.timeout_spin,
            self.rs485_mode_combo, self.pre_tx_spin, self.post_tx_spin,
            self.tx_port_combo, self.tx_baud_combo, self.enable_tx_chk, self.mon_rx_chk, self.mon_tx_chk
        ]:
            w.setEnabled(enabled)

    # ---------- channel table ----------
    def add_channel_row(self, default_name: str = "", default_addr: int = 0, default_dtype: str = "float32"):
        row = self.ch_table.rowCount()
        self.ch_table.insertRow(row)

        enabled_chk = QCheckBox()
        enabled_chk.setChecked(True)
        enabled_chk.setStyleSheet("margin-left:12px;")
        self.ch_table.setCellWidget(row, 0, enabled_chk)

        self.ch_table.setItem(row, 1, QTableWidgetItem(default_name or f"CH{row+1}"))
        self.ch_table.setItem(row, 2, QTableWidgetItem(str(default_addr)))

        dtype_combo = QComboBox()
        dtype_combo.addItems(list(DTYPE_INFO.keys()))
        dtype_combo.setCurrentText(default_dtype if default_dtype in DTYPE_INFO else "float32")
        self.ch_table.setCellWidget(row, 3, dtype_combo)

        byte_combo = QComboBox()
        byte_combo.addItems(["big", "little"])
        byte_combo.setCurrentText("big")
        self.ch_table.setCellWidget(row, 4, byte_combo)

        word_combo = QComboBox()
        word_combo.addItems(["big", "little"])
        word_combo.setCurrentText("big")
        self.ch_table.setCellWidget(row, 5, word_combo)

        self.ch_table.setItem(row, 6, QTableWidgetItem("-0.01"))

        self._refresh_friction_channel_options()
    def delete_selected_rows(self):
        rows = sorted({idx.row() for idx in self.ch_table.selectedIndexes()}, reverse=True)
        for r in rows:
            self.ch_table.removeRow(r)

        self._refresh_friction_channel_options()
    def gather_channels(self) -> List[ChannelConfig]:
        channels: List[ChannelConfig] = []
        seen_names = set()
        for r in range(self.ch_table.rowCount()):
            enabled_widget = self.ch_table.cellWidget(r, 0)
            enabled = bool(enabled_widget.isChecked()) if enabled_widget else True
            name = (self.ch_table.item(r, 1).text() if self.ch_table.item(r, 1) else "").strip() or f"CH{r+1}"
            if name in seen_names:
                i = 2
                base = name
                while f"{base}_{i}" in seen_names:
                    i += 1
                name = f"{base}_{i}"
            seen_names.add(name)

            try:
                address = int((self.ch_table.item(r, 2).text() if self.ch_table.item(r, 2) else "0").strip())
            except Exception:
                address = 0

            dtype_combo = self.ch_table.cellWidget(r, 3)
            dtype = dtype_combo.currentText() if dtype_combo else "float32"

            byte_combo = self.ch_table.cellWidget(r, 4)
            byte_order = byte_combo.currentText() if byte_combo else "big"

            word_combo = self.ch_table.cellWidget(r, 5)
            word_order = word_combo.currentText() if word_combo else "big"

            try:
                scale = float((self.ch_table.item(r, 6).text() if self.ch_table.item(r, 6) else "1.0").strip())
            except Exception:
                scale = 1.0

            channels.append(ChannelConfig(enabled=enabled, name=name, address=address, dtype=dtype,
                                         byte_order=byte_order, word_order=word_order, scale=scale))
        return channels


    # ---------- plot/data ----------
    def clear_data(self):
        # reset ring buffers (size follows current max-points)
        try:
            size = int(self.max_points_spin.value())
        except Exception:
            size = int(getattr(self, '_buf_size', 100) or 100)

        self.channel_names.clear()
        self._alloc_ring_buffers(size, [], keep_last=False)

        # reset plot time base (relative seconds)
        self._t0_mono_ts = None
        self._last_sample_rel_ts = None
        self._last_sample_mono_ts = None

        # reset pause-compensation
        self._mono_pause_accum = 0.0
        self._mono_pause_start = None

        self.plot.clear()
        self.plot.addLegend()
        self.curves.clear()
        try:
            if hasattr(self, "friction_curve") and self.friction_curve is not None:
                self.friction_curve.setData([], [])
            if hasattr(self, "mu_curve") and self.mu_curve is not None:
                self.mu_curve.setData([], [])
        except Exception:
            pass
        self.set_status("已清空数据")

    def init_curves(self, channel_names: List[str]):
        self.plot.clear()
        self.plot.addLegend()
        self.curves.clear()

        # 1-7 通道配色：红 橙 黄 绿 青 蓝 紫（更高区分度）
        palette = [
            (220, 0, 0),      # red
            (255, 140, 0),    # orange
            (255, 200, 0),    # yellow (slightly darker on white)
            (0, 170, 0),      # green
            (0, 170, 170),    # cyan
            (0, 0, 220),      # blue
            (140, 0, 200),    # purple
        ]
        width = 2  # 线宽稍微粗一点

        for i, name in enumerate(channel_names):
            color = palette[i % len(palette)]
            pen = pg.mkPen(color=color, width=width)
            self.curves[name] = self.plot.plot([], [], name=name, pen=pen)
            item = self.curves.get(name)
            if item is not None:
                # Per-curve performance hints (safe across versions)
                try:
                    item.setClipToView(True)
                except Exception:
                    pass
                try:
                    item.setDownsampling(auto=True, mode='peak')
                except Exception:
                    try:
                        item.setDownsampling(auto=True, method='peak')
                    except Exception:
                        pass
                # Some versions support skipping finite checks for speed
                try:
                    item.setSkipFiniteCheck(True)
                except Exception:
                    pass



    @Slot(float, dict)
    def on_data_ready(self, ts: float, row: dict):
        # Convert incoming timestamp to a smooth, relative monotonic time base.
        mono_now = time.monotonic()
        if self._t0_mono_ts is None:
            self._t0_mono_ts = float(mono_now)
        pause_accum = float(getattr(self, '_mono_pause_accum', 0.0) or 0.0)
        rel_ts = float(mono_now - float(self._t0_mono_ts) - pause_accum)
        if rel_ts < 0.0:
            rel_ts = 0.0
        self._last_sample_rel_ts = rel_ts
        self._last_sample_mono_ts = float(mono_now)

        # Lazily init curves + buffers on first frame (兼容未点击开始采集时的数据)
        if not self.channel_names:
            self.channel_names = list(row.keys())
            self.init_curves(self.channel_names)
            try:
                size = int(self.max_points_spin.value())
            except Exception:
                size = int(getattr(self, '_buf_size', 100) or 100)
            self._alloc_ring_buffers(size, list(self.channel_names), keep_last=False)

        # Keep buffer size synced with UI
        try:
            want = int(self.max_points_spin.value())
        except Exception:
            want = int(getattr(self, '_buf_size', 0) or 0)
        if want and want != int(getattr(self, '_buf_size', 0) or 0):
            self._resize_ring_buffers(want)

        size = int(getattr(self, '_buf_size', 0) or 0)
        if size <= 0:
            return
        i = int(getattr(self, '_buf_idx', 0) or 0) % size

        # Append to ring buffer (array size == 当前窗口最大点数)
        if np is not None:
            try:
                self._ts_buf[i] = float(rel_ts)
            except Exception:
                self._ts_buf[i] = np.nan
            for name in self.channel_names:
                v = row.get(name, None)
                try:
                    self._val_buf_by_channel[name][i] = (np.nan if v is None else float(v))
                except Exception:
                    self._val_buf_by_channel[name][i] = np.nan
        else:
            self._ts_buf[i] = rel_ts
            for name in self.channel_names:
                self._val_buf_by_channel[name][i] = row.get(name, None)

        # update derived friction buffers
        self._update_friction_buffers_at_index(i, row)

        self._buf_idx = (i + 1) % size
        if int(getattr(self, '_buf_count', 0) or 0) < size:
            self._buf_count += 1

        self._plot_seq = int(getattr(self, '_plot_seq', 0) or 0) + 1
        self._plot_dirty = True


    def update_plot(self):
        """Update plot curves (buffered + 刷新率驱动)."""
        count = int(getattr(self, "_buf_count", 0) or 0)
        if count <= 0 or not self.channel_names:
            return

        last_seq = int(getattr(self, "_last_plotted_seq", -1))
        cur_seq = int(getattr(self, "_plot_seq", 0) or 0)
        new_data = (last_seq != cur_seq)

        size = int(getattr(self, "_buf_size", 0) or 0)
        if size <= 0 or self._ts_buf is None:
            return
        idx = int(getattr(self, "_buf_idx", 0) or 0) % size
        full = (count >= size)
        # Smooth X scrolling (live): drive the right edge by monotonic time.
        # When NOT acquiring (stopped/paused), freeze scrolling and do NOT
        # keep forcing XRange updates (so users can pan/zoom the last frame).
        scroll_live = bool(getattr(self, 'is_acquiring', False)) and (not bool(getattr(self, 'is_paused', False)))

        try:
            if scroll_live and self._t0_mono_ts is not None:
                pause_accum = float(getattr(self, '_mono_pause_accum', 0.0) or 0.0)
                now_rel = float(time.monotonic() - float(self._t0_mono_ts) - pause_accum)
            else:
                now_rel = float(self._last_sample_rel_ts) if self._last_sample_rel_ts is not None else 0.0
        except Exception:
            now_rel = float(self._last_sample_rel_ts) if self._last_sample_rel_ts is not None else 0.0

        # Visible window width in seconds: max_points * poll_interval
        try:
            poll_ms = int(self.poll_spin.value()) if hasattr(self, 'poll_spin') else 20
        except Exception:
            poll_ms = 20
        poll_s = max(0.001, float(poll_ms) / 1000.0)
        npts = int(min(count, size))
        span = max(0.02, max(1, npts - 1) * poll_s)
        x_left = now_rel - span
        x_right = now_rel

        # Fast path: if no new samples arrived, avoid re-uploading curve data.
        # Keep only X scrolling (smooth) while reducing CPU/GPU overhead.
        if not new_data:
            if scroll_live:
                try:
                    self.plot.setXRange(x_left, x_right, padding=0.0)
                except Exception:
                    pass
                try:
                    self.friction_plot.setXRange(x_left, x_right, padding=0.0)
                except Exception:
                    pass
                try:
                    self.mu_plot.setXRange(x_left, x_right, padding=0.0)
                except Exception:
                    pass
            return

        # Prepare ordered X view only when uploading new curve data.
        xs = None
        if new_data:
            if np is not None:
                if not full:
                    xs = self._ts_buf[:count]
                else:
                    first = size - idx
                    self._plot_x[:first] = self._ts_buf[idx:]
                    self._plot_x[first:] = self._ts_buf[:idx]
                    xs = self._plot_x
            else:
                if not full:
                    xs = list(self._ts_buf[:count])
                else:
                    xs = list(self._ts_buf[idx:]) + list(self._ts_buf[:idx])
                if not xs:
                    xs = None

        # Prevent multiple repaints during one update (helps on Windows).
        self.plot.setUpdatesEnabled(False)
        try:
            global_y_min = None
            global_y_max = None

            if new_data and xs is not None and np is not None:
                for name in self.channel_names:
                    buf = self._val_buf_by_channel.get(name)
                    if buf is None:
                        continue

                    if not full:
                        ys = buf[:count]
                    else:
                        ytmp = self._plot_y_by_channel.get(name)
                        if ytmp is None or getattr(ytmp, 'shape', (0,))[0] != size:
                            ytmp = np.empty(size, dtype=float)
                            self._plot_y_by_channel[name] = ytmp
                        first = size - idx
                        ytmp[:first] = buf[idx:]
                        ytmp[first:] = buf[:idx]
                        ys = ytmp
                    ys_use = ys

                    curve = self.curves.get(name)
                    if curve is not None:
                        # Limit points sent to renderer when buffer is huge.
                        xs_use, ys_use = xs, ys
                        try:
                            max_disp = int(getattr(self, '_max_display_points', 0) or 0)
                        except Exception:
                            max_disp = 0
                        if max_disp:
                            try:
                                n = int(len(xs_use))
                            except Exception:
                                n = 0
                            if n > max_disp:
                                step = max(1, int(n // max_disp))
                                if step > 1:
                                    try:
                                        xs_use = xs_use[::step]
                                        ys_use = ys_use[::step]
                                    except Exception:
                                        # fallback: no decimation
                                        xs_use, ys_use = xs, ys

                        # Prefer skipping finite checks when supported
                        try:
                            curve.setData(xs_use, ys_use, connect='finite', skipFiniteCheck=True)
                        except Exception:
                            curve.setData(xs_use, ys_use, connect='finite')

                    if self.autoscale_chk.isChecked():
                        finite = np.isfinite(ys_use)
                        if finite.any():
                            y_min = float(np.nanmin(ys_use))
                            y_max = float(np.nanmax(ys_use))
                            global_y_min = y_min if global_y_min is None else min(global_y_min, y_min)
                            global_y_max = y_max if global_y_max is None else max(global_y_max, y_max)
            elif new_data and xs is not None:
                # Fallback (no numpy)
                for name in self.channel_names:
                    buf = self._val_buf_by_channel.get(name, [])
                    if not full:
                        ys_raw = list(buf[:count])
                    else:
                        ys_raw = list(buf[idx:]) + list(buf[:idx])

                    xs2, ys2 = [], []
                    for t, v in zip(xs, ys_raw):
                        if v is None:
                            continue
                        xs2.append(t)
                        ys2.append(v)

                    curve = self.curves.get(name)
                    if curve is not None:
                        curve.setData(xs2, ys2)

                    if self.autoscale_chk.isChecked() and ys2:
                        y_min, y_max = min(ys2), max(ys2)
                        global_y_min = y_min if global_y_min is None else min(global_y_min, y_min)
                        global_y_max = y_max if global_y_max is None else max(global_y_max, y_max)

            now = time.monotonic()
            # Smooth scrolling: keep a fixed visible window ending at "now".
            # Only do this while acquiring; after stop/pause we freeze and let
            # the user inspect/pan without being overridden by the timer.
            if scroll_live:
                try:
                    self.plot.setXRange(x_left, x_right, padding=0.0)
                except Exception:
                    pass

            # Y range update with hysteresis to reduce jitter/flicker
            if self.autoscale_chk.isChecked() and global_y_min is not None and global_y_max is not None:
                if (now - float(getattr(self, "_last_yrange_update", 0.0))) >= 1.0:
                    if global_y_min == global_y_max:
                        pad = 1.0 if global_y_min == 0 else abs(global_y_min) * 0.05
                        new_min = global_y_min - pad
                        new_max = global_y_max + pad
                    else:
                        span = global_y_max - global_y_min
                        pad = span * 0.08
                        new_min = global_y_min - pad
                        new_max = global_y_max + pad

                    apply = True
                    try:
                        cur_min, cur_max = self.plot.viewRange()[1]
                        cur_span = cur_max - cur_min
                        if cur_span > 0:
                            margin = cur_span * 0.05
                            # If new range is mostly inside current range, skip update.
                            if (new_min >= (cur_min + margin)) and (new_max <= (cur_max - margin)):
                                apply = False
                    except Exception:
                        pass

                    if apply:
                        self.plot.setYRange(new_min, new_max, padding=0.0)
                        self._last_yrange_update = now
        finally:
            self.plot.setUpdatesEnabled(True)
            # Request a single repaint after updating all curves.
            self.plot.update()

        try:
            self._update_friction_plots(xs, idx, full, count, scroll_live, x_left, x_right)
        except Exception:
            pass

        self._last_plotted_seq = int(getattr(self, "_plot_seq", 0) or 0)


    # ---------- connect/acquire ----------
    def toggle_connect(self):
        if self.is_connected:
            self.disconnect_serial()
        else:
            self.connect_serial()

    def connect_serial(self):
        port = self.port_combo.currentData()
        if not port:
            QMessageBox.warning(self, "提示", "未选择有效串口。请刷新串口后选择。")
            return

        try:
            baud = int(self.baud_combo.currentText())
        except Exception:
            baud = 9600

        parity = self.parity_combo.currentText()
        stopbits = int(self.stopbits_combo.currentText())
        bytesize = int(self.bytesize_combo.currentText())
        timeout_s = float(self.timeout_spin.value())

        tx_enabled = bool(self.enable_tx_chk.isChecked())
        tx_port = self.tx_port_combo.currentData() if tx_enabled else ""
        try:
            tx_baud = int(self.tx_baud_combo.currentText())
        except Exception:
            tx_baud = 115200

        tx_interval_ms = int(self.tx_interval_spin.value()) if hasattr(self, 'tx_interval_spin') else 20
        if tx_enabled and not tx_port:
            QMessageBox.warning(self, "提示", "已启用发送串口，但未选择发送串口。")
            self._set_serial_widgets_enabled(True)
            self.connect_btn.setEnabled(True)
            return

        rs485_cfg = Rs485CtrlConfig(
            mode=self.rs485_mode_combo.currentText(),
            tx_level_high=True,
            rx_level_high=False,
            pre_tx_ms=int(self.pre_tx_spin.value()),
            post_tx_ms=int(self.post_tx_spin.value()),
        )

        self.worker = ModbusRtuWorker(
            port=port, baudrate=baud, parity=parity, stopbits=stopbits,
            bytesize=bytesize, timeout_s=timeout_s, rs485_cfg=rs485_cfg,
            tx_enabled=tx_enabled, tx_port=tx_port, tx_baudrate=tx_baud, tx_interval_ms=tx_interval_ms
        )
        self.worker.data_ready.connect(self.on_data_ready)
        self.worker.status.connect(self.set_status)
        self.worker.connected.connect(self.on_connected_state)
        self.worker.acquiring.connect(self.on_acquiring_state)
        self.worker.log_line.connect(self.append_monitor)
        self.worker.frame.connect(self.on_frame)
        try:
            self.worker.set_tx_tap_enabled(bool(self.mon_tx_chk.isChecked()))
        except Exception:
            pass
        self.worker.start()

        self.set_status("正在连接...")
        self._set_serial_widgets_enabled(False)
        self.connect_btn.setEnabled(False)

    def disconnect_serial(self):
        if self.worker:
            try:
                self.worker.set_acquiring(False)
                self.worker.stop_thread()
                self.worker.wait(1500)
            except Exception:
                pass
            self.worker = None

        self.is_connected = False
        self.is_acquiring = False
        self.is_paused = False
        self.connect_btn.setText("连接串口")
        self.connect_btn.setEnabled(True)
        self.acquire_btn.setText("开始采集")
        self.acquire_btn.setEnabled(False)
        self.pause_btn.setText("暂停")
        self.pause_btn.setEnabled(False)
        self._set_serial_widgets_enabled(True)
        self.set_status("已断开连接")
        self.update_custom_send_ports()
        try:
            self._update_plot_timer_running()
        except Exception:
            pass

    @Slot(bool)
    def on_connected_state(self, ok: bool):
        self.is_connected = bool(ok)
        if ok:
            self.connect_btn.setText("断开串口")
            self.connect_btn.setEnabled(True)
            self.acquire_btn.setEnabled(True)
            self.pause_btn.setEnabled(False)
            self.pause_btn.setText("暂停")
            self.is_paused = False
            self.set_status("已连接（未采集）")
            self.update_custom_send_ports()
        else:
            self.disconnect_serial()
            return

        try:
            self._update_plot_timer_running()
        except Exception:
            pass

    def toggle_acquire(self):
        if not self.is_connected or not self.worker:
            return
        if self.is_acquiring:
            self.stop_acquire()
        else:
            self.start_acquire()

    def start_acquire(self):
        if not self.worker:
            return

        unit_id = int(self.unit_spin.value())
        func_code = int(self.func_combo.currentData())
        poll_ms = int(self.poll_spin.value())
        address_base_1 = bool(self.addr_base1_chk.isChecked())

        channels = self.gather_channels()
        enabled_channels = [c for c in channels if c.enabled]
        if not enabled_channels:
            QMessageBox.warning(self, "提示", "至少启用一个通道。")
            return

        self.worker.update_runtime(unit_id, func_code, poll_ms, address_base_1, enabled_channels, tx_interval_ms=int(self.tx_interval_spin.value()))

        # reset plot data at start
        self.clear_data()
        # reset pause-compensation for timeline
        self._mono_pause_accum = 0.0
        self._mono_pause_start = None
        self.channel_names = [c.name for c in enabled_channels]
        self.init_curves(self.channel_names)
        # allocate ring buffer with current max points
        try:
            size = int(self.max_points_spin.value())
        except Exception:
            size = int(getattr(self, '_buf_size', 100) or 100)
        self._alloc_ring_buffers(size, list(self.channel_names), keep_last=False)

        self.worker.set_acquiring(True)
        self.is_acquiring = True
        self.is_paused = False
        self.pause_btn.setText("暂停")
        self.pause_btn.setEnabled(True)
        self.acquire_btn.setText("停止采集")
        self.set_status("采集中...（请看通讯监视窗口 TX/RX）")
        try:
            self._update_plot_timer_running()
        except Exception:
            pass

    def stop_acquire(self):
        if self.worker:
            self.worker.set_acquiring(False)
        self.is_acquiring = False
        self.is_paused = False
        self._mono_pause_start = None
        self.pause_btn.setText("暂停")
        self.pause_btn.setEnabled(False)
        self.acquire_btn.setText("开始采集")
        self.set_status("已连接（未采集）")
        try:
            self._update_plot_timer_running()
        except Exception:
            pass

    def toggle_pause(self):
        """
        暂停/继续采集：
        - 暂停：停止发送查询（保持连接），不清空数据，不绘图追加
        - 继续：恢复发送查询，继续在原数据后追加
        """
        if not self.is_connected or not self.worker:
            return
        # 只有在“已开始采集（acquire_btn 处于停止采集状态）”时才允许暂停
        if self.acquire_btn.text() != "停止采集":
            return

        if not self.is_paused:
            # pause
            self.worker.set_acquiring(False)
            self.is_acquiring = False
            self.is_paused = True
            self._mono_pause_start = time.monotonic()
            self.pause_btn.setText("继续")
            self.set_status("已暂停（保持连接）")
        else:
            # resume
            try:
                if getattr(self, '_mono_pause_start', None) is not None:
                    self._mono_pause_accum = float(getattr(self, '_mono_pause_accum', 0.0) or 0.0) + max(0.0, float(time.monotonic() - float(self._mono_pause_start)))
            except Exception:
                pass
            self._mono_pause_start = None
            self.worker.set_acquiring(True)
            self.is_acquiring = True
            self.is_paused = False
            self.pause_btn.setText("暂停")
            self.set_status("采集中...（请看通讯监视窗口 TX/RX）")

        try:
            self._update_plot_timer_running()
        except Exception:
            pass


    @Slot(bool)
    def on_acquiring_state(self, on: bool):
        self.is_acquiring = bool(on)


    # ---------- export ----------
    def save_xlsx(self):
        xs, ys_map = self._snapshot_ring()
        if not xs or not self.channel_names:
            QMessageBox.information(self, "提示", "当前没有可保存的数据。请先开始采集。")
            return

        path, _ = QFileDialog.getSaveFileName(self, "保存为 XLSX", "modbus_data.xlsx", "Excel Files (*.xlsx)")
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        try:
            wb = Workbook()
            ws_all = wb.active
            ws_all.title = "All"

            ws_all.cell(row=1, column=1, value="Time")
            for i, name in enumerate(self.channel_names, start=2):
                ws_all.cell(row=1, column=i, value=name)

            nrows = len(xs)
            for r, ts in enumerate(xs, start=2):
                t_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(float(ts)))
                ws_all.cell(row=r, column=1, value=t_str)
                for c, name in enumerate(self.channel_names, start=2):
                    col = ys_map.get(name, [])
                    v = col[r - 2] if (r - 2) < len(col) else None
                    ws_all.cell(row=r, column=c, value=v)

            for name in self.channel_names:
                ws = wb.create_sheet(title=self._safe_sheet_name(name))
                ws.cell(row=1, column=1, value="Time")
                ws.cell(row=1, column=2, value="Value")
                vals = ys_map.get(name, [])
                for r in range(nrows):
                    t_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(float(xs[r])))
                    ws.cell(row=r + 2, column=1, value=t_str)
                    ws.cell(row=r + 2, column=2, value=(vals[r] if r < len(vals) else None))

            for ws in wb.worksheets:
                self._autosize_sheet(ws)

            wb.save(path)
            self.set_status(f"已保存：{path}")
        except Exception as e:
            QMessageBox.critical(self, "保存失败", f"保存 xlsx 失败：\n{e}")

    @staticmethod
    def _safe_sheet_name(name: str) -> str:
        bad = ['\\', '/', '*', '[', ']', ':', '?']
        s = "".join("_" if ch in bad else ch for ch in name).strip() or "Sheet"
        return s[:31]

    @staticmethod
    def _autosize_sheet(ws):
        max_rows = min(ws.max_row, 200)
        for col in range(1, ws.max_column + 1):
            max_len = 0
            for row in range(1, max_rows + 1):
                v = ws.cell(row=row, column=col).value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[get_column_letter(col)].width = min(max(10, max_len + 2), 40)

    def closeEvent(self, event):
        # Persist workspace layout (dock positions / splitter sizes / window geometry)
        try:
            self._save_window_layout()
        except Exception:
            pass
        # Ensure serial threads are stopped before exit
        try:
            self.disconnect_serial()
        except Exception:
            pass
        super().closeEvent(event)


def main():
    # Light theme
    pg.setConfigOption("background", "w")
    pg.setConfigOption("foreground", "k")
    # Speed-focused config (Windows: big win for real-time plotting)
    # - antialias=False: avoids costly QPainter AA
    # - useOpenGL=True: lets pyqtgraph render curves via OpenGL when available
    try:
        pg.setConfigOptions(antialias=False, useOpenGL=True)
    except Exception:
        try:
            pg.setConfigOptions(antialias=False)
            pg.setConfigOption("useOpenGL", True)
        except Exception:
            pass

    app = QApplication(sys.argv)
    w = MainWindow()
    w.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
